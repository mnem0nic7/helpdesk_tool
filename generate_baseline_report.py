import json
import os
from datetime import datetime, timezone, timedelta
from collections import Counter, defaultdict
import statistics

# ── Load data ──
print("Loading ticket data...")
with open("/home/gallison/workspace/altlassian/oit_all_issues.json", "r") as f:
    issues = json.load(f)

print(f"Loaded {len(issues)} tickets")

REPORT_DIR = "/home/gallison/workspace/altlassian/reports/2026-03-03_0825"
REPORT_TS = "2026-03-03 08:25 UTC"

# ── Status mapping per baseline ──
ACTIVE_STATUSES = {"new", "open", "assigned", "in progress", "work in progress", "investigating"}
PAUSED_STATUSES = {"waiting for customer", "waiting for support", "pending", "pending customer",
                   "pending vendor", "scheduled", "on hold", "awaiting approval",
                   "waiting for approval"}
TERMINAL_STATUSES = {"resolved", "closed", "done", "cancelled", "declined", "canceled"}

def map_status_bucket(status_name):
    s = (status_name or "").strip().lower()
    if s in TERMINAL_STATUSES:
        return "Terminal"
    if s in PAUSED_STATUSES:
        return "Paused"
    if s in ACTIVE_STATUSES:
        return "Active"
    # Fuzzy match
    for p in PAUSED_STATUSES:
        if p in s or s in p:
            return "Paused"
    for a in ACTIVE_STATUSES:
        if a in s or s in a:
            return "Active"
    return "Active"  # default: count toward clock

def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except:
        return None

def hours_between(a, b):
    if a and b:
        delta = (b - a).total_seconds() / 3600.0
        return max(delta, 0)
    return None

def percentile(data, p):
    if not data:
        return None
    s = sorted(data)
    k = (len(s) - 1) * p / 100.0
    f = int(k)
    c = f + 1
    if c >= len(s):
        return s[f]
    return s[f] + (k - f) * (s[c] - s[f])

# ── Parse all tickets ──
print("Parsing tickets...")
now = datetime.now(timezone.utc)
tickets = []
for iss in issues:
    fields = iss.get("fields", {})
    key = iss.get("key", "")
    
    created = parse_dt(fields.get("created"))
    resolved = parse_dt(fields.get("resolutiondate"))
    updated = parse_dt(fields.get("updated"))
    
    status_name = ""
    status_cat = ""
    if fields.get("status"):
        status_name = fields["status"].get("name", "")
        sc = fields["status"].get("statusCategory", {})
        status_cat = sc.get("name", "") if sc else ""
    
    priority_name = ""
    if fields.get("priority"):
        priority_name = fields["priority"].get("name", "")
    
    issue_type = ""
    if fields.get("issuetype"):
        issue_type = fields["issuetype"].get("name", "")
    
    assignee_name = ""
    if fields.get("assignee"):
        assignee_name = fields["assignee"].get("displayName", "")
    
    reporter_name = ""
    if fields.get("reporter"):
        reporter_name = fields["reporter"].get("displayName", "")
    
    summary = fields.get("summary", "")
    labels = fields.get("labels", [])
    components = [c.get("name", "") for c in (fields.get("components") or [])]
    
    # Request type (JSM)
    request_type = ""
    crf = fields.get("customfield_10010")
    if crf and isinstance(crf, dict):
        request_type = crf.get("requestType", {}).get("name", "") if isinstance(crf.get("requestType"), dict) else ""
    
    status_bucket = map_status_bucket(status_name)
    is_open = status_bucket != "Terminal"
    
    calendar_ttr_h = hours_between(created, resolved)
    
    age_h = hours_between(created, now) if created else None
    staleness_h = hours_between(updated, now) if updated else None
    
    # Month bucket
    month_str = created.strftime("%Y-%m") if created else "Unknown"
    
    tickets.append({
        "key": key,
        "summary": summary,
        "issue_type": issue_type,
        "request_type": request_type,
        "status": status_name,
        "status_bucket": status_bucket,
        "status_cat": status_cat,
        "priority": priority_name,
        "assignee": assignee_name,
        "reporter": reporter_name,
        "created": created,
        "resolved": resolved,
        "updated": updated,
        "is_open": is_open,
        "calendar_ttr_h": calendar_ttr_h,
        "age_h": age_h,
        "staleness_h": staleness_h,
        "month": month_str,
        "labels": labels,
        "components": components,
    })

print(f"Parsed {len(tickets)} tickets")

# ══════════════════════════════════════════════
# SECTION A: TICKET TYPE CLASSIFICATION
# ══════════════════════════════════════════════
type_counts = Counter(t["issue_type"] for t in tickets)
request_type_counts = Counter(t["request_type"] for t in tickets if t["request_type"])

# ══════════════════════════════════════════════
# SECTION B: CORE KPIs
# ══════════════════════════════════════════════

# B1: Demand & Intake
total = len(tickets)
open_tickets = [t for t in tickets if t["is_open"]]
resolved_tickets = [t for t in tickets if not t["is_open"]]

monthly_created = Counter(t["month"] for t in tickets)
monthly_resolved = Counter(t["month"] for t in resolved_tickets)

# Ticket mix by type
type_mix = {k: v/total*100 for k, v in type_counts.items()}

# By priority
priority_counts = Counter(t["priority"] for t in tickets)
priority_open = Counter(t["priority"] for t in open_tickets)

# B2: Backlog Health
age_buckets = {"0-2d": 0, "3-7d": 0, "8-14d": 0, "15-30d": 0, "30+d": 0}
for t in open_tickets:
    if t["age_h"] is not None:
        days = t["age_h"] / 24
        if days <= 2: age_buckets["0-2d"] += 1
        elif days <= 7: age_buckets["3-7d"] += 1
        elif days <= 14: age_buckets["8-14d"] += 1
        elif days <= 30: age_buckets["15-30d"] += 1
        else: age_buckets["30+d"] += 1

# Stale tickets (no update in 5+ business days ~ 7 calendar days)
stale = [t for t in open_tickets if t["staleness_h"] and t["staleness_h"] > 7*24]
stale_by_assignee = Counter(t["assignee"] or "(Unassigned)" for t in stale)
stale_by_status = Counter(t["status"] for t in stale)

# Net flow by month
all_months = sorted(set(list(monthly_created.keys()) + list(monthly_resolved.keys())))
net_flow = {m: monthly_created.get(m, 0) - monthly_resolved.get(m, 0) for m in all_months}

# B3: Responsiveness — FRT not available without changelog, note it
# B4: Resolution speed
resolved_with_ttr = [t for t in resolved_tickets if t["calendar_ttr_h"] is not None]
ttr_values = [t["calendar_ttr_h"] for t in resolved_with_ttr]

ttr_stats = {}
if ttr_values:
    ttr_stats["count"] = len(ttr_values)
    ttr_stats["mean"] = statistics.mean(ttr_values)
    ttr_stats["median"] = statistics.median(ttr_values)
    ttr_stats["p90"] = percentile(ttr_values, 90)
    ttr_stats["p95"] = percentile(ttr_values, 95)
    ttr_stats["min"] = min(ttr_values)
    ttr_stats["max"] = max(ttr_values)

# TTR by type
ttr_by_type = defaultdict(list)
for t in resolved_with_ttr:
    ttr_by_type[t["issue_type"]].append(t["calendar_ttr_h"])

# TTR by priority
ttr_by_priority = defaultdict(list)
for t in resolved_with_ttr:
    ttr_by_priority[t["priority"]].append(t["calendar_ttr_h"])

# TTR by month
ttr_by_month = defaultdict(list)
for t in resolved_with_ttr:
    m = t["resolved"].strftime("%Y-%m") if t["resolved"] else t["month"]
    ttr_by_month[m].append(t["calendar_ttr_h"])

# B5: SLA — no formal SLA targets defined yet, track distribution
# B6: Quality
# Reopen detection: approximate by looking for tickets resolved then back to open
# (Limited with snapshot data - note this)

# Reassignment: not available without changelog

# B7: Customer Experience — CSAT not in data, note it

# B8: Escalation — need changelog for accurate tracking

# ══════════════════════════════════════════════
# SECTION C: EXTENDED — Productivity
# ══════════════════════════════════════════════
assignee_resolved = Counter(t["assignee"] or "(Unassigned)" for t in resolved_tickets)
assignee_open = Counter(t["assignee"] or "(Unassigned)" for t in open_tickets)

# Top 20 assignees by resolved
top_assignees = assignee_resolved.most_common(20)

# Open backlog by assignee (WIP)
open_by_assignee_status = defaultdict(lambda: Counter())
for t in open_tickets:
    a = t["assignee"] or "(Unassigned)"
    open_by_assignee_status[a][t["status"]] += 1

# ══════════════════════════════════════════════
# GENERATE EXCEL
# ══════════════════════════════════════════════
print("Generating Excel report...")
import sys
sys.path.insert(0, "/home/gallison/workspace/altlassian/.venv/lib/python3.12/site-packages")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ──
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor="2F5496")
section_font = Font(bold=True, size=13, color="1F3864")
subsection_font = Font(bold=True, size=11, color="2F5496")
metric_font = Font(bold=True, size=11)
warn_fill = PatternFill("solid", fgColor="FFC7CE")
good_fill = PatternFill("solid", fgColor="C6EFCE")
neutral_fill = PatternFill("solid", fgColor="D9E2F3")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx

def fmt_hours(h):
    if h is None: return "N/A"
    if h < 1: return f"{h*60:.0f}m"
    if h < 24: return f"{h:.1f}h"
    return f"{h/24:.1f}d"

# ═══════════════════════════════════════════
# Sheet 1: Executive Summary
# ═══════════════════════════════════════════
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = "2F5496"

r = 1
ws.cell(r, 1, "OIT HELPDESK BASELINE METRICS REPORT").font = Font(bold=True, size=16, color="1F3864")
r = 2
ws.cell(r, 1, f"Report generated: {REPORT_TS}").font = Font(italic=True, size=10, color="666666")
ws.cell(r+1, 1, f"Data range: {all_months[0]} to {all_months[-1]} ({len(all_months)} months)").font = Font(italic=True, size=10, color="666666")
ws.cell(r+2, 1, "Baseline framework: Helpdesk Metrics Baseline v1.0").font = Font(italic=True, size=10, color="666666")

r = 6
ws.cell(r, 1, "A. HEADLINE NUMBERS").font = section_font
r += 1
headlines = [
    ("Total Tickets", f"{total:,}"),
    ("Open (Backlog)", f"{len(open_tickets):,}"),
    ("Resolved/Closed", f"{len(resolved_tickets):,}"),
    ("Resolution Rate", f"{len(resolved_tickets)/total*100:.1f}%"),
    ("", ""),
    ("Calendar TTR — Median", fmt_hours(ttr_stats.get("median"))),
    ("Calendar TTR — Mean", fmt_hours(ttr_stats.get("mean"))),
    ("Calendar TTR — P90", fmt_hours(ttr_stats.get("p90"))),
    ("Calendar TTR — P95", fmt_hours(ttr_stats.get("p95"))),
    ("", ""),
    ("Open Backlog Aging — 0-2 days", str(age_buckets["0-2d"])),
    ("Open Backlog Aging — 3-7 days", str(age_buckets["3-7d"])),
    ("Open Backlog Aging — 8-14 days", str(age_buckets["8-14d"])),
    ("Open Backlog Aging — 15-30 days", str(age_buckets["15-30d"])),
    ("Open Backlog Aging — 30+ days", str(age_buckets["30+d"])),
    ("", ""),
    ("Stale Tickets (no update 7+ days)", str(len(stale))),
]
for label, val in headlines:
    if not label:
        r += 1
        continue
    ws.cell(r, 1, label).font = metric_font
    ws.cell(r, 2, val)
    ws.cell(r, 2).alignment = Alignment(horizontal="right")
    r += 1

r += 2
ws.cell(r, 1, "B. DATA AVAILABILITY NOTES").font = section_font
r += 1
notes = [
    "Calendar TTR: Available (Created → Resolved timestamps). Reported as median + P90/P95 per baseline.",
    "Active TTR: NOT available — requires status changelog to calculate time in Paused states.",
    "First Response Time (FRT): NOT available — requires changelog for first agent comment timestamp.",
    "Time to Assign: NOT available — requires changelog for assignment events.",
    "Reopen Rate: NOT available from snapshot — requires changelog or status transition history.",
    "Reassignment Rate: NOT available — requires changelog for assignment changes.",
    "CSAT: NOT present in ticket data — requires separate survey integration.",
    "SLA Compliance: No formal SLA targets defined yet — tracking TTR distribution as proxy.",
    "Escalation Metrics: NOT available — requires changelog for tier handoff events.",
    "",
    "RECOMMENDATION: Enable Jira changelog export to unlock Active TTR, FRT, Reopen, Reassignment, and Escalation KPIs.",
]
for note in notes:
    ws.cell(r, 1, note).font = Font(size=9, color="444444")
    r += 1

ws.column_dimensions["A"].width = 55
ws.column_dimensions["B"].width = 20

# ═══════════════════════════════════════════
# Sheet 2: Demand & Intake
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Demand & Intake")
ws2.sheet_properties.tabColor = "00B050"

r = 1
ws2.cell(r, 1, "DEMAND & INTAKE METRICS").font = section_font
r += 2

# Ticket type mix
ws2.cell(r, 1, "Ticket Type Mix").font = subsection_font
r += 1
ws2.cell(r, 1, "Issue Type"); ws2.cell(r, 2, "Count"); ws2.cell(r, 3, "% of Total")
style_header_row(ws2, r, 3)
r += 1
for typ, cnt in type_counts.most_common():
    ws2.cell(r, 1, typ)
    ws2.cell(r, 2, cnt)
    ws2.cell(r, 3, f"{cnt/total*100:.1f}%")
    r += 1

# Priority distribution
r += 1
ws2.cell(r, 1, "Priority Distribution").font = subsection_font
r += 1
ws2.cell(r, 1, "Priority"); ws2.cell(r, 2, "Total"); ws2.cell(r, 3, "Open"); ws2.cell(r, 4, "% of Total")
style_header_row(ws2, r, 4)
r += 1
prio_order = ["Highest", "High", "Medium", "Low", "Lowest"]
for p in prio_order:
    if p in priority_counts:
        ws2.cell(r, 1, p)
        ws2.cell(r, 2, priority_counts[p])
        ws2.cell(r, 3, priority_open.get(p, 0))
        ws2.cell(r, 4, f"{priority_counts[p]/total*100:.1f}%")
        r += 1
# Any not in standard order
for p, cnt in priority_counts.most_common():
    if p not in prio_order:
        ws2.cell(r, 1, p or "(None)")
        ws2.cell(r, 2, cnt)
        ws2.cell(r, 3, priority_open.get(p, 0))
        ws2.cell(r, 4, f"{cnt/total*100:.1f}%")
        r += 1

# JSM Request Types (top 25)
if request_type_counts:
    r += 1
    ws2.cell(r, 1, "Top JSM Request Types").font = subsection_font
    r += 1
    ws2.cell(r, 1, "Request Type"); ws2.cell(r, 2, "Count"); ws2.cell(r, 3, "% of Total")
    style_header_row(ws2, r, 3)
    r += 1
    for rt, cnt in request_type_counts.most_common(25):
        ws2.cell(r, 1, rt)
        ws2.cell(r, 2, cnt)
        ws2.cell(r, 3, f"{cnt/total*100:.1f}%")
        r += 1

# Monthly volume
r += 1
ws2.cell(r, 1, "Monthly Ticket Volume").font = subsection_font
r += 1
hdr_r = r
ws2.cell(r, 1, "Month"); ws2.cell(r, 2, "Created"); ws2.cell(r, 3, "Resolved"); ws2.cell(r, 4, "Net Flow"); ws2.cell(r, 5, "Cumulative Backlog Delta")
style_header_row(ws2, r, 5)
r += 1
cum_delta = 0
start_monthly = r
for m in all_months:
    c = monthly_created.get(m, 0)
    res = monthly_resolved.get(m, 0)
    nf = c - res
    cum_delta += nf
    ws2.cell(r, 1, m)
    ws2.cell(r, 2, c)
    ws2.cell(r, 3, res)
    ws2.cell(r, 4, nf)
    ws2.cell(r, 5, cum_delta)
    if nf > 0:
        ws2.cell(r, 4).fill = warn_fill
    elif nf < 0:
        ws2.cell(r, 4).fill = good_fill
    r += 1

# Monthly volume chart
chart = LineChart()
chart.title = "Monthly Ticket Volume"
chart.y_axis.title = "Tickets"
chart.x_axis.title = "Month"
chart.width = 28
chart.height = 14
data = Reference(ws2, min_col=2, max_col=3, min_row=hdr_r, max_row=r-1)
cats = Reference(ws2, min_col=1, min_row=hdr_r+1, max_row=r-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.line.width = 25000
chart.series[1].graphicalProperties.line.width = 25000
ws2.add_chart(chart, f"G2")

auto_width(ws2)

# ═══════════════════════════════════════════
# Sheet 3: Backlog Health
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Backlog Health")
ws3.sheet_properties.tabColor = "FFC000"

r = 1
ws3.cell(r, 1, "BACKLOG HEALTH").font = section_font
r += 2

# Open backlog summary
ws3.cell(r, 1, "Open Backlog Summary").font = subsection_font
r += 1
ws3.cell(r, 1, "Total Open"); ws3.cell(r, 2, len(open_tickets))
r += 1

# Aging buckets
ws3.cell(r, 1, "Aging Distribution").font = subsection_font
r += 1
ws3.cell(r, 1, "Age Bucket"); ws3.cell(r, 2, "Count"); ws3.cell(r, 3, "% of Open")
style_header_row(ws3, r, 3)
r += 1
age_start = r
for bucket, cnt in age_buckets.items():
    ws3.cell(r, 1, bucket)
    ws3.cell(r, 2, cnt)
    ws3.cell(r, 3, f"{cnt/max(len(open_tickets),1)*100:.1f}%")
    if bucket == "30+d" and cnt > 0:
        ws3.cell(r, 2).fill = warn_fill
    r += 1

# Aging pie chart
pie = PieChart()
pie.title = "Open Backlog by Age"
pie.width = 16
pie.height = 12
data = Reference(ws3, min_col=2, min_row=age_start-1, max_row=r-1)
cats = Reference(ws3, min_col=1, min_row=age_start, max_row=r-1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws3.add_chart(pie, "E3")

# Open by status
r += 1
ws3.cell(r, 1, "Open Tickets by Status").font = subsection_font
r += 1
ws3.cell(r, 1, "Status"); ws3.cell(r, 2, "Count"); ws3.cell(r, 3, "Status Bucket")
style_header_row(ws3, r, 3)
r += 1
open_status = Counter(t["status"] for t in open_tickets)
for s, cnt in open_status.most_common():
    ws3.cell(r, 1, s)
    ws3.cell(r, 2, cnt)
    ws3.cell(r, 3, map_status_bucket(s))
    r += 1

# Open by priority
r += 1
ws3.cell(r, 1, "Open Tickets by Priority").font = subsection_font
r += 1
ws3.cell(r, 1, "Priority"); ws3.cell(r, 2, "Count")
style_header_row(ws3, r, 2)
r += 1
for p in prio_order:
    if priority_open.get(p, 0) > 0:
        ws3.cell(r, 1, p)
        ws3.cell(r, 2, priority_open[p])
        r += 1
for p, cnt in priority_open.most_common():
    if p not in prio_order and cnt > 0:
        ws3.cell(r, 1, p or "(None)")
        ws3.cell(r, 2, cnt)
        r += 1

# Stale tickets
r += 1
ws3.cell(r, 1, f"Stale Tickets (no update in 7+ days): {len(stale)}").font = subsection_font
r += 1
if stale:
    ws3.cell(r, 1, "Stale by Assignee").font = subsection_font
    r += 1
    ws3.cell(r, 1, "Assignee"); ws3.cell(r, 2, "Stale Count")
    style_header_row(ws3, r, 2)
    r += 1
    for a, cnt in stale_by_assignee.most_common():
        ws3.cell(r, 1, a)
        ws3.cell(r, 2, cnt)
        r += 1
    
    r += 1
    ws3.cell(r, 1, "Stale by Status").font = subsection_font
    r += 1
    ws3.cell(r, 1, "Status"); ws3.cell(r, 2, "Stale Count")
    style_header_row(ws3, r, 2)
    r += 1
    for s, cnt in stale_by_status.most_common():
        ws3.cell(r, 1, s)
        ws3.cell(r, 2, cnt)
        r += 1

# Stale ticket detail
r += 1
ws3.cell(r, 1, "Stale Ticket Detail").font = subsection_font
r += 1
ws3.cell(r, 1, "Key"); ws3.cell(r, 2, "Summary"); ws3.cell(r, 3, "Status"); ws3.cell(r, 4, "Assignee"); ws3.cell(r, 5, "Priority"); ws3.cell(r, 6, "Days Since Update"); ws3.cell(r, 7, "Age (days)")
style_header_row(ws3, r, 7)
r += 1
for t in sorted(stale, key=lambda x: x["staleness_h"] or 0, reverse=True):
    ws3.cell(r, 1, t["key"])
    ws3.cell(r, 2, t["summary"][:80])
    ws3.cell(r, 3, t["status"])
    ws3.cell(r, 4, t["assignee"] or "(Unassigned)")
    ws3.cell(r, 5, t["priority"])
    ws3.cell(r, 6, f"{t['staleness_h']/24:.0f}" if t["staleness_h"] else "")
    ws3.cell(r, 7, f"{t['age_h']/24:.0f}" if t["age_h"] else "")
    r += 1

auto_width(ws3)

# ═══════════════════════════════════════════
# Sheet 4: Resolution Speed (Calendar TTR)
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Resolution Speed")
ws4.sheet_properties.tabColor = "FF0000"

r = 1
ws4.cell(r, 1, "RESOLUTION SPEED — Calendar TTR").font = section_font
ws4.cell(r+1, 1, "Note: This is Calendar TTR (Created → Resolved). Active TTR requires status changelog.").font = Font(italic=True, size=9, color="666666")
r = 4

# Overall stats
ws4.cell(r, 1, "Overall Calendar TTR").font = subsection_font
r += 1
ws4.cell(r, 1, "Statistic"); ws4.cell(r, 2, "Value (hours)"); ws4.cell(r, 3, "Human-readable")
style_header_row(ws4, r, 3)
r += 1
for label, key in [("Count", "count"), ("Mean", "mean"), ("Median", "median"), ("P90", "p90"), ("P95", "p95"), ("Min", "min"), ("Max", "max")]:
    ws4.cell(r, 1, label)
    v = ttr_stats.get(key)
    if key == "count":
        ws4.cell(r, 2, v)
        ws4.cell(r, 3, f"{v:,}")
    else:
        ws4.cell(r, 2, round(v, 2) if v else "N/A")
        ws4.cell(r, 3, fmt_hours(v))
    r += 1

# TTR by issue type
r += 1
ws4.cell(r, 1, "Calendar TTR by Issue Type").font = subsection_font
r += 1
ws4.cell(r, 1, "Issue Type"); ws4.cell(r, 2, "Count"); ws4.cell(r, 3, "Median (h)"); ws4.cell(r, 4, "Mean (h)"); ws4.cell(r, 5, "P90 (h)"); ws4.cell(r, 6, "P95 (h)")
style_header_row(ws4, r, 6)
r += 1
for typ in sorted(ttr_by_type.keys()):
    vals = ttr_by_type[typ]
    ws4.cell(r, 1, typ)
    ws4.cell(r, 2, len(vals))
    ws4.cell(r, 3, round(statistics.median(vals), 2))
    ws4.cell(r, 4, round(statistics.mean(vals), 2))
    ws4.cell(r, 5, round(percentile(vals, 90), 2))
    ws4.cell(r, 6, round(percentile(vals, 95), 2))
    r += 1

# TTR by priority
r += 1
ws4.cell(r, 1, "Calendar TTR by Priority").font = subsection_font
r += 1
ws4.cell(r, 1, "Priority"); ws4.cell(r, 2, "Count"); ws4.cell(r, 3, "Median (h)"); ws4.cell(r, 4, "Mean (h)"); ws4.cell(r, 5, "P90 (h)"); ws4.cell(r, 6, "P95 (h)")
style_header_row(ws4, r, 6)
r += 1
for p in prio_order + [p for p in ttr_by_priority if p not in prio_order]:
    if p in ttr_by_priority:
        vals = ttr_by_priority[p]
        ws4.cell(r, 1, p or "(None)")
        ws4.cell(r, 2, len(vals))
        ws4.cell(r, 3, round(statistics.median(vals), 2))
        ws4.cell(r, 4, round(statistics.mean(vals), 2))
        ws4.cell(r, 5, round(percentile(vals, 90), 2))
        ws4.cell(r, 6, round(percentile(vals, 95), 2))
        r += 1

# Monthly TTR trend
r += 1
ws4.cell(r, 1, "Monthly Calendar TTR Trend").font = subsection_font
r += 1
hdr_r4 = r
ws4.cell(r, 1, "Month"); ws4.cell(r, 2, "Resolved"); ws4.cell(r, 3, "Median TTR (h)"); ws4.cell(r, 4, "P90 TTR (h)"); ws4.cell(r, 5, "P95 TTR (h)")
style_header_row(ws4, r, 5)
r += 1
start_r4 = r
for m in all_months:
    if m in ttr_by_month and ttr_by_month[m]:
        vals = ttr_by_month[m]
        ws4.cell(r, 1, m)
        ws4.cell(r, 2, len(vals))
        ws4.cell(r, 3, round(statistics.median(vals), 2))
        ws4.cell(r, 4, round(percentile(vals, 90), 2))
        ws4.cell(r, 5, round(percentile(vals, 95), 2))
        r += 1

# TTR trend chart
chart4 = LineChart()
chart4.title = "Monthly Calendar TTR (Median & P90)"
chart4.y_axis.title = "Hours"
chart4.width = 28
chart4.height = 14
data = Reference(ws4, min_col=3, max_col=4, min_row=hdr_r4, max_row=r-1)
cats = Reference(ws4, min_col=1, min_row=hdr_r4+1, max_row=r-1)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
ws4.add_chart(chart4, "G4")

# TTR distribution buckets
r += 2
ws4.cell(r, 1, "Resolution Time Distribution").font = subsection_font
r += 1
ttr_dist = {"< 1h": 0, "1-4h": 0, "4-8h": 0, "8-24h": 0, "1-3d": 0, "3-7d": 0, "7-14d": 0, "14-30d": 0, "30+d": 0}
for h in ttr_values:
    if h < 1: ttr_dist["< 1h"] += 1
    elif h < 4: ttr_dist["1-4h"] += 1
    elif h < 8: ttr_dist["4-8h"] += 1
    elif h < 24: ttr_dist["8-24h"] += 1
    elif h < 72: ttr_dist["1-3d"] += 1
    elif h < 168: ttr_dist["3-7d"] += 1
    elif h < 336: ttr_dist["7-14d"] += 1
    elif h < 720: ttr_dist["14-30d"] += 1
    else: ttr_dist["30+d"] += 1

ws4.cell(r, 1, "Time Bucket"); ws4.cell(r, 2, "Count"); ws4.cell(r, 3, "% of Resolved"); ws4.cell(r, 4, "Cumulative %")
style_header_row(ws4, r, 4)
r += 1
cum = 0
for bucket, cnt in ttr_dist.items():
    cum += cnt
    ws4.cell(r, 1, bucket)
    ws4.cell(r, 2, cnt)
    ws4.cell(r, 3, f"{cnt/max(len(ttr_values),1)*100:.1f}%")
    ws4.cell(r, 4, f"{cum/max(len(ttr_values),1)*100:.1f}%")
    r += 1

auto_width(ws4)

# ═══════════════════════════════════════════
# Sheet 5: Assignee Workload
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("Assignee Workload")
ws5.sheet_properties.tabColor = "7030A0"

r = 1
ws5.cell(r, 1, "ASSIGNEE WORKLOAD & PRODUCTIVITY").font = section_font
r += 2

ws5.cell(r, 1, "Assignee"); ws5.cell(r, 2, "Total Resolved"); ws5.cell(r, 3, "Currently Open")
ws5.cell(r, 4, "Median TTR (h)"); ws5.cell(r, 5, "P90 TTR (h)"); ws5.cell(r, 6, "Stale Open")
style_header_row(ws5, r, 6)
r += 1

# Build assignee-level TTR
assignee_ttr = defaultdict(list)
for t in resolved_with_ttr:
    assignee_ttr[t["assignee"] or "(Unassigned)"].append(t["calendar_ttr_h"])

all_assignees = set(list(assignee_resolved.keys()) + list(assignee_open.keys()))
assignee_data = []
for a in all_assignees:
    assignee_data.append((
        a,
        assignee_resolved.get(a, 0),
        assignee_open.get(a, 0),
        statistics.median(assignee_ttr[a]) if assignee_ttr.get(a) else None,
        percentile(assignee_ttr[a], 90) if assignee_ttr.get(a) else None,
        stale_by_assignee.get(a, 0)
    ))
assignee_data.sort(key=lambda x: x[1], reverse=True)

for a, res, opn, med, p90, stl in assignee_data[:30]:
    ws5.cell(r, 1, a)
    ws5.cell(r, 2, res)
    ws5.cell(r, 3, opn)
    ws5.cell(r, 4, round(med, 2) if med else "N/A")
    ws5.cell(r, 5, round(p90, 2) if p90 else "N/A")
    ws5.cell(r, 6, stl)
    if stl > 0:
        ws5.cell(r, 6).fill = warn_fill
    r += 1

auto_width(ws5)

# ═══════════════════════════════════════════
# Sheet 6: All Tickets (raw data)
# ═══════════════════════════════════════════
ws6 = wb.create_sheet("All Tickets")
ws6.sheet_properties.tabColor = "808080"

headers = ["Key", "Summary", "Issue Type", "Request Type", "Status", "Status Bucket",
           "Priority", "Assignee", "Reporter", "Created", "Resolved", "Updated",
           "Calendar TTR (h)", "Age (days)", "Days Since Update"]
for c, h in enumerate(headers, 1):
    ws6.cell(1, c, h)
style_header_row(ws6, 1, len(headers))

for i, t in enumerate(tickets, 2):
    ws6.cell(i, 1, t["key"])
    ws6.cell(i, 2, t["summary"][:120])
    ws6.cell(i, 3, t["issue_type"])
    ws6.cell(i, 4, t["request_type"])
    ws6.cell(i, 5, t["status"])
    ws6.cell(i, 6, t["status_bucket"])
    ws6.cell(i, 7, t["priority"])
    ws6.cell(i, 8, t["assignee"])
    ws6.cell(i, 9, t["reporter"])
    ws6.cell(i, 10, t["created"].strftime("%Y-%m-%d %H:%M") if t["created"] else "")
    ws6.cell(i, 11, t["resolved"].strftime("%Y-%m-%d %H:%M") if t["resolved"] else "")
    ws6.cell(i, 12, t["updated"].strftime("%Y-%m-%d %H:%M") if t["updated"] else "")
    ws6.cell(i, 13, round(t["calendar_ttr_h"], 2) if t["calendar_ttr_h"] else "")
    ws6.cell(i, 14, round(t["age_h"]/24, 1) if t["age_h"] else "")
    ws6.cell(i, 15, round(t["staleness_h"]/24, 1) if t["staleness_h"] else "")

ws6.auto_filter.ref = f"A1:O{len(tickets)+1}"

# ═══════════════════════════════════════════
# Sheet 7: Status Mapping
# ═══════════════════════════════════════════
ws7 = wb.create_sheet("Status Mapping")
ws7.sheet_properties.tabColor = "00B0F0"

r = 1
ws7.cell(r, 1, "STATUS MAPPING — Baseline Framework").font = section_font
r += 2
ws7.cell(r, 1, "Jira Status"); ws7.cell(r, 2, "Baseline Bucket"); ws7.cell(r, 3, "Clock"); ws7.cell(r, 4, "Count (All)"); ws7.cell(r, 5, "Count (Open)")
style_header_row(ws7, r, 5)
r += 1

all_statuses = Counter(t["status"] for t in tickets)
open_statuses = Counter(t["status"] for t in open_tickets)
for s, cnt in all_statuses.most_common():
    bucket = map_status_bucket(s)
    ws7.cell(r, 1, s)
    ws7.cell(r, 2, bucket)
    ws7.cell(r, 3, "Running" if bucket == "Active" else ("Paused" if bucket == "Paused" else "Stopped"))
    ws7.cell(r, 4, cnt)
    ws7.cell(r, 5, open_statuses.get(s, 0))
    if bucket == "Active":
        ws7.cell(r, 3).fill = good_fill
    elif bucket == "Paused":
        ws7.cell(r, 3).fill = PatternFill("solid", fgColor="FFEB9C")
    r += 1

auto_width(ws7)

# ═══════════════════════════════════════════
# Sheet 8: Metric Dictionary
# ═══════════════════════════════════════════
ws8 = wb.create_sheet("Metric Dictionary")
ws8.sheet_properties.tabColor = "1F3864"

r = 1
ws8.cell(r, 1, "METRIC DICTIONARY — Baseline Definitions").font = section_font
r += 2

dict_headers = ["Metric Name", "Category", "Formula / Definition", "Start", "Stop", "Clock Pauses", "Statistic", "Available?"]
for c, h in enumerate(dict_headers, 1):
    ws8.cell(r, c, h)
style_header_row(ws8, r, len(dict_headers))
r += 1

metrics_dict = [
    ("New Tickets", "Demand", "Count created in period", "N/A", "N/A", "N/A", "Count", "YES"),
    ("Ticket Mix", "Demand", "% by type (Incident/Request/Task)", "N/A", "N/A", "N/A", "Rate", "YES"),
    ("Open Backlog", "Backlog", "Count not Resolved/Closed", "N/A", "N/A", "N/A", "Count", "YES"),
    ("Backlog Aging", "Backlog", "Open grouped by age (0-2d, 3-7d, etc.)", "Created", "Now", "N/A", "Count by bucket", "YES"),
    ("Stale Tickets", "Backlog", "Open, no update in 7+ days", "Last update", "Now", "N/A", "Count", "YES"),
    ("Throughput", "Backlog", "Tickets Resolved in period", "N/A", "N/A", "N/A", "Count", "YES"),
    ("Net Flow", "Backlog", "New - Resolved per period", "N/A", "N/A", "N/A", "Count", "YES"),
    ("First Response Time", "Responsiveness", "Created → first agent response", "Created", "First response", "N/A", "Median + P90", "NO — needs changelog"),
    ("Time to Assign", "Responsiveness", "Created → first assignment", "Created", "Assignment", "N/A", "Median + P90", "NO — needs changelog"),
    ("Calendar TTR", "Resolution", "Created → Resolved (all time)", "Created", "Resolved", "None (includes all)", "Median + P90/P95", "YES"),
    ("Active TTR", "Resolution", "Time in Active states only", "Created", "Resolved", "Pending Customer, Vendor, Scheduled", "Median + P90/P95", "NO — needs changelog"),
    ("TTR Distribution", "Resolution", "Resolved grouped by TTR bucket", "N/A", "N/A", "N/A", "Count by bucket", "YES"),
    ("SLA Response Compliance", "SLA", "% meeting response target by priority", "N/A", "N/A", "N/A", "Rate", "NO — no targets set"),
    ("SLA Resolution Compliance", "SLA", "% meeting resolution target by priority", "N/A", "N/A", "N/A", "Rate", "NO — no targets set"),
    ("FCR", "Quality", "Resolved L1, no escalation, not reopened 7d", "N/A", "N/A", "N/A", "Rate", "NO — needs changelog"),
    ("Reopen Rate", "Quality", "Reopened / Resolved", "N/A", "N/A", "N/A", "Rate", "NO — needs changelog"),
    ("Reassignment Rate", "Quality", "Avg assignment changes per ticket", "N/A", "N/A", "N/A", "Average", "NO — needs changelog"),
    ("CSAT", "Customer", "Avg score + % satisfied", "N/A", "N/A", "N/A", "Average + Rate", "NO — not in data"),
    ("Escalation Rate", "Escalation", "Escalated to L2/L3 / Resolved", "N/A", "N/A", "N/A", "Rate", "NO — needs changelog"),
    ("Resolved per Agent", "Productivity", "Agent resolved count / period", "N/A", "N/A", "N/A", "Count", "YES"),
    ("WIP per Agent", "Productivity", "In Progress assigned to agent", "N/A", "N/A", "N/A", "Count", "YES"),
]

for row_data in metrics_dict:
    for c, val in enumerate(row_data, 1):
        cell = ws8.cell(r, c, val)
        if val == "NO — needs changelog":
            cell.fill = warn_fill
        elif val == "YES":
            cell.fill = good_fill
    r += 1

auto_width(ws8, max_w=50)

# ── Save ──
output_path = os.path.join(REPORT_DIR, "OIT_Baseline_Metrics_2026-03-03_0825.xlsx")
wb.save(output_path)
print(f"\nReport saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Total tickets processed: {len(tickets):,}")
print(f"Resolved with TTR data: {len(resolved_with_ttr):,}")
print(f"Open backlog: {len(open_tickets):,}")
print(f"Stale tickets: {len(stale):,}")
