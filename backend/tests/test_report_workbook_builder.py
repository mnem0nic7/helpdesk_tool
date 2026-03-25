from __future__ import annotations

import zipfile
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from models import ReportConfig, ReportTemplate
from report_workbook_builder import ReportWorkbookBuilder, _template_readiness


def _make_issue(
    *,
    key: str = "OIT-123",
    priority: str = "Low",
    status: str = "Open",
    status_category: str = "To Do",
    created: str = "2026-03-10T00:00:00+00:00",
    updated: str = "2026-03-10T00:00:00+00:00",
    resolved: str | None = None,
    labels: list[str] | None = None,
    components: list[str] | None = None,
    request_type: str = "Business Application Support",
    sla_response_status: str | None = None,
    sla_resolution_status: str | None = None,
) -> dict:
    def _sla_field(status_value: str | None, elapsed_millis: int) -> dict | None:
        if not status_value:
            return None
        normalized = status_value.strip().lower()
        if normalized == "(none)":
            return None
        if normalized in {"met", "breached"}:
            return {
                "completedCycles": [
                    {
                        "elapsedTime": {"millis": elapsed_millis},
                        "breached": normalized == "breached",
                    }
                ],
                "ongoingCycle": None,
            }
        return {
            "completedCycles": [],
            "ongoingCycle": {
                "paused": normalized == "paused",
                "elapsedTime": {"millis": elapsed_millis},
            },
        }

    return {
        "key": key,
        "fields": {
            "summary": "Example issue",
            "status": {"name": status, "statusCategory": {"name": status_category}},
            "priority": {"name": priority},
            "assignee": {"displayName": "Taylor Ops", "accountId": "acc-taylor"},
            "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
            "issuetype": {"name": "Service request"},
            "resolution": {"name": "Done"} if resolved else None,
            "created": created,
            "updated": updated,
            "resolutiondate": resolved,
            "labels": labels or [],
            "components": [{"name": value} for value in (components or [])],
            "customfield_11266": _sla_field(sla_response_status, 3_600_000),
            "customfield_11264": _sla_field(sla_resolution_status, 7_200_000),
            "customfield_11239": "Service requests",
            "customfield_10010": {"requestType": {"name": request_type}},
            "customfield_10700": [],
            "attachment": [],
            "comment": {"total": 2, "comments": []},
        },
    }


def _make_template(
    *,
    id: str,
    name: str,
    category: str,
    group_by: str,
    sort_field: str,
    readiness: str = "ready",
    notes: str = "",
) -> ReportTemplate:
    return ReportTemplate(
        id=id,
        site_scope="primary",
        name=name,
        description="",
        category=category,
        notes=notes,
        readiness=readiness,
        is_seed=True,
        include_in_master_export=True,
        created_at="2026-03-24T00:00:00+00:00",
        updated_at="2026-03-24T00:00:00+00:00",
        created_by_email="",
        created_by_name="",
        updated_by_email="",
        updated_by_name="",
        config=ReportConfig(group_by=group_by, sort_field=sort_field),
    )


def test_apply_changelog_tracks_reopens_priority_increases_and_assignee_changes():
    builder = ReportWorkbookBuilder(
        all_issues=[_make_issue()],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    fact = builder._facts_by_key["OIT-123"]

    builder._apply_changelog(
        fact,
        [
            {
                "created": "2026-03-10T01:00:00+00:00",
                "items": [{"field": "assignee", "fromString": "Taylor Ops", "toString": "Jordan Tech"}],
            },
            {
                "created": "2026-03-10T02:00:00+00:00",
                "items": [{"field": "priority", "fromString": "Low", "toString": "High"}],
            },
            {
                "created": "2026-03-10T03:00:00+00:00",
                "items": [{"field": "status", "fromString": "Open", "toString": "Resolved"}],
            },
            {
                "created": "2026-03-10T04:00:00+00:00",
                "items": [{"field": "status", "fromString": "Resolved", "toString": "In Progress"}],
            },
            {
                "created": "2026-03-10T05:00:00+00:00",
                "items": [{"field": "assignee", "fromString": "Jordan Tech", "toString": "Morgan Lead"}],
            },
        ],
    )

    assert fact.assignee_change_count == 2
    assert fact.priority_increase_count == 1
    assert fact.reopen_count == 1
    assert fact.first_resolved_dt is not None
    assert fact.is_escalated is True
    assert "Reassigned more than once" in fact.escalation_reasons
    assert "Priority increased" in fact.escalation_reasons
    assert date(2026, 3, 10) in fact.escalation_event_dates


def test_build_data_gaps_marks_csat_as_gap():
    builder = ReportWorkbookBuilder(
        all_issues=[_make_issue(key="OIT-555")],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )

    gaps = builder._build_data_gaps(template=None, report_name="Customer Satisfaction (CSAT)", facts=list(builder._facts_by_key.values()))

    assert any(gap.readiness == "gap" and "survey" in gap.limitation.lower() for gap in gaps)


def test_trend_rows_count_escalations_by_event_day_not_last_update_day():
    builder = ReportWorkbookBuilder(
        all_issues=[
            _make_issue(
                key="OIT-201",
                created="2026-03-01T00:00:00+00:00",
                updated="2026-03-06T00:00:00+00:00",
                resolved="2026-03-06T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
            )
        ],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    fact = builder._facts_by_key["OIT-201"]
    builder._apply_changelog(
        fact,
        [
            {
                "created": "2026-03-03T01:00:00+00:00",
                "items": [{"field": "assignee", "fromString": "Taylor Ops", "toString": "Jordan Tech"}],
            },
            {
                "created": "2026-03-05T02:00:00+00:00",
                "items": [{"field": "assignee", "fromString": "Jordan Tech", "toString": "Morgan Lead"}],
            },
        ],
    )

    trend_rows = {row["date"]: row for row in builder._build_trend_rows(list(builder._facts_by_key.values()))}

    assert trend_rows["2026-03-05"]["escalation_count"] == 1
    assert trend_rows["2026-03-06"]["escalation_count"] == 0


def test_master_changelog_prefetch_is_skipped_for_large_exports(monkeypatch):
    builder = ReportWorkbookBuilder(
        all_issues=[
            _make_issue(
                key=f"OIT-ESC-{idx}",
                created="2026-03-05T00:00:00+00:00",
                updated="2026-03-23T00:00:00+00:00",
                resolved="2026-03-23T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
            )
            for idx in range(300)
        ],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=True,
    )
    templates = [
        _make_template(
            id="tpl-escalation",
            name="Escalation Rate",
            category="Operational",
            group_by="assignee",
            sort_field="updated",
            readiness="proxy",
        ),
    ]
    called = False

    def _unexpected_fetch(_: object) -> None:
        nonlocal called
        called = True

    monkeypatch.setattr(builder, "ensure_changelogs", _unexpected_fetch)

    builder._prepare_master_changelogs(templates)

    assert called is False
    fact = builder._facts_by_key["OIT-ESC-0"]
    assert fact.changelog_loaded is True
    assert "Skipped Jira changelog fetch for large master export" in fact.changelog_error
    assert _template_readiness(templates[0], facts=list(builder._facts_by_key.values())) == "proxy"
    gaps = builder._build_data_gaps(
        template=templates[0],
        report_name=templates[0].name,
        facts=list(builder._facts_by_key.values()),
    )
    assert any("prefetch was skipped" in gap.limitation for gap in gaps)
    assert not any("Some Jira changelog fetches failed" in gap.limitation for gap in gaps)


def test_template_readiness_marks_first_response_as_proxy_when_elapsed_is_missing():
    builder = ReportWorkbookBuilder(
        all_issues=[_make_issue(key="OIT-301", created="2026-03-20T00:00:00+00:00")],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    template = ReportTemplate(
        id="tpl-fr",
        site_scope="primary",
        name="First Response Time",
        description="",
        category="Operational",
        notes="Ready when Jira response timers are present.",
        readiness="ready",
        is_seed=True,
        include_in_master_export=True,
        created_at="2026-03-24T00:00:00+00:00",
        updated_at="2026-03-24T00:00:00+00:00",
        created_by_email="",
        created_by_name="",
        updated_by_email="",
        updated_by_name="",
        config=ReportConfig(group_by="sla_first_response_status"),
    )

    assert _template_readiness(template, facts=list(builder._facts_by_key.values())) == "proxy"


def test_master_workbook_hides_dashboard_helper_columns_and_aligns_first_response_gap(tmp_path: Path):
    builder = ReportWorkbookBuilder(
        all_issues=[_make_issue(key="OIT-401", created="2026-03-20T00:00:00+00:00")],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    template = ReportTemplate(
        id="tpl-fr",
        site_scope="primary",
        name="First Response Time",
        description="",
        category="Operational",
        notes="Ready when Jira response timers are present.",
        readiness="ready",
        is_seed=True,
        include_in_master_export=True,
        created_at="2026-03-24T00:00:00+00:00",
        updated_at="2026-03-24T00:00:00+00:00",
        created_by_email="",
        created_by_name="",
        updated_by_email="",
        updated_by_name="",
        config=ReportConfig(group_by="sla_first_response_status"),
    )
    path = tmp_path / "master.xlsx"

    builder.build_master_report(path=str(path), templates=[template])

    workbook = load_workbook(path)
    dashboard = workbook["Executive Dashboard"]
    report_index = workbook["Report Index"]
    data_gaps = workbook["Data Gaps"]
    helper_sheet = workbook["Executive Dashboard Data"]
    detail_sheet = workbook["First Response Time 7d"]
    trends = workbook["Trends"]

    assert dashboard["I1"].value is None
    assert dashboard["M1"].value is None
    assert helper_sheet.sheet_state == "hidden"
    assert report_index["A1"].value == "Status"
    assert report_index["A2"].value == "⚠️"
    assert report_index["I2"].value == "proxy"
    assert dashboard["F6"].value == "Trend"
    assert dashboard["G6"].value == "Key Findings & Actions"
    assert data_gaps["C4"].value == "proxy"
    assert detail_sheet["B7"].value == "proxy"
    assert detail_sheet["B7"].comment is not None
    assert detail_sheet["H13"].value == "Δ Count vs Prior"
    assert detail_sheet["A15"].value == "Total"
    assert isinstance(detail_sheet["B15"].value, str) and detail_sheet["B15"].value.startswith("=")
    assert detail_sheet["A16"].value == "First Response SLA Met %"
    assert trends["I4"].value == "Created (7d MA)"
    assert trends["L4"].value == "Day"


def test_master_workbook_adds_percent_columns_total_rows_and_dashboard_formulas(tmp_path: Path):
    builder = ReportWorkbookBuilder(
        all_issues=[
            _make_issue(
                key="OIT-501",
                created="2026-03-20T00:00:00+00:00",
                updated="2026-03-22T00:00:00+00:00",
                resolved="2026-03-22T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
                priority="High",
                request_type="Access Request",
                sla_response_status="Met",
                sla_resolution_status="Met",
            ),
            _make_issue(
                key="OIT-502",
                created="2026-03-21T00:00:00+00:00",
                updated="2026-03-23T00:00:00+00:00",
                resolved="2026-03-23T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
                priority="Low",
                request_type="Access Request",
                sla_response_status="BREACHED",
                sla_resolution_status="BREACHED",
            ),
            _make_issue(
                key="OIT-503",
                created="2026-03-21T00:00:00+00:00",
                status="In Progress",
                status_category="In Progress",
                priority="Medium",
                request_type="Laptop",
                sla_response_status="Running",
                sla_resolution_status="Running",
            ),
        ],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    templates = [
        _make_template(id="tpl-mttr", name="Mean Time to Resolution", category="Executive", group_by="priority", sort_field="resolved"),
        _make_template(id="tpl-sla", name="SLA Compliance Rate", category="Executive", group_by="sla_resolution_status", sort_field="created"),
        _make_template(id="tpl-volume", name="Ticket Volume by Category", category="Executive", group_by="request_type", sort_field="created"),
        _make_template(id="tpl-backlog", name="Backlog Size & Aging", category="Operational", group_by="status", sort_field="created"),
        _make_template(id="tpl-escalation", name="Escalation Rate", category="Operational", group_by="assignee", sort_field="updated", readiness="proxy", notes="Proxy escalation review."),
        _make_template(id="tpl-fr", name="First Response Time", category="Operational", group_by="sla_first_response_status", sort_field="created"),
        _make_template(id="tpl-fcr", name="First Contact Resolution", category="Quality", group_by="request_type", sort_field="resolved", readiness="proxy", notes="Proxy FCR review."),
    ]
    path = tmp_path / "master-structure.xlsx"

    builder.build_master_report(path=str(path), templates=templates)

    workbook = load_workbook(path)
    ticket_volume = workbook["Ticket Volume by Category 7d"]
    backlog_7 = workbook["Backlog Size & Aging 7d"]
    backlog_30 = workbook["Backlog Size & Aging 30d"]
    escalation = workbook["Escalation Rate 30d"]
    fcr = workbook["First Contact Resolution 30d"]
    dashboard = workbook["Executive Dashboard"]

    ticket_total_row = next(row for row in range(14, ticket_volume.max_row + 1) if ticket_volume[f"A{row}"].value == "Total")
    assert ticket_volume["I13"].value == "% of Total"
    assert ticket_volume[f"I{ticket_total_row}"].number_format == "0.0%"
    assert isinstance(ticket_volume[f"I{ticket_total_row}"].value, str) and ticket_volume[f"I{ticket_total_row}"].value.startswith("=")
    assert backlog_7["C13"].value == "In Progress"
    assert "Acknowledged" not in [backlog_7.cell(13, idx).value for idx in range(1, 13)]
    assert backlog_7["L13"].value == "Δ Count vs Prior"
    assert backlog_30["C13"].value == "Acknowledged"
    assert backlog_30["M13"].value == "Δ Count vs Prior"
    assert escalation["I14"].number_format == "0.0%"
    assert fcr["I14"].number_format == "0.0%"
    assert isinstance(dashboard["B7"].value, str) and dashboard["B7"].value.startswith("=")
    assert isinstance(dashboard["B11"].value, str) and dashboard["B11"].value.startswith("=")
    assert isinstance(dashboard["F7"].value, str) and dashboard["F7"].value.startswith("=")
    assert dashboard["G6"].value == "Key Findings & Actions"

    with zipfile.ZipFile(path) as workbook_zip:
        worksheet_xml = "\n".join(
            workbook_zip.read(name).decode("utf-8", "ignore")
            for name in workbook_zip.namelist()
            if name.startswith("xl/worksheets/sheet")
        )

    assert "sparklineGroup" in worksheet_xml


def test_master_workbook_flags_escalation_anomaly(tmp_path: Path):
    issues = []
    for idx in range(120):
        issues.append(
            _make_issue(
                key=f"OIT-A-{idx}",
                created="2026-03-01T00:00:00+00:00",
                updated="2026-03-06T00:00:00+00:00",
                resolved="2026-03-06T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
            )
        )
    for idx in range(5):
        issues.append(
            _make_issue(
                key=f"OIT-B-{idx}",
                created="2026-03-01T00:00:00+00:00",
                updated="2026-03-05T00:00:00+00:00",
                resolved="2026-03-05T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
            )
        )

    builder = ReportWorkbookBuilder(
        all_issues=issues,
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    for idx in range(120):
        fact = builder._facts_by_key[f"OIT-A-{idx}"]
        builder._apply_changelog(
            fact,
            [
                {
                    "created": "2026-03-06T01:00:00+00:00",
                    "items": [{"field": "assignee", "fromString": "Taylor Ops", "toString": "Jordan Tech"}],
                },
                {
                    "created": "2026-03-06T02:00:00+00:00",
                    "items": [{"field": "assignee", "fromString": "Jordan Tech", "toString": "Morgan Lead"}],
                },
            ],
        )
        fact.changelog_loaded = True
    for idx in range(5):
        fact = builder._facts_by_key[f"OIT-B-{idx}"]
        builder._apply_changelog(
            fact,
            [
                {
                    "created": "2026-03-05T01:00:00+00:00",
                    "items": [{"field": "assignee", "fromString": "Taylor Ops", "toString": "Jordan Tech"}],
                },
                {
                    "created": "2026-03-05T02:00:00+00:00",
                    "items": [{"field": "assignee", "fromString": "Jordan Tech", "toString": "Morgan Lead"}],
                },
            ],
        )
        fact.changelog_loaded = True

    template = ReportTemplate(
        id="tpl-escalation",
        site_scope="primary",
        name="Escalation Rate",
        description="",
        category="Operational",
        notes="Proxy escalation review.",
        readiness="proxy",
        is_seed=True,
        include_in_master_export=True,
        created_at="2026-03-24T00:00:00+00:00",
        updated_at="2026-03-24T00:00:00+00:00",
        created_by_email="",
        created_by_name="",
        updated_by_email="",
        updated_by_name="",
        config=ReportConfig(group_by="assignee", sort_field="updated"),
    )
    path = tmp_path / "master-escalation.xlsx"

    builder.build_master_report(path=str(path), templates=[template])

    workbook = load_workbook(path)
    trends = workbook["Trends"]
    dashboard = workbook["Executive Dashboard"]
    data_gaps = workbook["Data Gaps"]

    anomaly_row = None
    for row_idx in range(5, trends.max_row + 1):
        if trends[f"A{row_idx}"].value == "2026-03-06":
            anomaly_row = row_idx
            break

    assert anomaly_row is not None
    assert trends[f"H{anomaly_row}"].comment is not None
    assert "suspected outlier" in dashboard["G11"].value.lower()
    assert any(data_gaps[f"C{row_idx}"].value == "anomaly" for row_idx in range(4, data_gaps.max_row + 1))


def test_master_workbook_chart_xml_uses_dashed_secondary_mttr_axis(tmp_path: Path):
    builder = ReportWorkbookBuilder(
        all_issues=[
            _make_issue(
                key=f"OIT-MTTR-{idx}",
                created="2026-03-20T00:00:00+00:00",
                updated="2026-03-21T00:00:00+00:00",
                resolved="2026-03-21T00:00:00+00:00",
                status="Resolved",
                status_category="Done",
            )
            for idx in range(3)
        ],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    template = ReportTemplate(
        id="tpl-mttr",
        site_scope="primary",
        name="Mean Time to Resolution",
        description="",
        category="Executive",
        notes="",
        readiness="ready",
        is_seed=True,
        include_in_master_export=True,
        created_at="2026-03-24T00:00:00+00:00",
        updated_at="2026-03-24T00:00:00+00:00",
        created_by_email="",
        created_by_name="",
        updated_by_email="",
        updated_by_name="",
        config=ReportConfig(group_by="priority", sort_field="resolved"),
    )
    path = tmp_path / "chart-master.xlsx"

    builder.build_master_report(path=str(path), templates=[template])

    with zipfile.ZipFile(path) as workbook_zip:
        chart_xml = "\n".join(
            workbook_zip.read(name).decode("utf-8", "ignore")
            for name in workbook_zip.namelist()
            if name.startswith("xl/charts/chart")
        )

    assert "MTTR P95 (hours)" in chart_xml
    assert "dash" in chart_xml


def test_followup_template_readiness_and_gaps_do_not_depend_on_report_name():
    builder = ReportWorkbookBuilder(
        all_issues=[_make_issue(key="OIT-FU-1", created="2026-03-20T00:00:00+00:00")],
        site_scope="primary",
        today=date(2026, 3, 24),
        enable_changelog_fetch=False,
    )
    template = ReportTemplate(
        id="tpl-followup",
        site_scope="primary",
        name="Ticket Touch Discipline",
        description="",
        category="Operational",
        notes="Proxy touch review.",
        readiness="proxy",
        is_seed=True,
        include_in_master_export=False,
        created_at="2026-03-24T00:00:00+00:00",
        updated_at="2026-03-24T00:00:00+00:00",
        created_by_email="",
        created_by_name="",
        updated_by_email="",
        updated_by_name="",
        config=ReportConfig(group_by="response_followup_status", sort_field="created"),
    )

    facts = list(builder._facts_by_key.values())
    assert _template_readiness(template, facts=facts) == "proxy"

    gaps = builder._build_data_gaps(
        template=template,
        report_name=template.name,
        facts=facts,
    )

    assert any("daily public follow-up" in gap.limitation.lower() for gap in gaps)
