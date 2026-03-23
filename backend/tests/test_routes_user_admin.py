from __future__ import annotations

from io import BytesIO
from unittest.mock import MagicMock

from openpyxl import load_workbook


def test_user_admin_capabilities_return_primary_payload(test_client, monkeypatch):
    import routes_user_admin

    mock_providers = MagicMock()
    mock_providers.get_capabilities.return_value = {
        "can_manage_users": True,
        "enabled_providers": {"entra": True, "mailbox": False, "device_management": True},
        "supported_actions": ["disable_sign_in", "revoke_sessions"],
        "license_catalog": [{"sku_id": "sku-1", "sku_part_number": "M365_BUSINESS_PREMIUM", "display_name": "M365 Business Premium"}],
        "group_catalog": [],
        "role_catalog": [],
        "conditional_access_exception_groups": [],
    }
    monkeypatch.setattr(routes_user_admin, "user_admin_providers", mock_providers)

    resp = test_client.get("/api/user-admin/capabilities", headers={"host": "it-app.movedocs.com"})

    assert resp.status_code == 200
    data = resp.json()
    assert data["can_manage_users"] is True
    assert data["enabled_providers"]["entra"] is True
    assert data["enabled_providers"]["mailbox"] is False


def test_user_admin_routes_are_not_available_on_azure_or_oasis(test_client):
    azure = test_client.get("/api/user-admin/capabilities", headers={"host": "azure.movedocs.com"})
    oasis = test_client.get("/api/user-admin/capabilities", headers={"host": "oasisdev.movedocs.com"})

    assert azure.status_code == 404
    assert oasis.status_code == 404


def test_user_admin_capabilities_require_auth(monkeypatch):
    import issue_cache
    import routes_metrics
    import routes_tickets
    import routes_chart
    import routes_export
    import routes_cache
    import routes_triage
    import routes_azure
    import routes_user_admin
    import azure_cache as azure_cache_module
    import user_admin_jobs as user_admin_jobs_module
    import user_admin_providers as user_admin_providers_module
    from unittest.mock import AsyncMock

    mock_cache = MagicMock()
    mock_cache.get_filtered_issues.return_value = []
    mock_cache.get_all_issues.return_value = []
    mock_cache.initialized = True
    mock_cache.refreshing = False
    mock_cache.last_refresh = None
    mock_cache.status.return_value = {
        "initialized": True,
        "refreshing": False,
        "issue_count": 0,
        "filtered_count": 0,
        "last_refresh": None,
    }
    for mod in [issue_cache, routes_metrics, routes_tickets, routes_chart, routes_export, routes_cache, routes_triage]:
        monkeypatch.setattr(mod, "cache", mock_cache)

    mock_azure_cache = MagicMock()
    mock_azure_cache.start_background_refresh = AsyncMock()
    mock_azure_cache.stop_background_refresh = AsyncMock()
    mock_azure_cache.status.return_value = {
        "configured": False,
        "initialized": True,
        "refreshing": False,
        "last_refresh": None,
        "datasets": [],
    }
    mock_azure_cache.get_overview.return_value = {
        "subscriptions": 0,
        "management_groups": 0,
        "resources": 0,
        "role_assignments": 0,
        "users": 0,
        "groups": 0,
        "enterprise_apps": 0,
        "app_registrations": 0,
        "directory_roles": 0,
        "cost": {
            "lookback_days": 30,
            "total_cost": 0.0,
            "currency": "USD",
            "top_service": "",
            "top_subscription": "",
            "top_resource_group": "",
            "recommendation_count": 0,
            "potential_monthly_savings": 0.0,
        },
        "datasets": [],
        "last_refresh": None,
    }
    mock_azure_cache.get_cost_summary.return_value = mock_azure_cache.get_overview.return_value["cost"]
    mock_azure_cache.get_cost_trend.return_value = []
    mock_azure_cache.get_cost_breakdown.return_value = []
    mock_azure_cache.get_advisor.return_value = []
    mock_azure_cache.list_resources.return_value = {"resources": [], "matched_count": 0, "total_count": 0}
    mock_azure_cache.list_directory_objects.return_value = []
    mock_azure_cache.get_grounding_context.return_value = {}
    monkeypatch.setattr(azure_cache_module, "azure_cache", mock_azure_cache)
    monkeypatch.setattr(routes_azure, "azure_cache", mock_azure_cache)

    mock_user_admin_jobs = MagicMock()
    mock_user_admin_jobs.start_worker = AsyncMock()
    mock_user_admin_jobs.stop_worker = AsyncMock()
    monkeypatch.setattr(user_admin_jobs_module, "user_admin_jobs", mock_user_admin_jobs)
    monkeypatch.setattr(routes_user_admin, "user_admin_jobs", mock_user_admin_jobs)

    mock_user_admin_providers = MagicMock()
    monkeypatch.setattr(user_admin_providers_module, "user_admin_providers", mock_user_admin_providers)
    monkeypatch.setattr(routes_user_admin, "user_admin_providers", mock_user_admin_providers)

    import main
    mock_ai_work_scheduler = MagicMock()
    mock_ai_work_scheduler.start_worker = AsyncMock()
    mock_ai_work_scheduler.stop_worker = AsyncMock()
    monkeypatch.setattr(main, "ai_work_scheduler", mock_ai_work_scheduler)
    mock_technician_scoring_manager = MagicMock()
    mock_technician_scoring_manager.start_worker = AsyncMock()
    mock_technician_scoring_manager.stop_worker = AsyncMock()
    monkeypatch.setattr(main, "technician_scoring_manager", mock_technician_scoring_manager)
    mock_kb_store = MagicMock()
    mock_kb_store.ensure_seed_articles.return_value = 0
    monkeypatch.setattr(main, "kb_store", mock_kb_store)
    from starlette.testclient import TestClient

    client = TestClient(main.app)
    resp = client.get("/api/user-admin/capabilities", headers={"host": "it-app.movedocs.com"})

    assert resp.status_code == 401


def test_create_and_poll_user_admin_job(test_client, monkeypatch):
    import routes_user_admin

    mock_jobs = MagicMock()
    mock_jobs.create_job.return_value = {
        "job_id": "job-123",
        "status": "queued",
        "action_type": "disable_sign_in",
        "provider": "entra",
        "target_user_ids": ["user-1"],
        "requested_by_email": "test@example.com",
        "requested_by_name": "Test User",
        "requested_at": "2026-03-19T00:00:00Z",
        "started_at": None,
        "completed_at": None,
        "progress_current": 0,
        "progress_total": 1,
        "progress_message": "Queued",
        "success_count": 0,
        "failure_count": 0,
        "results_ready": False,
        "error": "",
        "one_time_results_available": False,
    }
    mock_jobs.job_belongs_to.return_value = True
    mock_jobs.get_job.return_value = {
        "job_id": "job-123",
        "status": "completed",
        "action_type": "disable_sign_in",
        "provider": "entra",
        "target_user_ids": ["user-1"],
        "requested_by_email": "test@example.com",
        "requested_by_name": "Test User",
        "requested_at": "2026-03-19T00:00:00Z",
        "started_at": "2026-03-19T00:01:00Z",
        "completed_at": "2026-03-19T00:02:00Z",
        "progress_current": 1,
        "progress_total": 1,
        "progress_message": "Completed",
        "success_count": 1,
        "failure_count": 0,
        "results_ready": True,
        "error": "",
        "one_time_results_available": True,
    }
    mock_jobs.get_job_results.return_value = [
        {
            "target_user_id": "user-1",
            "target_display_name": "Ada Lovelace",
            "provider": "entra",
            "success": True,
            "summary": "Disabled sign-in",
            "error": "",
            "before_summary": {"enabled": True},
            "after_summary": {"enabled": False},
            "one_time_secret": None,
        }
    ]
    monkeypatch.setattr(routes_user_admin, "user_admin_jobs", mock_jobs)

    create_resp = test_client.post(
        "/api/user-admin/jobs",
        headers={"host": "it-app.movedocs.com"},
        json={"action_type": "disable_sign_in", "target_user_ids": ["user-1"], "params": {}},
    )
    status_resp = test_client.get("/api/user-admin/jobs/job-123", headers={"host": "it-app.movedocs.com"})
    results_resp = test_client.get("/api/user-admin/jobs/job-123/results", headers={"host": "it-app.movedocs.com"})

    assert create_resp.status_code == 200
    assert status_resp.status_code == 200
    assert results_resp.status_code == 200
    assert results_resp.json()[0]["target_display_name"] == "Ada Lovelace"


def test_user_activity_and_audit_read_from_job_store(test_client, monkeypatch):
    import routes_user_admin

    mock_jobs = MagicMock()
    mock_jobs.list_audit.side_effect = [
        [
            {
                "audit_id": "audit-1",
                "job_id": "job-1",
                "actor_email": "tech@example.com",
                "actor_name": "Tech User",
                "target_user_id": "user-1",
                "target_display_name": "Ada Lovelace",
                "provider": "entra",
                "action_type": "revoke_sessions",
                "params_summary": {},
                "before_summary": {},
                "after_summary": {"sessions_revoked": True},
                "status": "success",
                "error": "",
                "created_at": "2026-03-19T00:00:00Z",
            }
        ],
        [
            {
                "audit_id": "audit-2",
                "job_id": "job-2",
                "actor_email": "tech@example.com",
                "actor_name": "Tech User",
                "target_user_id": "user-2",
                "target_display_name": "Grace Hopper",
                "provider": "entra",
                "action_type": "disable_sign_in",
                "params_summary": {},
                "before_summary": {"enabled": True},
                "after_summary": {"enabled": False},
                "status": "success",
                "error": "",
                "created_at": "2026-03-19T01:00:00Z",
            }
        ],
    ]
    monkeypatch.setattr(routes_user_admin, "user_admin_jobs", mock_jobs)

    activity_resp = test_client.get("/api/user-admin/users/user-1/activity", headers={"host": "it-app.movedocs.com"})
    audit_resp = test_client.get("/api/user-admin/audit", headers={"host": "it-app.movedocs.com"})

    assert activity_resp.status_code == 200
    assert audit_resp.status_code == 200
    assert activity_resp.json()[0]["target_user_id"] == "user-1"
    assert audit_resp.json()[0]["target_user_id"] == "user-2"


def test_user_export_csv_applies_report_filter(test_client, monkeypatch):
    import routes_user_admin

    mock_cache = MagicMock()
    mock_cache.list_directory_objects.return_value = [
        {
            "id": "user-1",
            "display_name": "Disabled Licensed",
            "object_type": "user",
            "principal_name": "disabled@example.com",
            "mail": "disabled@example.com",
            "app_id": "",
            "enabled": False,
            "extra": {
                "user_type": "Member",
                "is_licensed": "true",
                "license_count": "2",
                "sku_part_numbers": "M365_BUSINESS_PREMIUM, EMS",
                "last_successful_utc": "",
                "last_successful_local": "",
            },
        },
        {
            "id": "user-2",
            "display_name": "Enabled Licensed",
            "object_type": "user",
            "principal_name": "enabled@example.com",
            "mail": "enabled@example.com",
            "app_id": "",
            "enabled": True,
            "extra": {
                "user_type": "Member",
                "is_licensed": "true",
                "license_count": "1",
                "sku_part_numbers": "M365_BUSINESS_BASIC",
                "last_successful_utc": "2026-03-18T00:00:00+00:00",
                "last_successful_local": "2026-03-17 17:00 PT",
            },
        },
    ]
    monkeypatch.setattr(routes_user_admin, "azure_cache", mock_cache)

    resp = test_client.get(
        "/api/user-admin/users/export.csv",
        headers={"host": "it-app.movedocs.com"},
        params={"report_filter": "disabled_licensed", "scope": "filtered"},
    )

    assert resp.status_code == 200
    assert "Disabled Licensed" in resp.text
    assert "Enabled Licensed" not in resp.text
    mock_cache.list_directory_objects.assert_called_once_with("users", search="")


def test_user_export_xlsx_supports_scope_all(test_client, monkeypatch):
    import routes_user_admin

    mock_cache = MagicMock()
    mock_cache.list_directory_objects.return_value = [
        {
            "id": "user-1",
            "display_name": "Ada Lovelace",
            "object_type": "user",
            "principal_name": "ada@example.com",
            "mail": "ada@example.com",
            "app_id": "",
            "enabled": False,
            "extra": {
                "user_type": "Member",
                "is_licensed": "true",
                "license_count": "1",
                "sku_part_numbers": "M365_BUSINESS_PREMIUM",
                "last_successful_utc": "",
                "last_successful_local": "",
            },
        },
        {
            "id": "user-2",
            "display_name": "Grace Hopper",
            "object_type": "user",
            "principal_name": "grace@example.com",
            "mail": "grace@example.com",
            "app_id": "",
            "enabled": True,
            "extra": {
                "user_type": "Member",
                "is_licensed": "",
                "license_count": "0",
                "sku_part_numbers": "",
                "last_successful_utc": "2026-03-19T00:00:00+00:00",
                "last_successful_local": "2026-03-18 17:00 PT",
            },
        },
    ]
    monkeypatch.setattr(routes_user_admin, "azure_cache", mock_cache)

    resp = test_client.get(
        "/api/user-admin/users/export.xlsx",
        headers={"host": "it-app.movedocs.com"},
        params={"search": "ada", "status": "disabled", "scope": "all"},
    )

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(resp.content))
    sheet = workbook.active
    names = [sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)]
    assert "Ada Lovelace" in names
    assert "Grace Hopper" in names
    mock_cache.list_directory_objects.assert_called_once_with("users", search="")
