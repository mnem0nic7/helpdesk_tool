"""Tests for report builder and export routes (~7 tests)."""

from __future__ import annotations

from io import BytesIO
import zipfile
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from report_template_store import ReportTemplateStore


def _make_workload_issue(
    *,
    key: str,
    summary: str,
    status: str,
    status_category: str,
    assignee: str,
    created: str,
    updated: str,
    resolved: str | None = None,
    priority: str = "Medium",
    request_type: str = "Business Application Support",
    components: list[str] | None = None,
    work_category: str = "Service requests",
    oasisdev: bool = True,
) -> dict:
    return {
        "key": key,
        "fields": {
            "summary": summary,
            "status": {"name": status, "statusCategory": {"name": status_category}},
            "priority": {"name": priority},
            "assignee": {
                "displayName": assignee,
                "accountId": f"acc-{assignee.lower().replace(' ', '-')}",
            },
            "reporter": {
                "displayName": "Grant Reviewer",
                "accountId": "acc-grant-reviewer",
            },
            "issuetype": {"name": "[System] Service request"},
            "resolution": {"name": "Done"} if resolved else None,
            "created": created,
            "updated": updated,
            "resolutiondate": resolved,
            "labels": ["oasisdev"] if oasisdev else [],
            "components": [{"name": name} for name in (components or [])],
            "customfield_11239": work_category,
            "customfield_10010": {"requestType": {"name": request_type}},
            "customfield_11266": None,
            "customfield_11264": None,
            "customfield_10700": [],
            "attachment": [],
        },
    }


class TestReportPreview:
    """POST /api/report/preview"""

    def test_flat_preview(self, test_client):
        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": [],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["grouped"] is False
        assert isinstance(data["rows"], list)
        assert data["total_count"] >= 0

    def test_custom_columns(self, test_client):
        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "summary", "status"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        rows = resp.json()["rows"]
        if rows:
            assert set(rows[0].keys()) == {"key", "summary", "status"}

    def test_grouped_preview(self, test_client):
        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": [],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": "priority",
            "include_excluded": False,
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["grouped"] is True
        if data["rows"]:
            assert "group" in data["rows"][0]
            assert "count" in data["rows"][0]

    def test_with_filters(self, test_client):
        resp = test_client.post("/api/report/preview", json={
            "filters": {"priority": "High"},
            "columns": [],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200

    def test_preview_includes_comment_metadata_when_requested(self, test_client, mock_cache):
        issue = {
            "key": "OIT-777",
            "fields": {
                "summary": "Comment-heavy ticket",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "Medium"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-01T10:00:00+00:00",
                "updated": "2026-03-02T12:00:00+00:00",
                "resolutiondate": "2026-03-02T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": None,
                "customfield_11239": "Service requests",
                "customfield_11266": None,
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
                "comment": {
                    "total": 2,
                    "comments": [
                        {
                            "created": "2026-03-01T11:00:00+00:00",
                            "updated": "2026-03-01T11:00:00+00:00",
                            "author": {"displayName": "Jordan Commenter"},
                        },
                        {
                            "created": "2026-03-02T12:00:00+00:00",
                            "updated": "2026-03-02T12:00:00+00:00",
                            "author": {"displayName": "Taylor Resolver"},
                        },
                    ],
                },
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "comment_count", "last_comment_author", "last_comment_date"],
            "sort_field": "comment_count",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert row["comment_count"] == 2
        assert row["last_comment_author"] == "Taylor Resolver"
        assert row["last_comment_date"] == "2026-03-02T12:00:00+00:00"

    def test_preview_includes_occ_ticket_id_when_requested(self, test_client, mock_cache):
        issue = {
            "key": "OIT-779",
            "fields": {
                "summary": "Imported from OCC",
                "description": "OCC Ticket Created By: Libra PhishER | OCC Ticket ID: LIBRA-SR-075203",
                "status": {"name": "Open", "statusCategory": {"name": "To Do"}},
                "priority": {"name": "Medium"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": None,
                "created": "2026-03-02T10:00:00+00:00",
                "updated": "2026-03-02T10:00:00+00:00",
                "resolutiondate": None,
                "labels": [],
                "components": [],
                "customfield_10010": None,
                "customfield_11239": "Service requests",
                "customfield_11266": None,
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "occ_ticket_id"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert row["occ_ticket_id"] == "LIBRA-SR-075203"

    def test_preview_includes_occ_ticket_id_from_comment_history_when_requested(self, test_client, mock_cache):
        issue = {
            "key": "OIT-779A",
            "fields": {
                "summary": "Imported from OCC comment",
                "description": "Awaiting OCC reference.",
                "status": {"name": "Open", "statusCategory": {"name": "To Do"}},
                "priority": {"name": "Medium"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": None,
                "created": "2026-03-02T10:00:00+00:00",
                "updated": "2026-03-02T10:00:00+00:00",
                "resolutiondate": None,
                "labels": [],
                "components": [],
                "customfield_10010": None,
                "customfield_11239": "Service requests",
                "customfield_11266": None,
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
                "comment": {
                    "total": 1,
                    "comments": [
                        {
                            "created": "2026-03-27T17:07:33+00:00",
                            "updated": "2026-03-27T17:07:33+00:00",
                            "author": {"displayName": "OSIJIRAOCC", "accountId": "acc-occ"},
                            "body": "Successfully OCC ticket Created with Ticket Id: LIBRA-SR-075206",
                            "public": True,
                        }
                    ],
                },
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "occ_ticket_id"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert row["occ_ticket_id"] == "LIBRA-SR-075206"

    def test_preview_includes_first_contact_when_requested(self, test_client, mock_cache):
        issue = {
            "key": "OIT-780",
            "fields": {
                "summary": "Needs response timeline",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-02T08:00:00+00:00",
                "updated": "2026-03-02T12:00:00+00:00",
                "resolutiondate": "2026-03-02T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": None,
                "customfield_11239": "Service requests",
                "customfield_11266": {
                    "completedCycles": [
                        {
                            "breached": False,
                            "stopTime": {"iso8601": "2026-03-02T09:00:07+00:00"},
                        }
                    ]
                },
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
                "comment": {"total": 0, "comments": []},
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "created", "first_contact_date"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert row["created"] == "2026-03-02T08:00:00+00:00"
        assert row["first_contact_date"] == "2026-03-02T09:00:07+00:00"

    def test_preview_prefers_outreach_comment_for_first_contact_when_available(self, test_client, mock_cache):
        issue = {
            "key": "OIT-780A",
            "fields": {
                "summary": "Outreach note should drive first contact",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-02T08:00:00+00:00",
                "updated": "2026-03-02T12:00:00+00:00",
                "resolutiondate": "2026-03-02T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": None,
                "customfield_11239": "Service requests",
                "customfield_11266": {
                    "completedCycles": [
                        {
                            "breached": False,
                            "stopTime": {"iso8601": "2026-03-02T08:10:00+00:00"},
                        }
                    ]
                },
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
                "comment": {
                    "total": 3,
                    "comments": [
                        {
                            "created": "2026-03-02T08:05:00+00:00",
                            "updated": "2026-03-02T08:05:00+00:00",
                            "author": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                            "body": "Any update?",
                            "public": True,
                        },
                        {
                            "created": "2026-03-02T08:10:00+00:00",
                            "updated": "2026-03-02T08:10:00+00:00",
                            "author": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                            "body": "Acknowledged, investigating now.",
                            "public": True,
                        },
                        {
                            "created": "2026-03-02T08:22:00+00:00",
                            "updated": "2026-03-02T08:22:00+00:00",
                            "author": {"displayName": "Alex Agent", "accountId": "acc-alex-agent"},
                            "body": "Reached out to the user and left voicemail.",
                            "public": False,
                        },
                    ],
                },
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "created", "first_contact_date"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        row = resp.json()["rows"][0]
        assert row["first_contact_date"] == "2026-03-02T08:22:00+00:00"

    def test_grouped_preview_supports_response_followup_compliance(self, test_client, mock_cache):
        issue = {
            "key": "OIT-778",
            "fields": {
                "summary": "Needs a prompt response",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-02T08:00:00+00:00",
                "updated": "2026-03-02T12:00:00+00:00",
                "resolutiondate": "2026-03-02T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": {"completedCycles": [{"breached": False}]},
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
                "comment": {
                    "total": 1,
                    "comments": [
                        {
                            "created": "2026-03-02T09:00:00+00:00",
                            "updated": "2026-03-02T09:00:00+00:00",
                            "author": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                        },
                    ],
                },
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": [],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": "response_followup_status",
            "include_excluded": False,
        })

        assert resp.status_code == 200
        payload = resp.json()
        assert payload["grouped"] is True
        assert payload["rows"][0]["group"] == "Met"
        assert payload["rows"][0]["count"] == 1

    def test_preview_excludes_non_tracked_moved_tickets(self, test_client, mock_cache):
        tracked_issue = {
            "key": "OIT-779",
            "fields": {
                "summary": "Tracked ticket",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-02T08:00:00+00:00",
                "updated": "2026-03-02T12:00:00+00:00",
                "resolutiondate": "2026-03-02T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": None,
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
            },
        }
        moved_issue = {
            "key": "MSD-779",
            "fields": {
                "summary": "Moved ticket",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "Low"},
                "assignee": {"displayName": "Moved User", "accountId": "acc-moved"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-02T08:00:00+00:00",
                "updated": "2026-03-02T12:00:00+00:00",
                "resolutiondate": "2026-03-02T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": None,
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
            },
        }
        mock_cache.get_all_issues.return_value = [tracked_issue, moved_issue]
        mock_cache.get_filtered_issues.return_value = [tracked_issue, moved_issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": True,
        })

        assert resp.status_code == 200
        keys = [row["key"] for row in resp.json()["rows"]]
        assert keys == ["OIT-779"]

    def test_preview_honors_libra_support_filter(self, test_client, mock_cache):
        libra_issue = _make_workload_issue(
            key="OIT-LIBRA-1",
            summary="Libra support ticket",
            status="Open",
            status_category="To Do",
            assignee="Taylor Ops",
            created="2026-03-02T08:00:00+00:00",
            updated="2026-03-02T12:00:00+00:00",
            oasisdev=False,
        )
        libra_issue["fields"]["labels"] = ["Libra_Support"]
        normal_issue = _make_workload_issue(
            key="OIT-LIBRA-2",
            summary="Standard ticket",
            status="Open",
            status_category="To Do",
            assignee="Taylor Ops",
            created="2026-03-02T08:00:00+00:00",
            updated="2026-03-02T12:00:00+00:00",
            oasisdev=False,
        )
        normal_issue["fields"]["labels"] = []
        mock_cache.get_all_issues.return_value = [libra_issue, normal_issue]
        mock_cache.get_filtered_issues.return_value = [libra_issue, normal_issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {"libra_support": "libra_support"},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })

        assert resp.status_code == 200
        keys = [row["key"] for row in resp.json()["rows"]]
        assert keys == ["OIT-LIBRA-1"]

    def test_preview_primary_include_excluded_still_ignores_oasisdev(self, test_client, mock_cache):
        primary_issue = _make_workload_issue(
            key="OIT-PRIMARY-1",
            summary="Primary scoped ticket",
            status="Open",
            status_category="To Do",
            assignee="Taylor Ops",
            created="2026-03-02T08:00:00+00:00",
            updated="2026-03-02T12:00:00+00:00",
            oasisdev=False,
        )
        oasis_issue = _make_workload_issue(
            key="OIT-OASIS-1",
            summary="OasisDev excluded ticket",
            status="Open",
            status_category="To Do",
            assignee="Taylor Ops",
            created="2026-03-02T08:00:00+00:00",
            updated="2026-03-02T12:00:00+00:00",
            oasisdev=True,
        )
        mock_cache.get_all_issues.return_value = [primary_issue, oasis_issue]
        mock_cache.get_filtered_issues.return_value = [primary_issue, oasis_issue]

        resp = test_client.post("/api/report/preview", json={
            "filters": {},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": True,
        })

        assert resp.status_code == 200
        keys = [row["key"] for row in resp.json()["rows"]]
        assert keys == ["OIT-PRIMARY-1"]

    def test_preview_rejects_invalid_custom_window(self, test_client):
        resp = test_client.post("/api/report/preview", json={
            "filters": {"created_after": "2026-03-10"},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
            "window_mode": "custom",
        })

        assert resp.status_code == 400
        assert "Created Before" in resp.json()["detail"]


class TestReportExport:
    """POST /api/report/export"""

    def test_report_export_build_runs_via_asyncio_to_thread(self, test_client, monkeypatch):
        calls: list[str] = []

        async def fake_to_thread(func, /, *args, **kwargs):
            calls.append(getattr(func, "__name__", ""))
            return func(*args, **kwargs)

        monkeypatch.setattr("routes_export.asyncio.to_thread", fake_to_thread)

        resp = test_client.post("/api/report/export", json={
            "filters": {},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        assert "_write_single_report_workbook_file" in calls

    def test_returns_excel(self, test_client):
        resp = test_client.post("/api/report/export", json={
            "filters": {},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")
        workbook = load_workbook(BytesIO(resp.content))
        assert workbook.sheetnames[0] == "Summary"
        assert "Trends" in workbook.sheetnames
        assert "30 Day" in workbook.sheetnames
        assert "7 Day" not in workbook.sheetnames
        assert "Data Gaps" in workbook.sheetnames

        workbook_zip = zipfile.ZipFile(BytesIO(resp.content))
        assert any(name.startswith("xl/charts/chart") for name in workbook_zip.namelist())

    def test_detail_export_uses_selected_created_window(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 10))

        issues = [
            _make_workload_issue(
                key=f"OIT-WIN-{offset}",
                summary=f"Windowed issue {offset}",
                status="Open",
                status_category="To Do",
                assignee="Taylor Ops",
                created=f"{(date(2026, 3, 10) - timedelta(days=offset)).isoformat()}T10:00:00+00:00",
                updated=f"{(date(2026, 3, 10) - timedelta(days=offset)).isoformat()}T10:00:00+00:00",
                oasisdev=False,
            )
            for offset in range(8)
        ]
        mock_cache.get_all_issues.return_value = issues
        mock_cache.get_filtered_issues.return_value = issues

        resp = test_client.post("/api/report/export", json={
            "filters": {},
            "columns": ["key", "summary"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
            "window_mode": "7d",
        })
        assert resp.status_code == 200

        workbook = load_workbook(BytesIO(resp.content))
        seven_day = workbook["7 Day"]

        assert seven_day["A1"].value == "Report"
        assert seven_day.row_dimensions[1].hidden is True
        assert seven_day["A2"].value == "Window"
        assert seven_day["B2"].value == "7 Day"
        assert seven_day["A3"].value == "Window Field"
        assert seven_day["B3"].value == "Created"
        assert seven_day["A13"].value == "Key"
        assert "30 Day" not in workbook.sheetnames

        seven_day_keys = [cell.value for cell in seven_day["A"] if isinstance(cell.value, str) and cell.value.startswith("OIT-WIN-")]
        assert len(seven_day_keys) == 7

    def test_grouped_export_uses_live_config_window_field_when_template_id_is_provided(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 10))

        issues = [
            _make_workload_issue(
                key="OIT-MTTR-RECENT",
                summary="Recently resolved issue",
                status="Resolved",
                status_category="Done",
                priority="High",
                assignee="Taylor Ops",
                created="2026-01-05T10:00:00+00:00",
                updated="2026-03-09T12:00:00+00:00",
                resolved="2026-03-09T12:00:00+00:00",
                oasisdev=False,
            ),
            _make_workload_issue(
                key="OIT-MTTR-OLDER",
                summary="Resolved within 30 days only",
                status="Resolved",
                status_category="Done",
                priority="Medium",
                assignee="Taylor Ops",
                created="2026-01-02T09:00:00+00:00",
                updated="2026-03-01T12:00:00+00:00",
                resolved="2026-03-01T12:00:00+00:00",
                oasisdev=False,
            ),
        ]
        mock_cache.get_all_issues.return_value = issues
        mock_cache.get_filtered_issues.return_value = issues

        templates_resp = test_client.get("/api/report/templates")
        assert templates_resp.status_code == 200
        mttr_template = next(template for template in templates_resp.json() if template["name"] == "Mean Time to Resolution")
        export_config = dict(mttr_template["config"])
        export_config["sort_field"] = "created"
        export_config["window_mode"] = "7d"

        resp = test_client.post(
            f"/api/report/export?template_id={mttr_template['id']}",
            json=export_config,
        )
        assert resp.status_code == 200

        workbook = load_workbook(BytesIO(resp.content))
        seven_day = workbook["7 Day"]
        assert seven_day["B1"].value == "Mean Time to Resolution"
        assert seven_day["B3"].value == "Created"
        assert seven_day["H13"].value == "Δ Count vs Prior"
        assert seven_day["D13"].value == "Avg TTR (h)"
        assert seven_day["E13"].value == "Median TTR (h)"
        assert seven_day["F13"].value == "P95 TTR (h)"
        assert seven_day["G13"].value == "P99 TTR (h)"

        seven_day_counts = {
            row[0].value: row[1].value
            for row in seven_day.iter_rows(min_row=14, max_col=2)
            if row[0].value in {"High", "Medium"}
        }
        assert seven_day_counts == {}

    def test_grouped_export_includes_raw_percentiles(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 10))

        issues = [
            _make_workload_issue(
                key="OIT-P95-1",
                summary="Resolved in 1 hour",
                status="Resolved",
                status_category="Done",
                priority="High",
                assignee="Taylor Ops",
                created="2026-03-09T00:00:00+00:00",
                updated="2026-03-09T01:00:00+00:00",
                resolved="2026-03-09T01:00:00+00:00",
                oasisdev=False,
            ),
            _make_workload_issue(
                key="OIT-P95-2",
                summary="Resolved in 2 hours",
                status="Resolved",
                status_category="Done",
                priority="High",
                assignee="Taylor Ops",
                created="2026-03-08T00:00:00+00:00",
                updated="2026-03-08T02:00:00+00:00",
                resolved="2026-03-08T02:00:00+00:00",
                oasisdev=False,
            ),
            _make_workload_issue(
                key="OIT-P95-3",
                summary="Resolved in 100 hours",
                status="Resolved",
                status_category="Done",
                priority="High",
                assignee="Taylor Ops",
                created="2026-03-06T00:00:00+00:00",
                updated="2026-03-10T04:00:00+00:00",
                resolved="2026-03-10T04:00:00+00:00",
                oasisdev=False,
            ),
        ]
        mock_cache.get_all_issues.return_value = issues
        mock_cache.get_filtered_issues.return_value = issues

        resp = test_client.post("/api/report/export", json={
            "filters": {},
            "columns": ["key", "summary", "priority", "resolved", "calendar_ttr_hours"],
            "sort_field": "resolved",
            "sort_dir": "desc",
            "group_by": "priority",
            "include_excluded": False,
        })
        assert resp.status_code == 200
        workbook = load_workbook(BytesIO(resp.content), data_only=True)
        sheet = workbook["30 Day"]
        high_row = next(
            row for row in sheet.iter_rows(min_row=14, max_col=8)
            if row[0].value == "High"
        )
        assert high_row[1].value == 3
        assert round(float(high_row[3].value), 1) == 34.3
        assert round(float(high_row[4].value), 1) == 2.0
        assert round(float(high_row[5].value), 1) >= 90.0
        assert round(float(high_row[6].value), 1) >= 98.0

    def test_detail_export_uses_custom_created_range(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 10))

        issues = [
            _make_workload_issue(
                key=f"OIT-CUSTOM-{offset}",
                summary=f"Custom issue {offset}",
                status="Open",
                status_category="To Do",
                assignee="Taylor Ops",
                created=f"{(date(2026, 3, 10) - timedelta(days=offset)).isoformat()}T10:00:00+00:00",
                updated=f"{(date(2026, 3, 10) - timedelta(days=offset)).isoformat()}T10:00:00+00:00",
                oasisdev=False,
            )
            for offset in range(8)
        ]
        mock_cache.get_all_issues.return_value = issues
        mock_cache.get_filtered_issues.return_value = issues

        resp = test_client.post("/api/report/export", json={
            "filters": {
                "created_after": "2026-03-06",
                "created_before": "2026-03-08",
            },
            "columns": ["key", "summary"],
            "sort_field": "resolved",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
            "window_mode": "custom",
        })
        assert resp.status_code == 200

        workbook = load_workbook(BytesIO(resp.content))
        custom_sheet = workbook["Custom Range"]
        assert custom_sheet["B2"].value == "Custom Range"
        assert custom_sheet["B3"].value == "Created"
        assert custom_sheet["B4"].value == "2026-03-06"
        assert custom_sheet["B5"].value == "2026-03-08"
        exported_keys = [cell.value for cell in custom_sheet["A"] if isinstance(cell.value, str) and cell.value.startswith("OIT-CUSTOM-")]
        assert exported_keys == ["OIT-CUSTOM-2", "OIT-CUSTOM-3", "OIT-CUSTOM-4"]

    def test_detail_export_includes_occ_ticket_id_column(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 10))

        issue = _make_workload_issue(
            key="OIT-OCC-1",
            summary="Imported from OCC",
            status="Open",
            status_category="To Do",
            assignee="Taylor Ops",
            created="2026-03-09T10:00:00+00:00",
            updated="2026-03-09T10:00:00+00:00",
            oasisdev=False,
        )
        issue["fields"]["description"] = "OCC Ticket Created By: Libra PhishER | OCC Ticket ID: LIBRA-SR-075203"
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/export", json={
            "filters": {},
            "columns": ["key", "occ_ticket_id"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200

        workbook = load_workbook(BytesIO(resp.content))
        detail_sheet = workbook["30 Day"]
        assert detail_sheet["A13"].value == "Key"
        assert detail_sheet["B13"].value == "OCC Ticket ID"
        assert detail_sheet["A14"].value == "OIT-OCC-1"
        assert detail_sheet["B14"].value == "LIBRA-SR-075203"

    def test_detail_export_includes_first_contact_column(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 10))

        issue = _make_workload_issue(
            key="OIT-FC-1",
            summary="Investigate response breach",
            status="Resolved",
            status_category="Done",
            assignee="Taylor Ops",
            created="2026-03-09T08:00:00+00:00",
            updated="2026-03-09T10:30:00+00:00",
            resolved="2026-03-09T10:30:00+00:00",
            oasisdev=False,
        )
        issue["fields"]["customfield_11266"] = {
            "completedCycles": [
                {
                    "breached": False,
                    "stopTime": {"iso8601": "2026-03-09T09:15:42+00:00"},
                }
            ]
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.post("/api/report/export", json={
            "filters": {},
            "columns": ["key", "created", "first_contact_date"],
            "sort_field": "created",
            "sort_dir": "desc",
            "group_by": None,
            "include_excluded": False,
        })
        assert resp.status_code == 200

        workbook = load_workbook(BytesIO(resp.content))
        detail_sheet = workbook["30 Day"]
        assert detail_sheet["A13"].value == "Key"
        assert detail_sheet["B13"].value == "Created"
        assert detail_sheet["C13"].value == "First Contact"
        assert detail_sheet["A14"].value == "OIT-FC-1"
        assert detail_sheet["B14"].value == "2026-03-09T08:00:00+00:00"
        assert detail_sheet["C14"].value == "2026-03-09T09:15:42+00:00"


class TestReportTemplates:
    def test_list_includes_seeded_primary_templates(self, test_client):
        resp = test_client.get("/api/report/templates")
        assert resp.status_code == 200
        rows = resp.json()
        names = {row["name"] for row in rows}
        assert "First Response Time" in names
        assert "Customer Satisfaction (CSAT)" in names
        followup = next(row for row in rows if row["name"] == "2-Hour Response & Daily Follow-Up")
        assert followup["readiness"] == "proxy"
        assert followup["include_in_master_export"] is False
        assert followup["config"]["group_by"] == "response_followup_status"

    def test_followup_template_becomes_ready_when_authoritative_fields_are_populated(self, test_client, mock_cache, monkeypatch):
        import metrics

        monkeypatch.setattr(metrics, "JIRA_FOLLOWUP_STATUS_FIELD_ID", "customfield_20001")
        monkeypatch.setattr(metrics, "JIRA_FOLLOWUP_LAST_PUBLIC_AGENT_TOUCH_FIELD_ID", "customfield_20002")
        monkeypatch.setattr(metrics, "JIRA_FOLLOWUP_PUBLIC_AGENT_TOUCH_COUNT_FIELD_ID", "customfield_20003")

        issue = {
            "key": "OIT-779",
            "fields": {
                "summary": "Authoritative follow-up coverage",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-23T08:00:00+00:00",
                "updated": "2026-03-23T12:00:00+00:00",
                "resolutiondate": "2026-03-23T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": {"completedCycles": [{"breached": False}]},
                "customfield_11264": None,
                "customfield_20001": {"value": "Met"},
                "customfield_20002": "2026-03-23T10:00:00+00:00",
                "customfield_20003": 1,
                "customfield_10700": [],
                "attachment": [],
                "comment": {"total": 0, "comments": []},
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.get("/api/report/templates")

        assert resp.status_code == 200
        followup = next(row for row in resp.json() if row["name"] == "2-Hour Response & Daily Follow-Up")
        assert followup["readiness"] == "ready"

    def test_followup_template_readiness_uses_current_30_day_window(self, test_client, mock_cache):
        old_issue = {
            "key": "OIT-101",
            "fields": {
                "summary": "Old follow-up history without authoritative cache",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2025-12-20T08:00:00+00:00",
                "updated": "2025-12-21T12:00:00+00:00",
                "resolutiondate": "2025-12-21T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": {"completedCycles": [{"breached": False}]},
                "customfield_11264": None,
                "customfield_10700": [],
                "attachment": [],
                "comment": {"total": 0, "comments": []},
            },
        }
        recent_issue = {
            "key": "OIT-102",
            "fields": {
                "summary": "Recent ticket with local authoritative follow-up cache",
                "status": {"name": "Resolved", "statusCategory": {"name": "Done"}},
                "priority": {"name": "High"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": {"name": "Done"},
                "created": "2026-03-23T08:00:00+00:00",
                "updated": "2026-03-23T12:00:00+00:00",
                "resolutiondate": "2026-03-23T12:00:00+00:00",
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": {"completedCycles": [{"breached": False}]},
                "customfield_11264": None,
                "_movedocs_followup_status": "Met",
                "_movedocs_followup_last_touch_at": "2026-03-23T10:00:00+00:00",
                "_movedocs_followup_touch_count": 1,
                "customfield_10700": [],
                "attachment": [],
                "comment": {"total": 0, "comments": []},
            },
        }
        mock_cache.get_all_issues.return_value = [old_issue, recent_issue]
        mock_cache.get_filtered_issues.return_value = [old_issue, recent_issue]

        resp = test_client.get("/api/report/templates")

        assert resp.status_code == 200
        followup = next(row for row in resp.json() if row["name"] == "2-Hour Response & Daily Follow-Up")
        assert followup["readiness"] == "ready"

    def test_followup_template_becomes_ready_with_public_comment_first_response_fallback(self, test_client, mock_cache):
        issue = {
            "key": "MSD-10117",
            "fields": {
                "summary": "Public agent reply exists but Jira SLA timer is blank",
                "status": {"name": "In Progress", "statusCategory": {"name": "In Progress"}},
                "priority": {"name": "Medium"},
                "assignee": {"displayName": "Alex Agent", "accountId": "acc-alex"},
                "reporter": {"displayName": "Riley Requester", "accountId": "acc-riley"},
                "issuetype": {"name": "[System] Service request"},
                "resolution": None,
                "created": "2026-03-23T08:00:00+00:00",
                "updated": "2026-03-23T09:00:00+00:00",
                "resolutiondate": None,
                "labels": [],
                "components": [],
                "customfield_10010": {"requestType": {"name": "Laptop"}},
                "customfield_11239": "Service requests",
                "customfield_11266": {
                    "id": "2",
                    "name": "Time to first response",
                    "_links": {"self": "https://keyjira.atlassian.net/rest/servicedeskapi/request/123/sla/2"},
                    "completedCycles": [],
                    "slaDisplayFormat": "NEW_SLA_FORMAT",
                },
                "customfield_11264": None,
                "_movedocs_followup_status": "Running",
                "_movedocs_followup_last_touch_at": "2026-03-23T08:45:00+00:00",
                "_movedocs_followup_touch_count": 1,
                "customfield_10700": [],
                "attachment": [],
                "comment": {
                    "total": 1,
                    "comments": [
                        {
                            "created": "2026-03-23T08:45:00+00:00",
                            "updated": "2026-03-23T08:45:00+00:00",
                            "author": {"displayName": "OSIJIRAOCC", "accountId": "acc-occ"},
                            "jsdPublic": True,
                            "body": "Initial public response",
                        },
                    ],
                },
            },
        }
        mock_cache.get_all_issues.return_value = [issue]
        mock_cache.get_filtered_issues.return_value = [issue]

        resp = test_client.get("/api/report/templates")

        assert resp.status_code == 200
        followup = next(row for row in resp.json() if row["name"] == "2-Hour Response & Daily Follow-Up")
        assert followup["readiness"] == "ready"

    def test_create_update_and_delete_custom_template(self, test_client):
        create_resp = test_client.post(
            "/api/report/templates",
            json={
                "name": "Leadership Weekly Snapshot",
                "description": "Weekly leadership report.",
                "category": "Executive",
                "notes": "Use in Monday leadership sync.",
                "include_in_master_export": False,
                "config": {
                    "filters": {"open_only": True},
                    "columns": ["key", "summary", "priority", "status"],
                    "sort_field": "created",
                    "sort_dir": "desc",
                    "group_by": "priority",
                    "include_excluded": False,
                },
            },
        )
        assert create_resp.status_code == 200
        created = create_resp.json()
        assert created["name"] == "Leadership Weekly Snapshot"
        assert created["is_seed"] is False
        assert created["include_in_master_export"] is False
        assert created["created_by_email"] == "test@example.com"

        update_resp = test_client.put(
            f"/api/report/templates/{created['id']}",
            json={
                "name": "Leadership Weekly Snapshot",
                "description": "Updated description.",
                "category": "Executive",
                "notes": "Updated notes.",
                "include_in_master_export": True,
                "config": {
                    "filters": {"open_only": True, "stale_only": True},
                    "columns": ["key", "summary", "priority", "age_days"],
                    "sort_field": "age_days",
                    "sort_dir": "desc",
                    "group_by": "status",
                    "include_excluded": False,
                },
            },
        )
        assert update_resp.status_code == 200
        updated = update_resp.json()
        assert updated["description"] == "Updated description."
        assert updated["include_in_master_export"] is True
        assert updated["config"]["sort_field"] == "age_days"
        assert updated["config"]["filters"]["stale_only"] is True

        delete_resp = test_client.delete(f"/api/report/templates/{created['id']}")
        assert delete_resp.status_code == 200
        assert delete_resp.json() == {"deleted": True}

        list_resp = test_client.get("/api/report/templates")
        assert list_resp.status_code == 200
        assert all(row["id"] != created["id"] for row in list_resp.json())

    def test_create_template_rejects_invalid_custom_window(self, test_client):
        create_resp = test_client.post(
            "/api/report/templates",
            json={
                "name": "Broken Custom Window",
                "description": "",
                "category": "Executive",
                "notes": "",
                "include_in_master_export": False,
                "config": {
                    "filters": {"created_after": "2026-03-10"},
                    "columns": ["key", "summary"],
                    "sort_field": "created",
                    "sort_dir": "desc",
                    "group_by": None,
                    "include_excluded": False,
                    "window_mode": "custom",
                },
            },
        )

        assert create_resp.status_code == 400
        assert "Created Before" in create_resp.json()["detail"]

    def test_seed_templates_can_be_updated_and_deleted(self, test_client):
        list_resp = test_client.get("/api/report/templates")
        assert list_resp.status_code == 200
        seed_template = next(row for row in list_resp.json() if row["is_seed"])
        seed_id = seed_template["id"]

        update_resp = test_client.put(
            f"/api/report/templates/{seed_id}",
            json={
                "name": "Updated Starter Template",
                "description": "Editable now.",
                "category": "Executive",
                "notes": "Updated by test.",
                "config": {
                    "filters": {},
                    "columns": ["key"],
                    "sort_field": "created",
                    "sort_dir": "desc",
                    "group_by": None,
                    "include_excluded": False,
                },
            },
        )
        assert update_resp.status_code == 200
        updated = update_resp.json()
        assert updated["id"] == seed_id
        assert updated["name"] == "Updated Starter Template"
        assert updated["description"] == "Editable now."

        delete_resp = test_client.delete(f"/api/report/templates/{seed_id}")
        assert delete_resp.status_code == 200
        assert delete_resp.json() == {"deleted": True}

        list_after_delete = test_client.get("/api/report/templates")
        assert list_after_delete.status_code == 200
        assert all(row["id"] != seed_id for row in list_after_delete.json())

    def test_deleted_seed_template_is_not_recreated_on_store_restart(self, tmp_path):
        store = ReportTemplateStore(str(tmp_path / "report_templates.db"))
        initial_seed = next(template for template in store.list_templates("primary") if template.is_seed)

        store.delete_template(initial_seed.id, "primary")

        reloaded_store = ReportTemplateStore(str(tmp_path / "report_templates.db"))
        assert all(template.id != initial_seed.id for template in reloaded_store.list_templates("primary"))

    def test_master_workbook_export_includes_index_and_seeded_reports(self, test_client):
        resp = test_client.get("/api/report/templates/master.xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

        workbook = load_workbook(BytesIO(resp.content))
        assert "Report Index" in workbook.sheetnames
        assert "Executive Dashboard" in workbook.sheetnames
        assert "Trends" in workbook.sheetnames
        assert "Data Gaps" in workbook.sheetnames
        assert "First Response Time 30 Day" in workbook.sheetnames
        assert "Backlog Size & Aging 30 Day" in workbook.sheetnames

        index_ws = workbook["Report Index"]
        headers = [index_ws.cell(row=1, column=idx).value for idx in range(1, 14)]
        assert headers == [
            "Status",
            "Report",
            "Sheet",
            "Window",
            "Window Field",
            "Window Start",
            "Window End",
            "Category",
            "Readiness",
            "View",
            "Rows",
            "Description",
            "Notes",
        ]
        report_names = [index_ws.cell(row=row_idx, column=2).value for row_idx in range(2, index_ws.max_row + 1)]
        assert "First Response Time" in report_names
        assert index_ws["A2"].value in {"✅", "⚠️", "🔴"}
        assert index_ws["C2"].hyperlink is not None

        frt_ws = workbook["First Response Time 30 Day"]
        assert frt_ws["A1"].value == "Report"
        assert frt_ws["B1"].value == "First Response Time"
        assert frt_ws["A2"].value == "Window"
        assert frt_ws["B2"].value == "30 Day"
        assert frt_ws["A3"].value == "Window Field"
        assert frt_ws["B3"].value == "Created"
        assert frt_ws["A7"].value == "Readiness"
        assert frt_ws["H13"].value == "Δ Count vs Prior"
        labels = [frt_ws.cell(row=row_idx, column=1).value for row_idx in range(14, frt_ws.max_row + 1)]
        assert "Total" in labels
        assert "First Response SLA Met %" in labels

    def test_master_workbook_build_runs_via_asyncio_to_thread(self, test_client, monkeypatch):
        calls: list[str] = []

        async def fake_to_thread(func, /, *args, **kwargs):
            calls.append(getattr(func, "__name__", ""))
            return func(*args, **kwargs)

        monkeypatch.setattr("routes_export.asyncio.to_thread", fake_to_thread)

        resp = test_client.get("/api/report/templates/master.xlsx")
        assert resp.status_code == 200
        assert "_write_master_report_workbook_file" in calls

    def test_master_workbook_only_includes_templates_marked_for_export(self, test_client):
        templates_resp = test_client.get("/api/report/templates")
        assert templates_resp.status_code == 200
        first_response_template = next(
            template for template in templates_resp.json() if template["name"] == "First Response Time"
        )

        toggle_resp = test_client.post(
            f"/api/report/templates/{first_response_template['id']}/export-selection",
            json={"include_in_master_export": False},
        )
        assert toggle_resp.status_code == 200
        assert toggle_resp.json()["include_in_master_export"] is False

        workbook_resp = test_client.get("/api/report/templates/master.xlsx")
        assert workbook_resp.status_code == 200
        workbook = load_workbook(BytesIO(workbook_resp.content))
        assert "Report Index" in workbook.sheetnames
        assert "Executive Dashboard" in workbook.sheetnames
        assert "First Response Time 30 Day" not in workbook.sheetnames

        index_ws = workbook["Report Index"]
        report_names = [index_ws.cell(row=row_idx, column=2).value for row_idx in range(2, index_ws.max_row + 1)]
        assert "First Response Time" not in report_names

    def test_master_workbook_only_passes_selected_templates_to_summary_lookup_and_writer(self, test_client, monkeypatch):
        import routes_export

        templates_resp = test_client.get("/api/report/templates")
        assert templates_resp.status_code == 200
        first_response_template = next(
            template for template in templates_resp.json() if template["name"] == "First Response Time"
        )

        toggle_resp = test_client.post(
            f"/api/report/templates/{first_response_template['id']}/export-selection",
            json={"include_in_master_export": False},
        )
        assert toggle_resp.status_code == 200

        captured: dict[str, list[str]] = {}

        async def fake_to_thread(func, /, *args, **kwargs):
            return func(*args, **kwargs)

        def fake_get_current_master_summaries(site_scope: str, templates):
            assert site_scope == "primary"
            captured["summary_lookup"] = [template.name for template in templates]
            return {}

        def fake_write_master_report_workbook_file(
            *,
            path: str,
            templates,
            site_scope: str,
            all_issues,
            today,
            ai_template_summaries=None,
        ) -> None:
            del site_scope, all_issues, today, ai_template_summaries
            captured["writer_templates"] = [template.name for template in templates]
            with open(path, "wb") as handle:
                handle.write(b"master-workbook")

        monkeypatch.setattr(routes_export.asyncio, "to_thread", fake_to_thread)
        monkeypatch.setattr(
            routes_export.report_ai_summary_service,
            "get_current_master_summaries",
            fake_get_current_master_summaries,
        )
        monkeypatch.setattr(
            routes_export,
            "_write_master_report_workbook_file",
            fake_write_master_report_workbook_file,
        )

        workbook_resp = test_client.get("/api/report/templates/master.xlsx")

        assert workbook_resp.status_code == 200
        assert "First Response Time" not in captured["summary_lookup"]
        assert captured["summary_lookup"] == captured["writer_templates"]

    def test_seed_template_export_selection_can_be_toggled_without_full_edit(self, test_client):
        templates_resp = test_client.get("/api/report/templates")
        assert templates_resp.status_code == 200
        seed_template = next(template for template in templates_resp.json() if template["is_seed"])

        toggle_resp = test_client.post(
            f"/api/report/templates/{seed_template['id']}/export-selection",
            json={"include_in_master_export": False},
        )
        assert toggle_resp.status_code == 200
        payload = toggle_resp.json()
        assert payload["id"] == seed_template["id"]
        assert payload["is_seed"] is True
        assert payload["include_in_master_export"] is False

    def test_template_insights_return_window_aware_metrics(self, test_client, mock_cache, monkeypatch):
        monkeypatch.setattr("routes_export._today_utc", lambda: date(2026, 3, 4))

        issues = []
        for offset in range(10):
            created_day = date(2026, 3, 4) - timedelta(days=offset)
            created_iso = f"{created_day.isoformat()}T10:00:00+00:00"
            for count_idx in range(2):
                issues.append(
                    _make_workload_issue(
                        key=f"OIT-OPEN-{offset}-{count_idx}",
                        summary=f"Open report issue {offset}-{count_idx}",
                        status="Open",
                        status_category="To Do",
                        assignee="Taylor Ops",
                        created=created_iso,
                        updated=created_iso,
                        oasisdev=False,
                    )
                )

        for offset in range(5):
            resolved_day = date(2026, 3, 4) - timedelta(days=offset)
            issues.append(
                _make_workload_issue(
                    key=f"OIT-RES-{offset}",
                    summary=f"Resolved issue {offset}",
                    status="Resolved",
                    status_category="Done",
                    assignee="Taylor Ops",
                    created="2026-01-10T10:00:00+00:00",
                    updated=f"{resolved_day.isoformat()}T13:00:00+00:00",
                    resolved=f"{resolved_day.isoformat()}T13:00:00+00:00",
                    oasisdev=False,
                )
            )

        mock_cache.get_all_issues.return_value = issues

        resp = test_client.get("/api/report/templates/insights")
        assert resp.status_code == 200
        data = resp.json()

        frt = next(item for item in data if item["template_name"] == "First Response Time")
        assert frt["window_mode"] == "30d"
        assert frt["window_label"] == "30 Day"
        assert frt["window_field"] == "created"
        assert frt["window_field_label"] == "Created"
        assert frt["window_start"] == "2026-02-03"
        assert frt["window_end"] == "2026-03-04"
        assert frt["count_in_window"] == 20
        assert frt["p95_daily_count"] == 2.0
        assert len(frt["trend"]) == 30

        mttr = next(item for item in data if item["template_name"] == "Mean Time to Resolution")
        assert mttr["window_mode"] == "30d"
        assert mttr["window_label"] == "30 Day"
        assert mttr["window_field"] == "resolved"
        assert mttr["window_field_label"] == "Resolved"
        assert mttr["count_in_window"] == 5

    def test_master_workbook_contains_dashboard_and_detail_charts(self, test_client):
        resp = test_client.get("/api/report/templates/master.xlsx")
        assert resp.status_code == 200

        workbook_zip = zipfile.ZipFile(BytesIO(resp.content))
        chart_parts = [name for name in workbook_zip.namelist() if name.startswith("xl/charts/chart")]
        drawing_parts = [name for name in workbook_zip.namelist() if name.startswith("xl/drawings/drawing")]
        assert len(chart_parts) >= 5
        assert drawing_parts


class TestReportAISummaries:
    def test_list_current_report_ai_summaries_returns_primary_summaries(self, test_client):
        import routes_export

        routes_export.report_ai_summary_service.list_current_summaries.return_value = [
            {
                "template_id": "tpl-1",
                "template_name": "Executive",
                "site_scope": "primary",
                "source": "manual",
                "status": "ready",
                "summary": "Summary text",
                "bullets": ["Bullet one"],
                "fallback_used": False,
                "model_used": "nemotron-3-nano:4b",
                "generated_at": "2026-03-26T00:00:00+00:00",
                "template_version": "2026-03-26T00:00:00+00:00",
                "data_version": "2026-03-26T00:00:00+00:00",
                "error": "",
            },
        ]

        resp = test_client.get("/api/report/templates/ai-summaries")

        assert resp.status_code == 200
        assert resp.json()[0]["template_id"] == "tpl-1"

    def test_generate_report_ai_summaries_queues_manual_batch(self, test_client):
        import routes_export

        routes_export.report_ai_summary_service.start_manual_batch.return_value = {
            "batch_id": "batch-2",
            "site_scope": "primary",
            "status": "queued",
            "item_count": 3,
            "requested_at": "2026-03-26T00:00:00+00:00",
        }

        resp = test_client.post("/api/report/templates/ai-summaries/generate")

        assert resp.status_code == 200
        assert resp.json()["batch_id"] == "batch-2"

    def test_get_report_ai_summary_batch_status_returns_items(self, test_client):
        import routes_export

        routes_export.report_ai_summary_service.get_batch_status.return_value = {
            "batch_id": "batch-3",
            "site_scope": "primary",
            "status": "running",
            "item_count": 1,
            "requested_at": "2026-03-26T00:00:00+00:00",
            "started_at": "2026-03-26T00:01:00+00:00",
            "completed_at": None,
            "items": [
                {
                    "template_id": "tpl-1",
                    "template_name": "Executive",
                    "status": "running",
                    "source": "manual",
                    "summary": "",
                    "bullets": [],
                    "fallback_used": False,
                    "model_used": "",
                    "generated_at": None,
                    "error": "",
                },
            ],
        }

        resp = test_client.get("/api/report/templates/ai-summaries/batches/batch-3")

        assert resp.status_code == 200
        assert resp.json()["items"][0]["status"] == "running"

    def test_generate_report_ai_summaries_hidden_on_non_primary_host(self, test_client):
        resp = test_client.post(
            "/api/report/templates/ai-summaries/generate",
            headers={"host": "oasisdev.movedocs.com"},
        )

        assert resp.status_code == 404


class TestLegacyExport:
    """GET /api/export/excel"""

    def test_returns_excel(self, test_client):
        resp = test_client.get("/api/export/excel")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    def test_response_is_binary(self, test_client):
        resp = test_client.get("/api/export/excel")
        assert len(resp.content) > 0


class TestOasisDevWorkloadReport:
    def test_preview_returns_monthly_summary_for_oasisdev_host(self, test_client, mock_cache):
        mock_cache.get_all_issues.return_value = [
            _make_workload_issue(
                key="OIT-501",
                summary="oasisdev app issue",
                status="Acknowledged",
                status_category="In Progress",
                assignee="Tim Reckamp",
                created="2026-01-05T10:00:00+00:00",
                updated="2026-01-06T10:00:00+00:00",
                components=["Portal"],
            ),
            _make_workload_issue(
                key="OIT-502",
                summary="oasisdev resolved issue",
                status="Resolved",
                status_category="Done",
                assignee="Tim Reckamp",
                created="2026-02-14T10:00:00+00:00",
                updated="2026-02-20T10:00:00+00:00",
                resolved="2026-02-20T10:00:00+00:00",
                components=["VPN"],
            ),
            _make_workload_issue(
                key="OIT-503",
                summary="oasisdev march follow-up",
                status="Resolved",
                status_category="Done",
                assignee="Tim Reckamp",
                created="2026-03-02T10:00:00+00:00",
                updated="2026-03-03T10:00:00+00:00",
                resolved="2026-03-03T10:00:00+00:00",
                components=["Outlook"],
            ),
            _make_workload_issue(
                key="OIT-504",
                summary="oasisdev other technician",
                status="Open",
                status_category="To Do",
                assignee="Other Tech",
                created="2026-03-07T10:00:00+00:00",
                updated="2026-03-07T10:00:00+00:00",
            ),
            _make_workload_issue(
                key="OIT-999",
                summary="primary site ticket",
                status="Open",
                status_category="To Do",
                assignee="Tim Reckamp",
                created="2026-03-04T10:00:00+00:00",
                updated="2026-03-04T10:00:00+00:00",
                oasisdev=False,
            ),
        ]

        resp = test_client.post(
            "/api/report/oasisdev-workload",
            headers={"host": "oasisdev.movedocs.com"},
            json={
                "assignee": "Tim Reckamp",
                "report_start": "2026-01-01",
                "report_end": "2026-03-31",
                "last_report_date": "2026-03-01",
            },
        )
        assert resp.status_code == 200
        payload = resp.json()
        assert payload["summary"]["assignee"] == "Tim Reckamp"
        assert payload["summary"]["tickets_created_in_window"] == 3
        assert payload["summary"]["tickets_resolved_in_window"] == 2
        assert payload["monthly_status"]["grand_total"] == [1, 1, 1]
        assert payload["since_last_report"]["created_count"] == 1
        assert payload["since_last_report"]["resolved_count"] == 1
        assert payload["since_last_report"]["tickets"][0]["application"] == "Outlook"

    def test_preview_is_hidden_on_non_oasisdev_host(self, test_client):
        resp = test_client.post("/api/report/oasisdev-workload", json={})
        assert resp.status_code == 404

    def test_export_returns_excel_workbook(self, test_client, mock_cache):
        mock_cache.get_all_issues.return_value = [
            _make_workload_issue(
                key="OIT-601",
                summary="oasisdev export row",
                status="Resolved",
                status_category="Done",
                assignee="Tim Reckamp",
                created="2026-03-08T10:00:00+00:00",
                updated="2026-03-10T10:00:00+00:00",
                resolved="2026-03-10T10:00:00+00:00",
                components=["Portal", "VPN"],
            ),
        ]

        resp = test_client.post(
            "/api/report/oasisdev-workload/export",
            headers={"host": "oasisdev.movedocs.com"},
            json={
                "assignee": "Tim Reckamp",
                "report_start": "2026-01-01",
                "report_end": "2026-03-31",
                "last_report_date": "2026-03-01",
            },
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")
        workbook = load_workbook(filename=BytesIO(resp.content))
        assert workbook.sheetnames == [
            "Monthly Status",
            "Created vs Resolved",
            "Since Last Report",
        ]
        since_last_report = workbook["Since Last Report"]
        assert since_last_report["B3"].value == 1
        assert since_last_report["J13"].value == "Portal, VPN"
