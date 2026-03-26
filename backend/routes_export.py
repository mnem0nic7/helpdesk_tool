"""API routes for Excel report export and report builder."""

from __future__ import annotations

import asyncio
import logging
import os
import re
import statistics
import tempfile
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from issue_cache import cache
from jira_client import JiraClient
from metrics import issue_to_row, _extract_description, _all_comments_text
from metrics import percentile
from auth import require_authenticated_user
from models import (
    OasisDevWorkloadReportRequest,
    ReportConfig,
    ReportTemplate,
    ReportTemplateExportSelectionRequest,
    ReportTemplateInsight,
    ReportTemplateCreateRequest,
    ReportTemplateUpdateRequest,
)
from report_template_store import report_template_store
from report_workbook_builder import (
    ReportWorkbookBuilder,
    report_window_mode,
    resolve_report_window_spec,
)
from routes_tickets import _match
from site_context import get_current_site_scope, get_scoped_issues, get_site_profile

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")


def _calculation_scoped_issues(*, include_excluded_on_primary: bool = False) -> list[dict[str, Any]]:
    """Return issues eligible for report/chart calculations on the active host.

    On the primary host, OasisDev tickets stay out of calculations even when a
    caller asks to include excluded tickets.
    """
    scope = get_current_site_scope()
    if scope == "primary":
        return get_scoped_issues(include_excluded_on_primary=False)
    return get_scoped_issues(include_excluded_on_primary=include_excluded_on_primary)

def _write_single_report_workbook_file(
    *,
    path: str,
    config: ReportConfig,
    report_name: str,
    report_description: str,
    template: ReportTemplate | None,
    site_scope: str,
    all_issues: list[dict[str, Any]],
    today: date,
) -> None:
    builder = ReportWorkbookBuilder(
        all_issues=all_issues,
        site_scope=site_scope,
        today=today,
    )
    builder.build_single_report(
        path=path,
        config=config,
        report_name=report_name,
        report_description=report_description,
        template=template,
    )


def _write_master_report_workbook_file(
    *,
    path: str,
    templates: list[ReportTemplate],
    site_scope: str,
    all_issues: list[dict[str, Any]],
    today: date,
) -> None:
    builder = ReportWorkbookBuilder(
        all_issues=all_issues,
        site_scope=site_scope,
        today=today,
    )
    builder.build_master_report(path=path, templates=templates)


def _apply_runtime_template_readiness(
    templates: list[ReportTemplate],
    *,
    site_scope: str,
) -> list[ReportTemplate]:
    if not templates:
        return []
    builder = ReportWorkbookBuilder(
        all_issues=_calculation_scoped_issues(include_excluded_on_primary=True),
        site_scope=site_scope,
        today=_today_utc(),
        enable_changelog_fetch=False,
    )
    adjusted: list[ReportTemplate] = []
    for template in templates:
        try:
            runtime_readiness = builder.runtime_template_readiness(template)
        except ValueError:
            runtime_readiness = "gap"
        adjusted.append(template.model_copy(update={"readiness": runtime_readiness}))
    return adjusted


# ---------------------------------------------------------------------------
# Column metadata — maps TicketRow field keys to human-readable labels
# ---------------------------------------------------------------------------

FIELD_META: dict[str, dict[str, str]] = {
    "key": {"label": "Key", "description": "Jira issue key"},
    "summary": {"label": "Summary", "description": "Issue title"},
    "description": {"label": "Description", "description": "Issue description text"},
    "issue_type": {"label": "Type", "description": "Issue type"},
    "status": {"label": "Status", "description": "Current status"},
    "status_category": {"label": "Status Category", "description": "Status category (To Do / In Progress / Done)"},
    "priority": {"label": "Priority", "description": "Priority level"},
    "resolution": {"label": "Resolution", "description": "Resolution type"},
    "assignee": {"label": "Assignee", "description": "Assigned team member"},
    "assignee_account_id": {"label": "Assignee ID", "description": "Assignee Atlassian account ID"},
    "reporter": {"label": "Reporter", "description": "Person who created the ticket"},
    "created": {"label": "Created", "description": "Date ticket was created"},
    "updated": {"label": "Updated", "description": "Date ticket was last updated"},
    "resolved": {"label": "Resolved", "description": "Date ticket was resolved"},
    "request_type": {"label": "Request Type", "description": "JSM request type"},
    "work_category": {"label": "Work Category", "description": "Work category classification"},
    "calendar_ttr_hours": {"label": "TTR (h)", "description": "Calendar time-to-resolution in hours"},
    "age_days": {"label": "Age (d)", "description": "Age of open tickets in days"},
    "days_since_update": {"label": "Days Since Update", "description": "Days since last update"},
    "comment_count": {"label": "Comments", "description": "Number of comments"},
    "last_comment_date": {"label": "Last Comment", "description": "Date of last comment"},
    "last_comment_author": {"label": "Last Commenter", "description": "Author of last comment"},
    "excluded": {"label": "Excluded", "description": "Excluded from metrics"},
    "sla_first_response_status": {"label": "SLA Response", "description": "First-response SLA status"},
    "sla_resolution_status": {"label": "SLA Resolution", "description": "Resolution SLA status"},
    "response_followup_status": {"label": "Response + Follow-Up", "description": "Overall 2-hour response and daily follow-up compliance status"},
    "first_response_2h_status": {"label": "Response <=2h", "description": "Whether the first non-requester response landed within 2 hours"},
    "daily_followup_status": {"label": "Daily Follow-Up", "description": "Whether the ticket received at least one non-requester follow-up every 24 hours until resolution"},
    "last_support_touch_date": {"label": "Last Public Agent Touch", "description": "Most recent public comment from a configured OIT Jira agent"},
    "support_touch_count": {"label": "Public Agent Touch Count", "description": "Number of public comments from configured OIT Jira agents on the ticket"},
    "labels": {"label": "Labels", "description": "Issue labels"},
    "components": {"label": "Components", "description": "Issue components"},
    "organizations": {"label": "Organizations", "description": "Customer organizations"},
    "attachment_count": {"label": "Attachments", "description": "Number of attachments"},
    "comments_text": {"label": "All Comments", "description": "Full text of all comments"},
}

# All TicketRow field keys in display order
ALL_FIELDS: list[str] = list(FIELD_META.keys())

# Default columns shown when no selection is provided
DEFAULT_COLUMNS: list[str] = [
    "key", "summary", "issue_type", "status", "priority",
    "assignee", "created", "resolved", "calendar_ttr_hours",
]

_REPORT_WINDOW_LABELS: dict[str, str] = {
    "created": "Created",
    "updated": "Updated",
    "resolved": "Resolved",
}

_DETAIL_WIDTH_DEFAULTS: dict[str, int] = {
    "key": 12,
    "summary": 50,
    "description": 60,
    "issue_type": 18,
    "status": 22,
    "status_category": 18,
    "priority": 12,
    "resolution": 18,
    "assignee": 25,
    "assignee_account_id": 20,
    "reporter": 25,
    "created": 22,
    "updated": 22,
    "resolved": 22,
    "request_type": 25,
    "work_category": 20,
    "calendar_ttr_hours": 12,
    "age_days": 10,
    "days_since_update": 14,
    "comment_count": 10,
    "last_comment_date": 22,
    "last_comment_author": 25,
    "excluded": 10,
    "sla_first_response_status": 15,
    "sla_resolution_status": 15,
    "response_followup_status": 22,
    "first_response_2h_status": 16,
    "daily_followup_status": 18,
    "last_support_touch_date": 22,
    "support_touch_count": 12,
    "labels": 30,
    "components": 25,
    "organizations": 30,
    "attachment_count": 12,
    "comments_text": 60,
}

# ---------------------------------------------------------------------------
# Legacy column definitions for the old /export/excel endpoint
# ---------------------------------------------------------------------------

_COLUMNS: list[tuple[str, str]] = [
    ("Key", "key"),
    ("Summary", "summary"),
    ("Type", "issue_type"),
    ("Status", "status"),
    ("Priority", "priority"),
    ("Assignee", "assignee"),
    ("Reporter", "reporter"),
    ("Created", "created"),
    ("Updated", "updated"),
    ("Resolved", "resolved"),
    ("TTR (h)", "calendar_ttr_hours"),
    ("Age (d)", "age_days"),
    ("SLA Response", "sla_first_response_status"),
    ("SLA Resolution", "sla_resolution_status"),
    ("Response + Follow-Up", "response_followup_status"),
    ("Response <=2h", "first_response_2h_status"),
    ("Daily Follow-Up", "daily_followup_status"),
    ("Last Public Agent Touch", "last_support_touch_date"),
    ("Public Agent Touch Count", "support_touch_count"),
    ("Excluded", "excluded"),
]

# Header styling
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Report builder helpers
# ---------------------------------------------------------------------------


def _issues_matching_config(
    config: ReportConfig,
    *,
    issues: list[dict[str, Any]] | None = None,
    report_name: str = "",
    today: date | None = None,
) -> list[dict[str, Any]]:
    """Return scoped Jira issues that match the report config filters."""
    source_issues = issues if issues is not None else _calculation_scoped_issues(include_excluded_on_primary=config.include_excluded)
    source_issues = [iss for iss in source_issues if JiraClient.is_tracked_issue(iss)]
    filters = config.filters.model_dump(exclude_none=True)
    if report_window_mode(config) != "custom":
        filters.pop("created_after", None)
        filters.pop("created_before", None)
    for k in ("open_only", "stale_only"):
        if not filters.get(k):
            filters.pop(k, None)
    matched = [iss for iss in source_issues if _match(iss, **filters)]
    window_field = _date_field_for_report_window_config(config, report_name=report_name)
    window_spec = resolve_report_window_spec(config, today=today or _today_utc())
    return [
        issue
        for issue in matched
        if (issue_day := _issue_date_for_window(issue, window_field)) and window_spec.start <= issue_day <= window_spec.end
    ]


def _rows_from_issues(config: ReportConfig, issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Convert matching Jira issues to sorted flat report rows."""
    matched = issues

    requested_fields = set(config.columns or [])
    if config.sort_field:
        requested_fields.add(config.sort_field)
    if config.group_by:
        requested_fields.add(config.group_by)

    include_comment_meta = bool(
        requested_fields.intersection({"comment_count", "last_comment_date", "last_comment_author"})
    )
    include_description = "description" in requested_fields

    # Convert to flat rows
    rows = [
        issue_to_row(
            iss,
            include_comment_meta=include_comment_meta,
            include_description=include_description,
        )
        for iss in matched
    ]

    # Sort
    sort_field = config.sort_field or "created"
    reverse = config.sort_dir == "desc"

    def sort_key(row: dict[str, Any]) -> Any:
        val = row.get(sort_field)
        if isinstance(val, list):
            return ", ".join(str(part) for part in val)
        if val is None:
            return "" if isinstance(row.get(sort_field, ""), str) else -1
        return val

    rows.sort(key=sort_key, reverse=reverse)

    return rows


def _apply_config(config: ReportConfig) -> list[dict[str, Any]]:
    """Filter, convert to flat rows, and sort according to the config."""
    matched_issues = _issues_matching_config(config)
    return _rows_from_issues(config, matched_issues)


def _select_columns(row: dict[str, Any], columns: list[str]) -> dict[str, Any]:
    """Pick only the selected keys from a row dict."""
    return {k: row.get(k) for k in columns}


def _aggregate(rows: list[dict[str, Any]], group_by: str) -> list[dict[str, Any]]:
    """Group flat rows by a field and return summary rows."""
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        val = row.get(group_by)
        if isinstance(val, list):
            key = ", ".join(val) if val else "(none)"
        elif isinstance(val, bool):
            key = "Yes" if val else "No"
        else:
            key = str(val) if val else "(none)"
        groups[key].append(row)

    summary: list[dict[str, Any]] = []
    for group_val, group_rows in sorted(groups.items()):
        open_count = sum(
            1 for r in group_rows
            if r.get("status_category", "") != "Done"
        )
        ttrs = [
            r["calendar_ttr_hours"] for r in group_rows
            if r.get("calendar_ttr_hours") is not None
        ]
        avg_ttr = round(statistics.mean(ttrs), 1) if ttrs else None
        summary.append({
            "group": group_val,
            "count": len(group_rows),
            "open": open_count,
            "avg_ttr_hours": avg_ttr,
        })

    return summary


def _sanitize_for_excel(val: str) -> str:
    """Prevent Excel formula injection by prefixing dangerous strings with a tab."""
    if val and val[0] in ("=", "+", "-", "@"):
        return "\t" + val
    return val


def _cell_value(value: Any) -> Any:
    """Convert a value for Excel output, with formula injection protection."""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, list):
        return _sanitize_for_excel(", ".join(str(v) for v in value))
    if value is None:
        return ""
    if isinstance(value, str):
        return _sanitize_for_excel(value)
    return value


def _sanitize_sheet_name(name: str) -> str:
    """Return an Excel-safe worksheet title."""
    cleaned = re.sub(r"[\[\]\*:/\\?]", " ", str(name or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:31] or "Report"


def _sanitize_report_filename(name: str) -> str:
    """Return a filesystem-safe report filename segment."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name or "").strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._")
    return cleaned[:80] or "Report"


def _unique_sheet_name(name: str, used_names: set[str]) -> str:
    """Generate a unique Excel-safe worksheet title."""
    base = _sanitize_sheet_name(name)
    candidate = base
    suffix = 2
    while candidate in used_names:
        suffix_text = f" ({suffix})"
        candidate = f"{base[: max(1, 31 - len(suffix_text))]}{suffix_text}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _write_detail_sheet(ws, *, title: str, columns: list[str], rows: list[dict[str, Any]], metadata: list[tuple[str, Any]] | None = None) -> None:
    """Render a flat detail report worksheet."""
    ws.title = title
    start_row = 1
    if metadata:
        for row_idx, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=_cell_value(value))
        start_row = len(metadata) + 2
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 18)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 48)

    headers = [FIELD_META.get(c, {}).get("label", c) for c in columns]
    _write_header_row(ws, start_row, headers)
    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, col_key in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(row.get(col_key)))

    if columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{len(rows) + start_row}"
    ws.freeze_panes = f"A{start_row + 1}"
    for col_idx, col_key in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _DETAIL_WIDTH_DEFAULTS.get(col_key, 15)


def _write_grouped_sheet(
    ws,
    *,
    title: str,
    group_by: str,
    rows: list[dict[str, Any]],
    metadata: list[tuple[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Render an aggregated report worksheet and return the grouped rows."""
    ws.title = title
    start_row = 1
    if metadata:
        for row_idx, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=_cell_value(value))
        start_row = len(metadata) + 2
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 18)
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 48)

    grouped_rows = _aggregate(rows, group_by)
    headers = [
        FIELD_META.get(group_by, {}).get("label", group_by),
        "Count",
        "Open",
        "Avg TTR (h)",
    ]
    _write_header_row(ws, start_row, headers)
    for row_idx, agg_row in enumerate(grouped_rows, start_row + 1):
        ws.cell(row=row_idx, column=1, value=_cell_value(agg_row["group"]))
        ws.cell(row=row_idx, column=2, value=agg_row["count"])
        ws.cell(row=row_idx, column=3, value=agg_row["open"])
        ws.cell(row=row_idx, column=4, value=agg_row["avg_ttr_hours"] or "")

    ws.auto_filter.ref = f"A{start_row}:D{len(grouped_rows) + start_row}"
    ws.freeze_panes = f"A{start_row + 1}"
    _autofit_columns(ws, {1: 30, 2: 12, 3: 12, 4: 14})
    return grouped_rows


def _build_master_report_workbook(
    templates: list[ReportTemplate],
    *,
    today: date | None = None,
) -> Workbook:
    """Build a single workbook containing every saved report template for the current site."""
    now = datetime.now(timezone.utc)
    wb = Workbook()
    index_ws = wb.active
    index_ws.title = "Report Index"
    index_headers = [
        "Report",
        "Sheet",
        "Window",
        "Window Field",
        "Window Start",
        "Window End",
        "Category",
        "Readiness",
        "View",
        "Rows",
        "Description",
        "Notes",
    ]
    _write_header_row(index_ws, 1, index_headers)
    used_names = {"Report Index"}
    index_row = 2

    included_templates = [template for template in templates if template.include_in_master_export]

    if not included_templates:
        index_ws.cell(row=2, column=1, value="No report templates are currently included in the master export.")
        _autofit_columns(index_ws, {1: 44})
        return wb

    for template in included_templates:
        config = template.config if isinstance(template.config, ReportConfig) else ReportConfig()
        view_type = _report_view_type(config)
        window_field = _date_field_for_report_window(template)
        window_field_label = _REPORT_WINDOW_LABELS.get(window_field, "Created")
        base_issues = _calculation_scoped_issues(include_excluded_on_primary=config.include_excluded)
        window_spec = resolve_report_window_spec(config, today=today or _today_utc())
        rows, window_start, window_end = _windowed_rows_for_report(
            config,
            window_field=window_field,
            report_name=template.name,
            issues=base_issues,
            today=today,
        )
        sheet_name = _unique_sheet_name(f"{template.name} {window_spec.label}", used_names)
        metadata = [
            ("Report", template.name),
            ("Window", window_spec.label),
            ("Window Field", window_field_label),
            ("Window Start", window_start.isoformat()),
            ("Window End", window_end.isoformat()),
            ("Category", template.category or "Uncategorized"),
            ("Readiness", template.readiness or "custom"),
            ("View", view_type),
            ("Description", template.description or ""),
            ("Notes", template.notes or ""),
            ("Generated", now.isoformat()),
        ]

        sheet = wb.create_sheet(sheet_name)
        if config.group_by:
            rendered_rows = _write_grouped_sheet(
                sheet,
                title=sheet_name,
                group_by=config.group_by,
                rows=rows,
                metadata=metadata,
            )
            exported_row_count = len(rendered_rows)
        else:
            columns = config.columns or DEFAULT_COLUMNS
            _write_detail_sheet(
                sheet,
                title=sheet_name,
                columns=columns,
                rows=rows,
                metadata=metadata,
            )
            exported_row_count = len(rows)

        index_ws.cell(row=index_row, column=1, value=template.name)
        index_ws.cell(row=index_row, column=2, value=sheet_name)
        index_ws.cell(row=index_row, column=3, value=window_spec.label)
        index_ws.cell(row=index_row, column=4, value=window_field_label)
        index_ws.cell(row=index_row, column=5, value=window_start.isoformat())
        index_ws.cell(row=index_row, column=6, value=window_end.isoformat())
        index_ws.cell(row=index_row, column=7, value=template.category or "")
        index_ws.cell(row=index_row, column=8, value=template.readiness or "custom")
        index_ws.cell(row=index_row, column=9, value=view_type)
        index_ws.cell(row=index_row, column=10, value=exported_row_count)
        index_ws.cell(row=index_row, column=11, value=template.description or "")
        index_ws.cell(row=index_row, column=12, value=template.notes or "")
        index_row += 1

    index_ws.freeze_panes = "A2"
    index_ws.auto_filter.ref = f"A1:L{index_row - 1}"
    _autofit_columns(
        index_ws,
        {
            1: 30,
            2: 26,
            3: 12,
            4: 14,
            5: 14,
            6: 14,
            7: 18,
            8: 12,
            9: 12,
            10: 10,
            11: 48,
            12: 60,
        },
    )
    return wb


def _ensure_oasisdev_site() -> None:
    """Restrict OasisDev-only reports to the OasisDev host."""
    if get_current_site_scope() != "oasisdev":
        raise HTTPException(status_code=404, detail="Report not available on this site.")


def _parse_iso_date(raw: str | None, *, field_name: str, default: date) -> date:
    """Parse an ISO date or return the provided default."""
    if not raw:
        return default
    try:
        return date.fromisoformat(raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}: {raw}") from exc


def _today_utc() -> date:
    """Return today's UTC date."""
    return datetime.now(timezone.utc).date()


def _parse_issue_date(issue: dict[str, Any], field_name: str) -> date | None:
    """Return the Jira field as a date, ignoring malformed values."""
    raw = ((issue.get("fields") or {}).get(field_name) or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _date_field_for_report_window_config(config: ReportConfig, *, report_name: str = "") -> str:
    """Choose the most useful date field for rolling report windows."""
    if report_window_mode(config) == "custom":
        return "created"
    sort_field = str(config.sort_field or "").strip().lower()
    if sort_field in _REPORT_WINDOW_LABELS:
        return sort_field

    name = str(report_name or "").strip().lower()
    if any(token in name for token in ("resolution", "mttr", "reopen", "csat")):
        return "resolved"
    if any(token in name for token in ("escalation", "utilization")):
        return "updated"
    return "created"


def _date_field_for_report_window(template: ReportTemplate) -> str:
    """Choose the most useful date field for rolling report windows."""
    return _date_field_for_report_window_config(template.config, report_name=template.name)


def _issue_date_for_window(issue: dict[str, Any], window_field: str) -> date | None:
    """Return the issue date used for rolling report windows."""
    jira_field = {
        "created": "created",
        "updated": "updated",
        "resolved": "resolutiondate",
    }.get(window_field, "created")
    return _parse_issue_date(issue, jira_field)


def _window_bounds(window_days: int, *, today: date | None = None) -> tuple[date, date]:
    """Return inclusive UTC date bounds for a rolling report window."""
    effective_today = today or _today_utc()
    return effective_today - timedelta(days=max(window_days - 1, 0)), effective_today


def _filter_issues_for_report_window(
    issues: list[dict[str, Any]],
    *,
    window_field: str,
    window_days: int,
    today: date | None = None,
) -> tuple[list[dict[str, Any]], date, date]:
    """Filter matching issues to the requested rolling report window."""
    window_start, window_end = _window_bounds(window_days, today=today)
    filtered = []
    for issue in issues:
        issue_day = _issue_date_for_window(issue, window_field)
        if not issue_day or issue_day < window_start or issue_day > window_end:
            continue
        filtered.append(issue)
    return filtered, window_start, window_end


def _report_view_type(config: ReportConfig) -> str:
    """Return the human-readable report view mode."""
    return "Grouped" if config.group_by else "Detail"


def _validate_report_window(config: ReportConfig, *, template_name: str | None = None) -> None:
    try:
        resolve_report_window_spec(config, today=_today_utc())
    except ValueError as exc:
        prefix = f'Template "{template_name}": ' if template_name else ""
        raise HTTPException(status_code=400, detail=f"{prefix}{exc}") from exc


def _windowed_rows_for_report(
    config: ReportConfig,
    *,
    window_field: str,
    report_name: str = "",
    issues: list[dict[str, Any]] | None = None,
    today: date | None = None,
) -> tuple[list[dict[str, Any]], date, date]:
    """Return sorted report rows for the configured report window."""
    window_spec = resolve_report_window_spec(config, today=today or _today_utc())
    matched_issues = _issues_matching_config(config, issues=issues, report_name=report_name, today=today)
    return _rows_from_issues(config, matched_issues), window_spec.start, window_spec.end


def _report_filters_dict(config: ReportConfig) -> dict[str, Any]:
    """Normalize report filters for matching helpers."""
    filters = config.filters.model_dump(exclude_none=True)
    if report_window_mode(config) != "custom":
        filters.pop("created_after", None)
        filters.pop("created_before", None)
    for key in ("open_only", "stale_only"):
        if not filters.get(key):
            filters.pop(key, None)
    return filters


def _build_template_insight(template: ReportTemplate, issues: list[dict[str, Any]], *, today: date | None = None) -> ReportTemplateInsight:
    """Compute a window-aware summary for a saved template."""
    effective_today = today or _today_utc()
    window_spec = resolve_report_window_spec(template.config, today=effective_today)
    window_field = _date_field_for_report_window(template)
    days = [window_spec.start + timedelta(days=offset) for offset in range(window_spec.days)]
    daily_counts = {day.isoformat(): 0 for day in days}
    filters = _report_filters_dict(template.config)

    count_in_window = 0
    for issue in issues:
        if not _match(issue, **filters):
            continue
        issue_day = _issue_date_for_window(issue, window_field)
        if not issue_day or issue_day < window_spec.start or issue_day > window_spec.end:
            continue
        issue_key = issue_day.isoformat()
        daily_counts[issue_key] += 1
        count_in_window += 1

    ordered_counts = [daily_counts[day.isoformat()] for day in days]
    trend = [
        {"date": day.isoformat(), "count": daily_counts[day.isoformat()]}
        for day in days
    ]
    return ReportTemplateInsight(
        template_id=template.id,
        template_name=template.name,
        window_mode=report_window_mode(template.config),
        window_label=window_spec.label,
        window_field=window_field,
        window_field_label=_REPORT_WINDOW_LABELS.get(window_field, "Created"),
        window_start=window_spec.start.isoformat(),
        window_end=window_spec.end.isoformat(),
        count_in_window=count_in_window,
        p95_daily_count=round(float(percentile([float(value) for value in ordered_counts], 95) or 0.0), 1),
        trend=trend,
    )


def _month_key(value: date) -> str:
    """Return a stable YYYY-MM bucket key."""
    return value.strftime("%Y-%m")


def _month_label(month_key: str) -> str:
    """Return a human-readable month label."""
    return datetime.strptime(month_key, "%Y-%m").strftime("%b %Y")


def _iter_month_keys(start: date, end: date) -> list[str]:
    """Return inclusive month buckets for a date window."""
    month_keys: list[str] = []
    cursor = date(start.year, start.month, 1)
    last = date(end.year, end.month, 1)
    while cursor <= last:
        month_keys.append(_month_key(cursor))
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)
    return month_keys


def _autofit_columns(ws, widths: dict[int, int]) -> None:
    """Apply simple explicit widths to a worksheet."""
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_header_row(ws, row_idx: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT


def _build_oasisdev_workload_report(
    request: OasisDevWorkloadReportRequest,
) -> dict[str, Any]:
    """Build the OasisDev workload summary and export payload."""
    _ensure_oasisdev_site()

    today = datetime.now(timezone.utc).date()
    default_start = date(today.year, 1, 1)
    report_start = _parse_iso_date(request.report_start, field_name="report_start", default=default_start)
    report_end = _parse_iso_date(request.report_end, field_name="report_end", default=today)
    last_report_date = _parse_iso_date(
        request.last_report_date,
        field_name="last_report_date",
        default=report_start,
    )
    if report_start > report_end:
        raise HTTPException(status_code=400, detail="report_start must be on or before report_end.")
    if last_report_date > report_end:
        raise HTTPException(status_code=400, detail="last_report_date must be on or before report_end.")

    issues = get_scoped_issues()
    filtered_records: list[dict[str, Any]] = []
    for issue in issues:
        if request.assignee and not _match(issue, assignee=request.assignee):
            continue
        created_date = _parse_issue_date(issue, "created")
        if not created_date:
            continue
        resolved_date = _parse_issue_date(issue, "resolutiondate")
        row = issue_to_row(issue)
        filtered_records.append(
            {
                "issue": issue,
                "row": row,
                "created_date": created_date,
                "resolved_date": resolved_date,
            }
        )

    month_keys = _iter_month_keys(report_start, report_end)
    month_labels = [{"key": key, "label": _month_label(key)} for key in month_keys]

    created_window_records = [
        record
        for record in filtered_records
        if report_start <= record["created_date"] <= report_end
    ]
    created_since_last_report = [
        record
        for record in filtered_records
        if last_report_date <= record["created_date"] <= report_end
    ]

    created_by_month = {key: 0 for key in month_keys}
    resolved_by_month = {key: 0 for key in month_keys}
    status_month_counts: dict[str, Counter[str]] = defaultdict(Counter)

    for record in filtered_records:
        created_date = record["created_date"]
        resolved_date = record["resolved_date"]
        if report_start <= created_date <= report_end:
            created_key = _month_key(created_date)
            created_by_month[created_key] += 1
            status = record["row"].get("status") or "Unknown"
            status_month_counts[status][created_key] += 1
        if resolved_date and report_start <= resolved_date <= report_end:
            resolved_by_month[_month_key(resolved_date)] += 1

    status_rows = []
    for status in sorted(status_month_counts.keys()):
        counts = [status_month_counts[status].get(month_key, 0) for month_key in month_keys]
        status_rows.append(
            {
                "status": status,
                "counts": counts,
                "total": sum(counts),
            }
        )

    month_totals = [created_by_month[month_key] for month_key in month_keys]
    created_vs_resolved = [
        {
            "month_key": month_key,
            "month_label": _month_label(month_key),
            "created": created_by_month[month_key],
            "resolved": resolved_by_month[month_key],
            "net_flow": created_by_month[month_key] - resolved_by_month[month_key],
        }
        for month_key in month_keys
    ]

    since_status_counts = Counter(
        (record["row"].get("status") or "Unknown")
        for record in created_since_last_report
    )
    resolved_since_last_report = sum(
        1
        for record in created_since_last_report
        if record["resolved_date"] and record["resolved_date"] <= report_end
    )
    created_count = len(created_since_last_report)
    open_count = max(created_count - resolved_since_last_report, 0)
    resolution_rate = round((resolved_since_last_report / created_count) * 100, 1) if created_count else 0.0

    detail_rows = []
    for record in sorted(created_since_last_report, key=lambda item: item["created_date"], reverse=True):
        row = record["row"]
        detail_rows.append(
            {
                "key": row.get("key", ""),
                "summary": row.get("summary", ""),
                "status": row.get("status", ""),
                "priority": row.get("priority", ""),
                "assignee": row.get("assignee", ""),
                "reporter": row.get("reporter", ""),
                "created": row.get("created", ""),
                "resolved": row.get("resolved", ""),
                "request_type": row.get("request_type", ""),
                "application": ", ".join(row.get("components") or []),
                "operational_categorization": row.get("work_category", ""),
            }
        )

    status_breakdown = [
        {"status": status, "count": count}
        for status, count in sorted(
            since_status_counts.items(),
            key=lambda item: (-item[1], item[0].lower()),
        )
    ]

    assignee_label = request.assignee or "All assignees"
    return {
        "summary": {
            "assignee": assignee_label,
            "report_start": report_start.isoformat(),
            "report_end": report_end.isoformat(),
            "last_report_date": last_report_date.isoformat(),
            "tickets_created_in_window": len(created_window_records),
            "tickets_resolved_in_window": sum(resolved_by_month.values()),
        },
        "monthly_status": {
            "months": month_labels,
            "rows": status_rows,
            "grand_total": month_totals,
            "grand_total_overall": sum(month_totals),
        },
        "created_vs_resolved": created_vs_resolved,
        "since_last_report": {
            "created_count": created_count,
            "resolved_count": resolved_since_last_report,
            "open_count": open_count,
            "resolution_rate": resolution_rate,
            "status_breakdown": status_breakdown,
            "tickets": detail_rows,
        },
    }


def _build_oasisdev_workload_workbook(report: dict[str, Any]) -> Workbook:
    """Render the OasisDev workload report to Excel."""
    wb = Workbook()

    summary = report["summary"]
    monthly_status = report["monthly_status"]
    created_vs_resolved = report["created_vs_resolved"]
    since_last_report = report["since_last_report"]

    ws = wb.active
    ws.title = "Monthly Status"
    ws["A1"] = "Assignee"
    ws["B1"] = summary["assignee"]
    ws["A2"] = "Report Window"
    ws["B2"] = f"{summary['report_start']} to {summary['report_end']}"
    ws["A3"] = "Last Report Date"
    ws["B3"] = summary["last_report_date"]
    for cell_ref in ("A1", "A2", "A3"):
        ws[cell_ref].font = Font(bold=True)

    month_headers = [month["label"] for month in monthly_status["months"]]
    header_row_idx = 5
    _write_header_row(ws, header_row_idx, ["Status", *month_headers, "Total"])
    current_row = header_row_idx + 1
    for row in monthly_status["rows"]:
        ws.cell(row=current_row, column=1, value=row["status"])
        for col_idx, count in enumerate(row["counts"], 2):
            ws.cell(row=current_row, column=col_idx, value=count)
        ws.cell(row=current_row, column=len(month_headers) + 2, value=row["total"])
        current_row += 1
    ws.cell(row=current_row, column=1, value="Grand Total").font = Font(bold=True)
    for col_idx, count in enumerate(monthly_status["grand_total"], 2):
        ws.cell(row=current_row, column=col_idx, value=count).font = Font(bold=True)
    ws.cell(
        row=current_row,
        column=len(month_headers) + 2,
        value=monthly_status["grand_total_overall"],
    ).font = Font(bold=True)
    ws.freeze_panes = "A6"
    _autofit_columns(
        ws,
        {
            1: 24,
            **{idx + 2: 14 for idx in range(len(month_headers))},
            len(month_headers) + 2: 12,
        },
    )

    flow_ws = wb.create_sheet("Created vs Resolved")
    _write_header_row(flow_ws, 1, ["Month", "Created", "Resolved", "Net Flow"])
    for row_idx, row in enumerate(created_vs_resolved, 2):
        flow_ws.cell(row=row_idx, column=1, value=row["month_label"])
        flow_ws.cell(row=row_idx, column=2, value=row["created"])
        flow_ws.cell(row=row_idx, column=3, value=row["resolved"])
        flow_ws.cell(row=row_idx, column=4, value=row["net_flow"])
    flow_ws.freeze_panes = "A2"
    _autofit_columns(flow_ws, {1: 18, 2: 12, 3: 12, 4: 12})

    since_ws = wb.create_sheet("Since Last Report")
    since_ws["A1"] = "Assignee"
    since_ws["B1"] = summary["assignee"]
    since_ws["A2"] = "Last Report Date"
    since_ws["B2"] = summary["last_report_date"]
    since_ws["A3"] = "New Tickets"
    since_ws["B3"] = since_last_report["created_count"]
    since_ws["A4"] = "Resolved"
    since_ws["B4"] = since_last_report["resolved_count"]
    since_ws["A5"] = "Still Open"
    since_ws["B5"] = since_last_report["open_count"]
    since_ws["A6"] = "Resolution Rate"
    since_ws["B6"] = since_last_report["resolution_rate"] / 100 if since_last_report["created_count"] else 0
    since_ws["B6"].number_format = "0.0%"
    for cell_ref in ("A1", "A2", "A3", "A4", "A5", "A6"):
        since_ws[cell_ref].font = Font(bold=True)

    breakdown_start = 8
    _write_header_row(since_ws, breakdown_start, ["Current Status", "Count"])
    for row_idx, row in enumerate(since_last_report["status_breakdown"], breakdown_start + 1):
        since_ws.cell(row=row_idx, column=1, value=row["status"])
        since_ws.cell(row=row_idx, column=2, value=row["count"])

    detail_start = breakdown_start + len(since_last_report["status_breakdown"]) + 3
    detail_headers = [
        "Key",
        "Summary",
        "Status",
        "Priority",
        "Assignee",
        "Reporter",
        "Created",
        "Resolved",
        "Request Type",
        "Application",
        "Category",
    ]
    _write_header_row(since_ws, detail_start, detail_headers)
    for row_idx, row in enumerate(since_last_report["tickets"], detail_start + 1):
        since_ws.cell(row=row_idx, column=1, value=_cell_value(row["key"]))
        since_ws.cell(row=row_idx, column=2, value=_cell_value(row["summary"]))
        since_ws.cell(row=row_idx, column=3, value=_cell_value(row["status"]))
        since_ws.cell(row=row_idx, column=4, value=_cell_value(row["priority"]))
        since_ws.cell(row=row_idx, column=5, value=_cell_value(row["assignee"]))
        since_ws.cell(row=row_idx, column=6, value=_cell_value(row["reporter"]))
        since_ws.cell(row=row_idx, column=7, value=_cell_value(row["created"]))
        since_ws.cell(row=row_idx, column=8, value=_cell_value(row["resolved"]))
        since_ws.cell(row=row_idx, column=9, value=_cell_value(row["request_type"]))
        since_ws.cell(row=row_idx, column=10, value=_cell_value(row["application"]))
        since_ws.cell(
            row=row_idx,
            column=11,
            value=_cell_value(row["operational_categorization"]),
        )
    since_ws.freeze_panes = f"A{detail_start + 1}"
    _autofit_columns(
        since_ws,
        {
            1: 12,
            2: 48,
            3: 20,
            4: 12,
            5: 24,
            6: 24,
            7: 22,
            8: 22,
            9: 24,
            10: 28,
            11: 24,
        },
    )

    return wb


# ---------------------------------------------------------------------------
# Report builder endpoints
# ---------------------------------------------------------------------------


@router.post("/report/preview")
async def report_preview(config: ReportConfig) -> dict[str, Any]:
    """Return a preview of the report (capped at 100 rows)."""
    _validate_report_window(config)
    try:
        rows = _apply_config(config)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    columns = config.columns or DEFAULT_COLUMNS

    if config.group_by:
        agg = _aggregate(rows, config.group_by)
        return {
            "rows": agg[:100],
            "total_count": len(agg),
            "grouped": True,
        }

    selected = [_select_columns(r, columns) for r in rows[:100]]
    return {
        "rows": selected,
        "total_count": len(rows),
        "grouped": False,
    }


@router.post("/report/export")
async def report_export(config: ReportConfig, template_id: str | None = None) -> FileResponse:
    """Generate and return an Excel workbook from the report config."""
    _validate_report_window(config)
    report_prefix = get_site_profile()["report_prefix"]
    site_scope = get_current_site_scope()
    template = None
    if template_id:
        template = report_template_store.get_template(template_id, site_scope)
        if template is None:
            raise HTTPException(status_code=404, detail="Report template not found.")

    report_name = template.name if template else f"{report_prefix} Report"
    report_description = template.description if template else "Ad hoc report export from the current builder configuration."
    now = datetime.now(timezone.utc)
    filename = f"{report_prefix}_{_sanitize_report_filename(report_name)}_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    all_issues = _calculation_scoped_issues(include_excluded_on_primary=True)
    today = _today_utc()
    await asyncio.to_thread(
        _write_single_report_workbook_file,
        path=tmp_path,
        config=config,
        report_name=report_name,
        report_description=report_description,
        template=template,
        site_scope=site_scope,
        all_issues=all_issues,
        today=today,
    )
    logger.info("Executive report export saved to %s for %s", tmp_path, report_name)

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.unlink, tmp_path),
    )


@router.get("/report/templates", response_model=list[ReportTemplate])
async def list_report_templates(
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[ReportTemplate]:
    """Return saved report templates for the current site scope."""
    site_scope = get_current_site_scope()
    templates = report_template_store.list_templates(site_scope)
    return _apply_runtime_template_readiness(templates, site_scope=site_scope)


@router.get("/report/templates/insights", response_model=list[ReportTemplateInsight])
async def list_report_template_insights(
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[ReportTemplateInsight]:
    """Return rolling operational summaries for saved report templates."""
    templates = report_template_store.list_templates(get_current_site_scope())
    if not templates:
        return []
    today = _today_utc()
    insights: list[ReportTemplateInsight] = []
    for template in templates:
        try:
            insights.append(
                _build_template_insight(
                    template,
                    _calculation_scoped_issues(include_excluded_on_primary=template.config.include_excluded),
                    today=today,
                )
            )
        except ValueError:
            logger.warning("Skipping template insight for invalid window config: %s", template.name)
    return insights


@router.get("/report/templates/master.xlsx")
async def export_master_report_workbook(
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> FileResponse:
    """Export all saved report templates for the current site in a single workbook."""
    templates = report_template_store.list_templates(get_current_site_scope())
    for template in templates:
        if template.include_in_master_export:
            _validate_report_window(template.config, template_name=template.name)
    included_count = sum(1 for template in templates if template.include_in_master_export)
    report_prefix = get_site_profile()["report_prefix"]
    now = datetime.now(timezone.utc)
    filename = f"{report_prefix}_Master_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    site_scope = get_current_site_scope()
    all_issues = _calculation_scoped_issues(include_excluded_on_primary=True)
    today = _today_utc()
    await asyncio.to_thread(
        _write_master_report_workbook_file,
        path=tmp_path,
        templates=templates,
        site_scope=site_scope,
        all_issues=all_issues,
        today=today,
    )
    logger.info("Master report workbook saved to %s (%d templates included)", tmp_path, included_count)
    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.unlink, tmp_path),
    )


@router.post("/report/templates", response_model=ReportTemplate)
async def create_report_template(
    body: ReportTemplateCreateRequest,
    session: dict[str, Any] = Depends(require_authenticated_user),
) -> ReportTemplate:
    """Create a saved report template for the current site scope."""
    _validate_report_window(body.config)
    try:
        return report_template_store.create_template(
            site_scope=get_current_site_scope(),
            name=body.name,
            description=body.description,
            category=body.category,
            notes=body.notes,
            include_in_master_export=body.include_in_master_export,
            config=body.config,
            actor_email=str(session.get("email") or ""),
            actor_name=str(session.get("name") or ""),
        )
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@router.put("/report/templates/{template_id}", response_model=ReportTemplate)
async def update_report_template(
    template_id: str,
    body: ReportTemplateUpdateRequest,
    session: dict[str, Any] = Depends(require_authenticated_user),
) -> ReportTemplate:
    """Update a saved report template for the current site scope."""
    _validate_report_window(body.config)
    try:
        return report_template_store.update_template(
            template_id=template_id,
            site_scope=get_current_site_scope(),
            name=body.name,
            description=body.description,
            category=body.category,
            notes=body.notes,
            include_in_master_export=body.include_in_master_export,
            config=body.config,
            actor_email=str(session.get("email") or ""),
            actor_name=str(session.get("name") or ""),
        )
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@router.delete("/report/templates/{template_id}")
async def delete_report_template(
    template_id: str,
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, bool]:
    """Delete a saved custom report template."""
    try:
        report_template_store.delete_template(template_id, get_current_site_scope())
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc
    return {"deleted": True}


@router.post("/report/templates/{template_id}/export-selection", response_model=ReportTemplate)
async def update_report_template_export_selection(
    template_id: str,
    body: ReportTemplateExportSelectionRequest,
    session: dict[str, Any] = Depends(require_authenticated_user),
) -> ReportTemplate:
    """Update whether a template is included in the master workbook export."""
    try:
        return report_template_store.set_master_export_inclusion(
            template_id=template_id,
            site_scope=get_current_site_scope(),
            include_in_master_export=body.include_in_master_export,
            actor_email=str(session.get("email") or ""),
            actor_name=str(session.get("name") or ""),
        )
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.post("/report/oasisdev-workload")
async def oasisdev_workload_report_preview(
    request: OasisDevWorkloadReportRequest,
) -> dict[str, Any]:
    """Return the OasisDev workload report preview payload."""
    return _build_oasisdev_workload_report(request)


@router.post("/report/oasisdev-workload/export")
async def oasisdev_workload_report_export(
    request: OasisDevWorkloadReportRequest,
) -> FileResponse:
    """Generate and return the OasisDev workload report workbook."""
    report = _build_oasisdev_workload_report(request)
    wb = _build_oasisdev_workload_workbook(report)
    now = datetime.now(timezone.utc)
    filename = f"OasisDev_Workload_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb.save(tmp_path)
    logger.info("OasisDev workload export saved to %s", tmp_path)

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.unlink, tmp_path),
    )


# ---------------------------------------------------------------------------
# Full data export — all fields for all tickets
# ---------------------------------------------------------------------------

# All columns in display order for full export
_FULL_COLUMNS: list[str] = list(FIELD_META.keys())


@router.get("/export/all")
async def export_all_data(include_excluded: bool = False) -> FileResponse:
    """Export ALL ticket data with every available field as Excel."""
    issues = get_scoped_issues(include_excluded_on_primary=include_excluded)
    rows = []
    for iss in issues:
        row = issue_to_row(iss)
        fields = iss.get("fields", {})
        # Override description with full untruncated text
        row["description"] = _extract_description(fields, max_len=0)
        # Add full comment text
        row["comments_text"] = _all_comments_text(fields)
        rows.append(row)
    rows.sort(key=lambda r: r.get("created", ""), reverse=True)

    wb = Workbook()
    ws = wb.active
    report_prefix = get_site_profile()["report_prefix"]
    ws.title = f"{report_prefix} Tickets"

    columns = _FULL_COLUMNS
    headers = [FIELD_META.get(c, {}).get("label", c) for c in columns]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_key in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(row.get(col_key)))

    # Auto-filter & freeze
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    ws.freeze_panes = "A2"

    # Column widths
    _WIDTH_MAP: dict[str, int] = {
        "key": 12, "summary": 50, "description": 60, "issue_type": 18,
        "status": 22, "priority": 12, "assignee": 25, "reporter": 25,
        "created": 22, "updated": 22, "resolved": 22, "request_type": 25,
        "work_category": 20, "calendar_ttr_hours": 12, "age_days": 10,
        "comment_count": 10, "last_comment_date": 22, "last_comment_author": 25,
        "labels": 30, "components": 25, "organizations": 30,
        "comments_text": 60,
    }
    for col_idx, col_key in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = _WIDTH_MAP.get(col_key, 15)

    now = datetime.now(timezone.utc)
    filename = f"{report_prefix}_All_Data_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb.save(tmp_path)
    logger.info("Full data export: %d tickets, %d columns to %s", len(rows), len(columns), tmp_path)

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.unlink, tmp_path),
    )


# ---------------------------------------------------------------------------
# Legacy full export (unchanged)
# ---------------------------------------------------------------------------


@router.get("/export/excel")
async def export_excel() -> FileResponse:
    """Generate and return an Excel workbook with all OIT issues."""
    logger.info("Starting Excel export from cache")

    # Read from the current site scope.
    issues = _calculation_scoped_issues(include_excluded_on_primary=True)
    logger.info("Export: %d issues from cache", len(issues))

    # Create workbook
    wb = Workbook()
    ws = wb.active
    report_prefix = get_site_profile()["report_prefix"]
    ws.title = f"{report_prefix} Tickets"

    # Write header row
    for col_idx, (header_text, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    # Write data rows
    for row_idx, issue in enumerate(issues, start=2):
        row_data = issue_to_row(issue)
        for col_idx, (_, dict_key) in enumerate(_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(row_data.get(dict_key)))

    # Auto-filter on all columns
    last_col_letter = chr(ord("A") + len(_COLUMNS) - 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(issues) + 1}"

    # Freeze the top row
    ws.freeze_panes = "A2"

    # Adjust column widths for readability
    _WIDTHS: dict[str, int] = {
        "Key": 12,
        "Summary": 50,
        "Type": 18,
        "Status": 22,
        "Priority": 12,
        "Assignee": 25,
        "Reporter": 25,
        "Created": 22,
        "Updated": 22,
        "Resolved": 22,
        "TTR (h)": 12,
        "Age (d)": 10,
        "SLA Response": 15,
        "SLA Resolution": 15,
        "Excluded": 10,
    }
    for col_idx, (header_text, _) in enumerate(_COLUMNS, start=1):
        col_letter = chr(ord("A") + col_idx - 1)
        ws.column_dimensions[col_letter].width = _WIDTHS.get(header_text, 15)

    # Save to temp file
    now = datetime.now(timezone.utc)
    filename = f"{report_prefix}_Report_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()

    wb.save(tmp_path)
    logger.info("Excel workbook saved to %s (%d rows)", tmp_path, len(issues))

    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.unlink, tmp_path),
    )
