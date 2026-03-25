"""Primary-site user administration APIs."""

from __future__ import annotations

import csv
import os
import tempfile
from datetime import datetime, timedelta, timezone
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from starlette.background import BackgroundTask

from auth import require_can_manage_users, session_is_admin
from azure_cache import azure_cache
from models import (
    UserAdminAuditEntryResponse,
    UserAdminCapabilitiesResponse,
    UserAdminDeviceResponse,
    UserAdminGroupMembershipResponse,
    UserAdminJobCreateRequest,
    UserAdminJobResponse,
    UserAdminJobResultResponse,
    UserAdminLicenseResponse,
    UserAdminMailboxResponse,
    UserAdminRoleResponse,
    UserAdminUserDetailResponse,
    UserExitReportFilter,
)
from site_context import get_current_site_scope
from user_admin_jobs import user_admin_jobs
from user_admin_providers import UserAdminProviderError, user_admin_providers

router = APIRouter(prefix="/api/user-admin")

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_USER_EXPORT_HEADERS = [
    "Display Name",
    "User Principal Name",
    "Email",
    "Status",
    "User Type",
    "Department",
    "Job Title",
    "Directory",
    "On-Prem Synced",
    "On-Prem Domain",
    "On-Prem NetBIOS",
    "On-Prem SAM Account Name",
    "On-Prem Distinguished Name",
    "Created UTC",
    "Last Password Change UTC",
    "Licensed",
    "License Count",
    "SKU Part Numbers",
    "Last Interactive UTC",
    "Last Interactive PT",
    "Last Noninteractive UTC",
    "Last Noninteractive PT",
    "Last Successful UTC",
    "Last Successful PT",
    "Office Location",
    "Company",
    "City",
    "Country",
    "Mobile Phone",
    "Business Phones",
    "Proxy Addresses",
]


def _ensure_primary_site() -> None:
    if get_current_site_scope() != "primary":
        raise HTTPException(
            status_code=404,
            detail="User administration APIs are only available on it-app.movedocs.com",
        )


def _safe_export_text(value: Any) -> str:
    text = str(value or "")
    if text and text[0] in ("=", "+", "-", "@"):
        return "\t" + text
    return text


def _parse_datetime(value: str) -> datetime | None:
    if not value:
        return None
    normalized = value.strip()
    if not normalized:
        return None
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _directory_label(user: dict[str, Any]) -> str:
    extra = user.get("extra") if isinstance(user.get("extra"), dict) else {}
    if extra.get("on_prem_domain"):
        return str(extra.get("on_prem_domain") or "")
    if str(extra.get("user_type") or "") == "Guest":
        return "External"
    return "Cloud"


def _is_licensed_user(user: dict[str, Any]) -> bool:
    extra = user.get("extra") if isinstance(user.get("extra"), dict) else {}
    return str(extra.get("is_licensed") or "").strip().lower() == "true"


def _is_on_prem_synced_user(user: dict[str, Any]) -> bool:
    extra = user.get("extra") if isinstance(user.get("extra"), dict) else {}
    return str(extra.get("on_prem_sync") or "").strip().lower() == "true"


def _has_no_successful_sign_in_30d(user: dict[str, Any]) -> bool:
    if user.get("enabled") is not True:
        return False
    extra = user.get("extra") if isinstance(user.get("extra"), dict) else {}
    last_successful = _parse_datetime(str(extra.get("last_successful_utc") or ""))
    if last_successful is None:
        return True
    return last_successful <= datetime.now(timezone.utc) - timedelta(days=30)


def _matches_report_filter(user: dict[str, Any], report_filter: UserExitReportFilter) -> bool:
    if report_filter == "disabled_licensed":
        return user.get("enabled") is False and _is_licensed_user(user)
    if report_filter == "active_no_success_30d":
        return _has_no_successful_sign_in_30d(user)
    return True


def _filter_directory_users(
    *,
    search: str = "",
    status: str = "all",
    type: str = "all",
    license: str = "all",
    activity: str = "all",
    sync: str = "all",
    directory: str = "",
    report_filter: UserExitReportFilter = "",
    scope: str = "filtered",
) -> list[dict[str, Any]]:
    rows = azure_cache.list_directory_objects("users", search=search if scope == "filtered" else "")
    if scope == "all":
        return rows

    filtered: list[dict[str, Any]] = []
    status_value = status.strip().lower()
    type_value = type.strip().lower()
    license_value = license.strip().lower()
    activity_value = activity.strip().lower()
    sync_value = sync.strip().lower()
    directory_value = directory.strip().lower()
    for user in rows:
        extra = user.get("extra") if isinstance(user.get("extra"), dict) else {}
        if status_value == "enabled" and user.get("enabled") is not True:
            continue
        if status_value == "disabled" and user.get("enabled") is not False:
            continue
        if type_value == "member" and str(extra.get("user_type") or "") == "Guest":
            continue
        if type_value == "guest" and str(extra.get("user_type") or "") != "Guest":
            continue
        if license_value == "licensed" and not _is_licensed_user(user):
            continue
        if activity_value == "no_success_30d" and not _has_no_successful_sign_in_30d(user):
            continue
        if sync_value == "on_prem_synced" and not _is_on_prem_synced_user(user):
            continue
        if directory_value and _directory_label(user).lower() != directory_value:
            continue
        if report_filter and not _matches_report_filter(user, report_filter):
            continue
        filtered.append(user)
    return filtered


def _user_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    export_rows: list[dict[str, Any]] = []
    for user in rows:
        extra = user.get("extra") if isinstance(user.get("extra"), dict) else {}
        export_rows.append(
            {
                "Display Name": _safe_export_text(user.get("display_name") or ""),
                "User Principal Name": _safe_export_text(user.get("principal_name") or ""),
                "Email": _safe_export_text(user.get("mail") or ""),
                "Status": "Enabled" if user.get("enabled") is True else "Disabled" if user.get("enabled") is False else "Unknown",
                "User Type": _safe_export_text(extra.get("user_type") or ""),
                "Department": _safe_export_text(extra.get("department") or ""),
                "Job Title": _safe_export_text(extra.get("job_title") or ""),
                "Directory": _safe_export_text(_directory_label(user)),
                "On-Prem Synced": "Yes" if str(extra.get("on_prem_sync") or "").strip().lower() == "true" else "No",
                "On-Prem Domain": _safe_export_text(extra.get("on_prem_domain") or ""),
                "On-Prem NetBIOS": _safe_export_text(extra.get("on_prem_netbios") or ""),
                "On-Prem SAM Account Name": _safe_export_text(extra.get("on_prem_sam_account_name") or ""),
                "On-Prem Distinguished Name": _safe_export_text(extra.get("on_prem_distinguished_name") or ""),
                "Created UTC": _safe_export_text(extra.get("created_datetime") or ""),
                "Last Password Change UTC": _safe_export_text(extra.get("last_password_change") or ""),
                "Licensed": "Yes" if str(extra.get("is_licensed") or "").strip().lower() == "true" else "No",
                "License Count": int(extra.get("license_count") or 0) if str(extra.get("license_count") or "").strip() else 0,
                "SKU Part Numbers": _safe_export_text(extra.get("sku_part_numbers") or ""),
                "Last Interactive UTC": _safe_export_text(extra.get("last_interactive_utc") or ""),
                "Last Interactive PT": _safe_export_text(extra.get("last_interactive_local") or ""),
                "Last Noninteractive UTC": _safe_export_text(extra.get("last_noninteractive_utc") or ""),
                "Last Noninteractive PT": _safe_export_text(extra.get("last_noninteractive_local") or ""),
                "Last Successful UTC": _safe_export_text(extra.get("last_successful_utc") or ""),
                "Last Successful PT": _safe_export_text(extra.get("last_successful_local") or ""),
                "Office Location": _safe_export_text(extra.get("office_location") or ""),
                "Company": _safe_export_text(extra.get("company_name") or ""),
                "City": _safe_export_text(extra.get("city") or ""),
                "Country": _safe_export_text(extra.get("country") or ""),
                "Mobile Phone": _safe_export_text(extra.get("mobile_phone") or ""),
                "Business Phones": _safe_export_text(extra.get("business_phones") or ""),
                "Proxy Addresses": _safe_export_text(extra.get("proxy_addresses") or ""),
            }
        )
    return export_rows


def _delete_file(path: str) -> None:
    try:
        os.unlink(path)
    except FileNotFoundError:
        pass


def _build_file_response(path: str, filename: str, media_type: str) -> FileResponse:
    return FileResponse(
        path,
        filename=filename,
        media_type=media_type,
        background=BackgroundTask(_delete_file, path),
    )


def _write_export_workbook(title: str, rows: list[dict[str, Any]]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    headers = list(rows[0].keys()) if rows else list(_USER_EXPORT_HEADERS)

    for column_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for row_index, row in enumerate(rows, 2):
        for column_index, header in enumerate(headers, 1):
            ws.cell(row=row_index, column=column_index, value=row.get(header, ""))

    for index, header in enumerate(headers, 1):
        width = min(max(len(str(header)) + 4, 14), 36)
        ws.column_dimensions[get_column_letter(index)].width = width

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(tmp.name)
    finally:
        tmp.close()
    return tmp.name


@router.get("/capabilities", response_model=UserAdminCapabilitiesResponse)
def get_user_admin_capabilities(session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    return user_admin_providers.get_capabilities()


@router.get("/users/{user_id}/detail", response_model=UserAdminUserDetailResponse)
def get_user_detail(user_id: str, session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    try:
        return user_admin_providers.get_user_detail(user_id)
    except UserAdminProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/users/{user_id}/groups", response_model=list[UserAdminGroupMembershipResponse])
def get_user_groups(user_id: str, session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    try:
        return user_admin_providers.list_groups(user_id)
    except UserAdminProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/users/{user_id}/licenses", response_model=list[UserAdminLicenseResponse])
def get_user_licenses(user_id: str, session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    try:
        return user_admin_providers.list_licenses(user_id)
    except UserAdminProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/users/{user_id}/roles", response_model=list[UserAdminRoleResponse])
def get_user_roles(user_id: str, session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    try:
        return user_admin_providers.list_roles(user_id)
    except UserAdminProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/users/{user_id}/mailbox", response_model=UserAdminMailboxResponse)
def get_user_mailbox(user_id: str, session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    try:
        return user_admin_providers.get_mailbox(user_id)
    except UserAdminProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/users/{user_id}/devices", response_model=list[UserAdminDeviceResponse])
def get_user_devices(user_id: str, session: dict = Depends(require_can_manage_users)):
    del session
    _ensure_primary_site()
    try:
        return user_admin_providers.list_devices(user_id)
    except UserAdminProviderError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@router.get("/users/{user_id}/activity", response_model=list[UserAdminAuditEntryResponse])
def get_user_activity(
    user_id: str,
    limit: int = Query(50, ge=1, le=200),
    session: dict = Depends(require_can_manage_users),
):
    del session
    _ensure_primary_site()
    return user_admin_jobs.list_audit(limit=limit, target_user_id=user_id)


@router.post("/jobs", response_model=UserAdminJobResponse)
def create_user_admin_job(
    body: UserAdminJobCreateRequest,
    session: dict = Depends(require_can_manage_users),
):
    _ensure_primary_site()
    try:
        return user_admin_jobs.create_job(
            action_type=body.action_type,
            target_user_ids=body.target_user_ids,
            params=body.params,
            requested_by_email=str(session.get("email") or ""),
            requested_by_name=str(session.get("name") or ""),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/jobs/{job_id}", response_model=UserAdminJobResponse)
def get_user_admin_job(job_id: str, session: dict = Depends(require_can_manage_users)):
    _ensure_primary_site()
    if not user_admin_jobs.job_belongs_to(job_id, str(session.get("email") or ""), is_admin=session_is_admin(session)):
        raise HTTPException(status_code=404, detail="Job not found")
    job = user_admin_jobs.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.get("/jobs/{job_id}/results", response_model=list[UserAdminJobResultResponse])
def get_user_admin_job_results(job_id: str, session: dict = Depends(require_can_manage_users)):
    _ensure_primary_site()
    if not user_admin_jobs.job_belongs_to(job_id, str(session.get("email") or ""), is_admin=session_is_admin(session)):
        raise HTTPException(status_code=404, detail="Job not found")
    if not user_admin_jobs.get_job(job_id):
        raise HTTPException(status_code=404, detail="Job not found")
    return user_admin_jobs.get_job_results(job_id)


@router.get("/audit", response_model=list[UserAdminAuditEntryResponse])
def get_user_admin_audit(
    limit: int = Query(100, ge=1, le=500),
    session: dict = Depends(require_can_manage_users),
):
    del session
    _ensure_primary_site()
    return user_admin_jobs.list_audit(limit=limit)


@router.get("/users/export.csv")
def export_user_admin_users_csv(
    search: str = Query(default=""),
    status: str = Query(default="all"),
    type: str = Query(default="all"),
    license: str = Query(default="all"),
    activity: str = Query(default="all"),
    sync: str = Query(default="all"),
    directory: str = Query(default=""),
    report_filter: UserExitReportFilter = Query(default=""),
    scope: str = Query(default="filtered"),
    session: dict = Depends(require_can_manage_users),
) -> FileResponse:
    del session
    _ensure_primary_site()
    rows = _user_export_rows(
        _filter_directory_users(
            search=search,
            status=status,
            type=type,
            license=license,
            activity=activity,
            sync=sync,
            directory=directory,
            report_filter=report_filter,
            scope=scope,
        )
    )
    now = datetime.now(timezone.utc)
    filename = f"entra_users_{scope}_{now.strftime('%Y%m%d_%H%M')}.csv"
    headers = list(rows[0].keys()) if rows else list(_USER_EXPORT_HEADERS)
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    try:
        writer = csv.DictWriter(tmp, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        tmp.flush()
    finally:
        tmp.close()
    return _build_file_response(tmp.name, filename, "text/csv; charset=utf-8")


@router.get("/users/export.xlsx")
def export_user_admin_users_excel(
    search: str = Query(default=""),
    status: str = Query(default="all"),
    type: str = Query(default="all"),
    license: str = Query(default="all"),
    activity: str = Query(default="all"),
    sync: str = Query(default="all"),
    directory: str = Query(default=""),
    report_filter: UserExitReportFilter = Query(default=""),
    scope: str = Query(default="filtered"),
    session: dict = Depends(require_can_manage_users),
) -> FileResponse:
    del session
    _ensure_primary_site()
    rows = _user_export_rows(
        _filter_directory_users(
            search=search,
            status=status,
            type=type,
            license=license,
            activity=activity,
            sync=sync,
            directory=directory,
            report_filter=report_filter,
            scope=scope,
        )
    )
    now = datetime.now(timezone.utc)
    filename = f"entra_users_{scope}_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    path = _write_export_workbook("Entra Users", rows)
    return _build_file_response(path, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
