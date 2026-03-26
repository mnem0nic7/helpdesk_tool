"""API routes for the Azure portal site."""

from __future__ import annotations

import asyncio
import csv
import logging
import os
import tempfile
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from starlette.background import BackgroundTask

from ai_client import (
    answer_azure_cost_question,
    get_available_copilot_models,
    get_default_copilot_model_id,
)
from auth import require_admin, require_authenticated_user, session_is_admin
from azure_cache import azure_cache
from azure_cost_exports import azure_cost_export_service
from azure_finops import azure_finops_service
from azure_vm_export_jobs import azure_vm_export_jobs
from azure_alert_engine import send_recommendation_teams_alert
from jira_write_service import append_fallback_actor_block, get_jira_write_context
from config import (
    AZURE_APP_HOST,
    AZURE_FINOPS_RECOMMENDATION_JIRA_ISSUE_TYPE,
    AZURE_FINOPS_RECOMMENDATION_JIRA_PROJECT,
    AZURE_FINOPS_RECOMMENDATION_TEAMS_CHANNEL_LABEL,
    AZURE_FINOPS_RECOMMENDATION_TEAMS_WEBHOOK_URL,
    AZURE_REPORTING_COST_ANALYSIS_LABEL,
    AZURE_REPORTING_COST_ANALYSIS_URL,
    AZURE_REPORTING_POWER_BI_LABEL,
    AZURE_REPORTING_POWER_BI_URL,
    JIRA_BASE_URL,
)
from jira_client import JiraClient
from models import (
    AzureAllocationRuleRequest,
    AzureAllocationRunRequest,
    AzureCostChatRequest,
    AzureRecommendationActionContractResponse,
    AzureRecommendationActionStateRequest,
    AzureRecommendationCreateTicketRequest,
    AzureRecommendationCreateTicketResponse,
    AzureRecommendationRunSafeScriptRequest,
    AzureRecommendationRunSafeScriptResponse,
    AzureRecommendationSendAlertRequest,
    AzureRecommendationSendAlertResponse,
    AzureRecommendationDismissRequest,
    AzureRecommendationReopenRequest,
    AzureSavingsOpportunity,
    AzureSavingsSummary,
    AzureVirtualMachineCostExportJobCreateRequest,
    AzureVirtualMachineCostExportJobResponse,
    AzureVirtualMachineDetailResponse,
)
from site_context import get_current_site_scope

router = APIRouter(prefix="/api/azure")
logger = logging.getLogger(__name__)
_jira_client = JiraClient()

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _ensure_azure_site() -> None:
    if get_current_site_scope() != "azure":
        raise HTTPException(status_code=404, detail="Azure portal APIs are only available on azure.movedocs.com")


def _ensure_azure_or_primary_site() -> None:
    if get_current_site_scope() not in {"azure", "primary"}:
        raise HTTPException(
            status_code=404,
            detail="Azure directory user APIs are only available on azure.movedocs.com and it-app.movedocs.com",
        )


def _safe_export_text(value: Any) -> str:
    text = str(value or "")
    if text and text[0] in ("=", "+", "-", "@"):
        return "\t" + text
    return text


def _format_finops_currency(value: Any, currency: str = "USD") -> str:
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return "Unavailable"
    return f"{currency or 'USD'} {amount:,.2f}"


def _default_recommendation_ticket_summary(recommendation: dict[str, Any]) -> str:
    title = str(recommendation.get("title") or "Azure FinOps recommendation").strip()
    if title.lower().startswith("[finops]"):
        return title
    return f"[FinOps] {title}"


def _build_recommendation_ticket_description(
    recommendation: dict[str, Any],
    *,
    operator_note: str = "",
) -> str:
    currency = str(recommendation.get("currency") or "USD")
    lines = [
        "Azure FinOps recommendation follow-up",
        "",
        f"Title: {str(recommendation.get('title') or '').strip()}",
        f"Summary: {str(recommendation.get('summary') or '').strip()}",
        f"Category: {str(recommendation.get('category') or '').strip()}",
        f"Opportunity type: {str(recommendation.get('opportunity_type') or '').strip()}",
        f"Estimated monthly savings: {_format_finops_currency(recommendation.get('estimated_monthly_savings'), currency)}",
        f"Current monthly cost: {_format_finops_currency(recommendation.get('current_monthly_cost'), currency)}",
        f"Subscription: {str(recommendation.get('subscription_name') or recommendation.get('subscription_id') or 'Unavailable').strip()}",
        f"Resource group: {str(recommendation.get('resource_group') or 'Unavailable').strip()}",
        f"Resource name: {str(recommendation.get('resource_name') or 'Unavailable').strip()}",
        f"Resource ID: {str(recommendation.get('resource_id') or 'Unavailable').strip()}",
        f"Azure Portal: {str(recommendation.get('portal_url') or '').strip() or 'Unavailable'}",
    ]

    follow_up_route = str(recommendation.get("follow_up_route") or "").strip()
    if follow_up_route:
        lines.append(f"Dashboard follow-up: https://{AZURE_APP_HOST}{follow_up_route}")

    recommended_steps = recommendation.get("recommended_steps") or []
    if isinstance(recommended_steps, list) and recommended_steps:
        lines.extend(["", "Recommended steps:"])
        for index, step in enumerate(recommended_steps, start=1):
            step_text = str(step or "").strip()
            if step_text:
                lines.append(f"{index}. {step_text}")

    evidence_rows = recommendation.get("evidence") or []
    if isinstance(evidence_rows, list) and evidence_rows:
        lines.extend(["", "Evidence:"])
        for row in evidence_rows:
            if not isinstance(row, dict):
                continue
            label = str(row.get("label") or "").strip()
            value = str(row.get("value") or "").strip()
            if label or value:
                lines.append(f"- {label or 'Detail'}: {value or 'Unavailable'}")

    note_text = operator_note.strip()
    if note_text:
        lines.extend(["", "Operator note:", note_text])

    lines.extend(
        [
            "",
            "Created from the Azure FinOps recommendations workspace.",
            f"Generated at: {datetime.now(timezone.utc).isoformat()}",
        ]
    )
    return "\n".join(lines)


def _coverage_label(delta: Any) -> str:
    if delta is None:
        return "Unavailable"
    try:
        amount = int(delta)
    except (TypeError, ValueError):
        return "Unavailable"
    if amount > 0:
        return f"{amount} needed"
    if amount < 0:
        return f"{abs(amount)} excess"
    return "Balanced"


def _vm_coverage_export_rows() -> list[dict[str, Any]]:
    payload = azure_cache.list_virtual_machines()
    rows: list[dict[str, Any]] = []
    for item in payload.get("by_size") or []:
        rows.append(
            {
                "SKU": _safe_export_text(item.get("label") or ""),
                "Region": _safe_export_text(item.get("region") or ""),
                "VMs": int(item.get("vm_count") or 0),
                "Reserved Instances (RI)": (
                    int(item.get("reserved_instance_count") or 0)
                    if item.get("reserved_instance_count") is not None
                    else ""
                ),
                "Needed / Excess": _safe_export_text(_coverage_label(item.get("delta"))),
            }
        )
    return rows


def _vm_excess_export_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in azure_cache.get_vm_excess_reservation_report():
        rows.append(
            {
                "SKU": _safe_export_text(item.get("label") or ""),
                "Region": _safe_export_text(item.get("region") or ""),
                "VMs": int(item.get("vm_count") or 0),
                "Reserved Instances (RI)": int(item.get("reserved_instance_count") or 0),
                "Excess": int(item.get("excess_count") or 0),
                "Active Reservation Names": _safe_export_text(
                    "; ".join(item.get("active_reservation_names") or [])
                ),
            }
        )
    return rows


def _delete_file(path: str) -> None:
    try:
        os.unlink(path)
    except FileNotFoundError:
        pass


def _build_vm_coverage_file_response(path: str, filename: str, media_type: str) -> FileResponse:
    return FileResponse(
        path,
        filename=filename,
        media_type=media_type,
        background=BackgroundTask(_delete_file, path),
    )


def _savings_export_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    export_rows: list[dict[str, Any]] = []
    for item in rows:
        evidence = "; ".join(
            f"{row.get('label', '')}: {row.get('value', '')}"
            for row in (item.get("evidence") or [])
            if isinstance(row, dict)
        )
        export_rows.append(
            {
                "Title": _safe_export_text(item.get("title") or ""),
                "Category": _safe_export_text(item.get("category") or ""),
                "Opportunity Type": _safe_export_text(item.get("opportunity_type") or ""),
                "Source": _safe_export_text(item.get("source") or ""),
                "Subscription": _safe_export_text(item.get("subscription_name") or item.get("subscription_id") or ""),
                "Resource Group": _safe_export_text(item.get("resource_group") or ""),
                "Location": _safe_export_text(item.get("location") or ""),
                "Resource Name": _safe_export_text(item.get("resource_name") or ""),
                "Resource Type": _safe_export_text(item.get("resource_type") or ""),
                "Current Monthly Cost": item.get("current_monthly_cost") if item.get("current_monthly_cost") is not None else "",
                "Estimated Monthly Savings": (
                    item.get("estimated_monthly_savings")
                    if item.get("estimated_monthly_savings") is not None
                    else ""
                ),
                "Quantified": "Yes" if item.get("quantified") else "No",
                "Effort": _safe_export_text(item.get("effort") or ""),
                "Risk": _safe_export_text(item.get("risk") or ""),
                "Confidence": _safe_export_text(item.get("confidence") or ""),
                "Estimate Basis": _safe_export_text(item.get("estimate_basis") or ""),
                "Summary": _safe_export_text(item.get("summary") or ""),
                "Recommended Steps": _safe_export_text("; ".join(item.get("recommended_steps") or [])),
                "Evidence": _safe_export_text(evidence),
                "Portal URL": _safe_export_text(item.get("portal_url") or ""),
                "Follow Up Route": _safe_export_text(item.get("follow_up_route") or ""),
            }
        )
    return export_rows


def _list_filtered_savings_opportunities(
    *,
    search: str = "",
    category: str = "",
    opportunity_type: str = "",
    subscription_id: str = "",
    resource_group: str = "",
    effort: str = "",
    risk: str = "",
    confidence: str = "",
    quantified_only: bool = False,
) -> list[dict[str, Any]]:
    _refresh_finops_recommendations_from_cache()
    finops_rows = azure_finops_service.list_recommendations(
        search=search,
        category=category,
        opportunity_type=opportunity_type,
        subscription_id=subscription_id,
        resource_group=resource_group,
        effort=effort,
        risk=risk,
        confidence=confidence,
        quantified_only=quantified_only,
    )
    if finops_rows:
        return finops_rows
    return azure_cache.list_savings_opportunities(
        search=search,
        category=category,
        opportunity_type=opportunity_type,
        subscription_id=subscription_id,
        resource_group=resource_group,
        effort=effort,
        risk=risk,
        confidence=confidence,
        quantified_only=quantified_only,
    )


def _refresh_finops_recommendations_from_cache() -> dict[str, Any] | None:
    cache_rows = azure_cache.list_savings_opportunities()
    cache_resources = azure_cache._snapshot("resources") or []
    cache_status = azure_cache.status()
    cache_source_refreshed_at = str(cache_status.get("last_refresh") or "")
    cache_source_version = cache_source_refreshed_at or f"rows:{len(cache_rows)}"
    inventory_source_version = cache_source_refreshed_at or f"resources:{len(cache_resources)}"
    try:
        return azure_finops_service.refresh_recommendations_snapshot(
            cache_rows,
            cache_source_version=cache_source_version,
            cache_source_refreshed_at=cache_source_refreshed_at,
            cache_resources=cache_resources,
            inventory_source_version=inventory_source_version,
        )
    except Exception:
        logger.exception("Failed to refresh local FinOps recommendation snapshot from Azure cache")
        return None


def _get_savings_summary_payload() -> dict[str, Any]:
    _refresh_finops_recommendations_from_cache()
    finops_summary = azure_finops_service.get_recommendation_summary()
    if finops_summary:
        return _savings_summary_with_context(finops_summary, source="recommendations")
    return _savings_summary_with_context(azure_cache.get_savings_summary(), source="cache")


def _list_filtered_recommendations(
    *,
    search: str = "",
    category: str = "",
    opportunity_type: str = "",
    subscription_id: str = "",
    resource_group: str = "",
    effort: str = "",
    risk: str = "",
    confidence: str = "",
    quantified_only: bool = False,
) -> list[dict[str, Any]]:
    _refresh_finops_recommendations_from_cache()
    return azure_finops_service.list_recommendations(
        search=search,
        category=category,
        opportunity_type=opportunity_type,
        subscription_id=subscription_id,
        resource_group=resource_group,
        effort=effort,
        risk=risk,
        confidence=confidence,
        quantified_only=quantified_only,
    )


def _get_export_job_or_404(job_id: str) -> dict[str, Any]:
    job = azure_vm_export_jobs.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Azure VM cost export job was not found")
    return job


def _ensure_export_job_access(job_id: str, session: dict[str, Any]) -> dict[str, Any]:
    job = _get_export_job_or_404(job_id)
    if not azure_vm_export_jobs.job_belongs_to(
        job_id,
        str(session.get("email") or ""),
        is_admin=session_is_admin(session),
    ):
        raise HTTPException(status_code=403, detail="You do not have access to this Azure VM export job")
    return job


def _azure_status_with_exports(payload: dict[str, Any]) -> dict[str, Any]:
    enriched = dict(payload)
    enriched["cost_exports"] = azure_cost_export_service.status()
    finops_status = azure_finops_service.get_status()
    cost_context = _get_cost_context_payload()
    if "cost" in enriched:
        enriched["cost"] = _get_cost_summary_payload()
    enriched["reporting"] = {
        "power_bi": {
            "label": AZURE_REPORTING_POWER_BI_LABEL,
            "url": AZURE_REPORTING_POWER_BI_URL or None,
            "configured": bool(AZURE_REPORTING_POWER_BI_URL),
            "description": "Shared finance and showback reporting from governed Cost Management exports.",
        },
        "cost_analysis": {
            "label": AZURE_REPORTING_COST_ANALYSIS_LABEL,
            "url": AZURE_REPORTING_COST_ANALYSIS_URL or None,
            "configured": bool(AZURE_REPORTING_COST_ANALYSIS_URL),
            "description": "Interactive Azure-native cost exploration and saved views for drill-down analysis.",
        },
        "sources": {
            "overview": {
                "label": "Cached inventory + export-backed cost" if cost_context.get("export_backed") else "Cached app data",
                "description": (
                    "Overview inventory and identity metrics come from cached Azure snapshots, while cost metrics prefer local export-backed analytics."
                    if cost_context.get("export_backed")
                    else "Overview metrics come from the app's cached Azure snapshots and cost queries."
                ),
            },
            "cost": {
                "label": "Export-backed local analytics",
                "description": "Cost charts and tables prefer local FinOps analytics hydrated from Cost Management exports, with cache fallback when exports are unavailable.",
            },
            "savings": {
                "label": "Heuristic operational guidance",
                "description": "Savings recommendations blend cached Azure data, Advisor signals, and app heuristics.",
            },
            "exports": {
                "label": "Export-backed governed reporting",
                "description": "Shared reporting should come from Cost Management exports and governed BI assets.",
            },
        },
    }
    enriched["finops"] = {
        "available": bool(finops_status.get("available")),
        "record_count": int(finops_status.get("record_count") or 0),
        "coverage_start": finops_status.get("coverage_start"),
        "coverage_end": finops_status.get("coverage_end"),
        "field_coverage": finops_status.get("field_coverage") or {},
        "ai_usage": finops_status.get("ai_usage") or {},
        "cost_context": cost_context,
    }
    return enriched


def _get_finops_validation_payload() -> dict[str, Any]:
    return azure_finops_service.get_validation_report(
        azure_cache.get_cost_summary(),
        azure_cost_export_service.status(),
    )


def _get_cost_copilot_context() -> dict[str, Any]:
    context = dict(azure_cache.get_grounding_context())
    export_summary = azure_finops_service.get_cost_summary()
    if export_summary:
        context["export_cost_summary"] = export_summary
        context["export_cost_trend"] = azure_finops_service.get_cost_trend()
        context["export_cost_by_service"] = azure_finops_service.get_cost_breakdown("service")
    finops_status = azure_finops_service.get_status()
    context["finops_status"] = {
        "available": bool(finops_status.get("available")),
        "record_count": int(finops_status.get("record_count") or 0),
        "coverage_start": finops_status.get("coverage_start"),
        "coverage_end": finops_status.get("coverage_end"),
        "field_coverage": finops_status.get("field_coverage") or {},
        "ai_usage": finops_status.get("ai_usage") or {},
    }
    return context


def _cost_summary_with_source(payload: dict[str, Any], *, source: str) -> dict[str, Any]:
    total_cost = float(payload.get("total_cost") or 0.0)
    enriched = dict(payload)
    enriched.setdefault("total_actual_cost", total_cost)
    enriched.setdefault("total_amortized_cost", total_cost)
    enriched.setdefault("record_count", 0)
    enriched.setdefault("window_start", None)
    enriched.setdefault("window_end", None)
    enriched["source"] = source
    enriched["source_label"] = "Export-backed local analytics" if source == "exports" else "Cached app data"
    enriched["export_backed"] = source == "exports"
    return enriched


def _cost_trend_with_source(rows: list[dict[str, Any]], *, source: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        cost = float(row.get("cost") or 0.0)
        enriched = dict(row)
        enriched.setdefault("actual_cost", cost)
        enriched.setdefault("amortized_cost", cost)
        enriched.setdefault("currency", str(row.get("currency") or "USD"))
        enriched["source"] = source
        result.append(enriched)
    return result


def _cost_breakdown_with_source(rows: list[dict[str, Any]], *, source: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        amount = float(row.get("amount") or 0.0)
        enriched = dict(row)
        enriched.setdefault("actual_cost", amount)
        enriched.setdefault("amortized_cost", amount)
        enriched.setdefault("currency", str(row.get("currency") or "USD"))
        enriched["source"] = source
        result.append(enriched)
    return result


def _get_cost_context_payload() -> dict[str, Any]:
    export_summary = azure_finops_service.get_cost_summary()
    if export_summary:
        return {
            "available": True,
            "source": "exports",
            "source_label": "Export-backed local analytics",
            "source_description": (
                "Cost context comes from parsed Azure Cost Management exports in the local FinOps analytics store."
            ),
            "window_start": export_summary.get("window_start"),
            "window_end": export_summary.get("window_end"),
            "record_count": int(export_summary.get("record_count") or 0),
            "currency": str(export_summary.get("currency") or "USD"),
            "total_actual_cost": float(export_summary.get("total_actual_cost") or export_summary.get("total_cost") or 0.0),
            "total_amortized_cost": float(
                export_summary.get("total_amortized_cost") or export_summary.get("total_cost") or 0.0
            ),
            "export_backed": True,
        }

    cache_summary = azure_cache.get_cost_summary()
    total_cost = float(cache_summary.get("total_cost") or 0.0)
    return {
        "available": bool(cache_summary),
        "source": "cache",
        "source_label": "Cached app data",
        "source_description": (
            "Cost context is currently coming from cached Azure Cost Management query results because export-backed facts are unavailable."
        ),
        "window_start": None,
        "window_end": None,
        "record_count": 0,
        "currency": str(cache_summary.get("currency") or "USD"),
        "total_actual_cost": total_cost,
        "total_amortized_cost": total_cost,
        "export_backed": False,
    }


def _get_recommendation_workspace_context(*, source: str) -> dict[str, Any]:
    finops_status = azure_finops_service.get_status()
    recommendation_status = finops_status.get("recommendations") or {}
    last_refreshed_at = recommendation_status.get("last_refreshed_at")
    if last_refreshed_at is not None:
        last_refreshed_at = str(last_refreshed_at or "").strip() or None
    if source == "recommendations":
        return {
            "available": True,
            "source": "recommendations",
            "source_label": "Persisted recommendation workspace",
            "source_description": (
                "Recommendation lists and workflow state come from the local FinOps recommendation store, hydrated from cache heuristics, auxiliary export datasets, and the export-backed resource-cost bridge used for AKS visibility."
            ),
            "last_refreshed_at": last_refreshed_at,
        }
    return {
        "available": False,
        "source": "cache",
        "source_label": "Cached heuristic workspace",
        "source_description": (
            "Recommendation detail is currently coming directly from cache-backed heuristics because the persisted workspace is unavailable."
        ),
        "last_refreshed_at": last_refreshed_at,
    }


def _savings_summary_with_context(payload: dict[str, Any], *, source: str) -> dict[str, Any]:
    enriched = dict(payload)
    workspace = _get_recommendation_workspace_context(source=source)
    enriched["source"] = workspace["source"]
    enriched["source_label"] = workspace["source_label"]
    enriched["source_description"] = workspace["source_description"]
    enriched["last_refreshed_at"] = workspace["last_refreshed_at"]
    enriched["cost_context"] = _get_cost_context_payload()
    return enriched


def _get_cost_summary_payload() -> dict[str, Any]:
    export_summary = azure_finops_service.get_cost_summary()
    if export_summary:
        cached_summary = azure_cache.get_cost_summary()
        export_summary["recommendation_count"] = int(cached_summary.get("recommendation_count") or 0)
        export_summary["potential_monthly_savings"] = float(cached_summary.get("potential_monthly_savings") or 0.0)
        return _cost_summary_with_source(export_summary, source="exports")
    return _cost_summary_with_source(azure_cache.get_cost_summary(), source="cache")


def _get_cost_trend_payload() -> list[dict[str, Any]]:
    export_rows = azure_finops_service.get_cost_trend()
    if export_rows:
        return _cost_trend_with_source(export_rows, source="exports")
    return _cost_trend_with_source(azure_cache.get_cost_trend(), source="cache")


def _get_cost_breakdown_payload(group_by: str) -> list[dict[str, Any]]:
    export_rows = azure_finops_service.get_cost_breakdown(group_by)
    if export_rows:
        return _cost_breakdown_with_source(export_rows, source="exports")
    return _cost_breakdown_with_source(azure_cache.get_cost_breakdown(group_by), source="cache")


@router.get("/status")
async def get_status() -> dict[str, Any]:
    _ensure_azure_site()
    return _azure_status_with_exports(azure_cache.status())


@router.post("/refresh")
async def refresh_azure(
    background_tasks: BackgroundTasks,
    _admin: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    if azure_cache.status().get("refreshing"):
        return {**_azure_status_with_exports(azure_cache.status()), "message": "Refresh already in progress"}

    async def _run() -> None:
        await asyncio.get_running_loop().run_in_executor(None, azure_cache.trigger_refresh)

    background_tasks.add_task(_run)
    return {**_azure_status_with_exports(azure_cache.status()), "started": True}


@router.get("/overview")
async def get_overview() -> dict[str, Any]:
    _ensure_azure_site()
    return _azure_status_with_exports(azure_cache.get_overview())


@router.get("/search")
async def get_quick_search(search: str = Query(default="")) -> dict[str, Any]:
    _ensure_azure_site()
    return {"results": azure_cache.quick_search(search)}


@router.get("/subscriptions")
async def list_subscriptions() -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache._snapshot("subscriptions") or []


@router.get("/management-groups")
async def list_management_groups() -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache._snapshot("management_groups") or []


@router.get("/role-assignments")
async def list_role_assignments(
    search: str = Query(default=""),
    subscription_id: str = Query(default=""),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    rows = azure_cache._snapshot("role_assignments") or []
    search_lower = search.strip().lower()
    subscription_lower = subscription_id.strip().lower()
    result: list[dict[str, Any]] = []
    for item in rows:
        if subscription_lower and str(item.get("subscription_id") or "").lower() != subscription_lower:
            continue
        if search_lower:
            haystack = " ".join(
                [
                    str(item.get("scope") or ""),
                    str(item.get("principal_id") or ""),
                    str(item.get("principal_type") or ""),
                    str(item.get("role_name") or ""),
                ]
            ).lower()
            if search_lower not in haystack:
                continue
        result.append(item)
    return result


@router.get("/resources")
async def get_resources(
    search: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    resource_type: str = Query(default=""),
    location: str = Query(default=""),
    state: str = Query(default=""),
    tag_key: str = Query(default=""),
    tag_value: str = Query(default=""),
) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_cache.list_resources(
        search=search,
        subscription_id=subscription_id,
        resource_group=resource_group,
        resource_type=resource_type,
        location=location,
        state=state,
        tag_key=tag_key,
        tag_value=tag_value,
    )


@router.get("/vms")
async def get_virtual_machines(
    search: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    location: str = Query(default=""),
    state: str = Query(default=""),
    size: str = Query(default=""),
) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_cache.list_virtual_machines(
        search=search,
        subscription_id=subscription_id,
        resource_group=resource_group,
        location=location,
        state=state,
        size=size,
    )


@router.get("/virtual-desktops/removal-candidates")
async def get_virtual_desktop_removal_candidates(
    search: str = Query(default=""),
    removal_only: bool = Query(default=False),
    under_utilized_only: bool = Query(default=False),
    over_utilized_only: bool = Query(default=False),
) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_cache.list_virtual_desktop_removal_candidates(
        search=search,
        removal_only=removal_only,
        under_utilized_only=under_utilized_only,
        over_utilized_only=over_utilized_only,
    )


@router.get("/virtual-desktops/detail")
async def get_virtual_desktop_detail(resource_id: str = Query(default="")) -> dict[str, Any]:
    _ensure_azure_site()
    detail = azure_cache.get_virtual_desktop_detail(resource_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Virtual desktop was not found in the Azure cache")
    return detail


@router.get("/vms/detail")
async def get_virtual_machine_detail(resource_id: str = Query(default="")) -> AzureVirtualMachineDetailResponse:
    _ensure_azure_site()
    detail = azure_cache.get_virtual_machine_detail(resource_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Virtual machine was not found in the Azure cache")
    return AzureVirtualMachineDetailResponse.model_validate(detail)


@router.post("/vms/cost-export-jobs", status_code=202)
async def create_virtual_machine_cost_export_job(
    body: AzureVirtualMachineCostExportJobCreateRequest,
    session: dict[str, Any] = Depends(require_authenticated_user),
) -> AzureVirtualMachineCostExportJobResponse:
    _ensure_azure_site()
    try:
        job = azure_vm_export_jobs.create_job(
            recipient_email=str(session.get("email") or ""),
            requester_name=str(session.get("name") or ""),
            scope=body.scope,
            lookback_days=int(body.lookback_days),
            filters=body.filters.model_dump(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return AzureVirtualMachineCostExportJobResponse.model_validate(job)


@router.get("/vms/cost-export-jobs/{job_id}")
async def get_virtual_machine_cost_export_job(
    job_id: str,
    session: dict[str, Any] = Depends(require_authenticated_user),
) -> AzureVirtualMachineCostExportJobResponse:
    _ensure_azure_site()
    job = _ensure_export_job_access(job_id, session)
    return AzureVirtualMachineCostExportJobResponse.model_validate(job)


@router.get("/vms/cost-export-jobs/{job_id}/download")
async def download_virtual_machine_cost_export_job(
    job_id: str,
    session: dict[str, Any] = Depends(require_authenticated_user),
) -> FileResponse:
    _ensure_azure_site()
    job = _ensure_export_job_access(job_id, session)
    if str(job.get("status") or "") != "completed":
        raise HTTPException(status_code=409, detail="Azure VM export is not ready yet")

    file_path = str(job.get("file_path") or "")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=410, detail="Azure VM export file is no longer available")

    return FileResponse(
        file_path,
        filename=str(job.get("file_name") or os.path.basename(file_path)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/vms/coverage/export.csv")
async def export_virtual_machine_coverage_csv() -> FileResponse:
    _ensure_azure_site()
    rows = _vm_coverage_export_rows()
    now = datetime.now(timezone.utc)
    filename = f"azure_vm_coverage_{now.strftime('%Y%m%d_%H%M')}.csv"
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    fieldnames = ["SKU", "Region", "VMs", "Reserved Instances (RI)", "Needed / Excess"]
    try:
        writer = csv.DictWriter(tmp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        tmp.flush()
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(tmp.name, filename, "text/csv; charset=utf-8")


@router.get("/vms/coverage/export.xlsx")
async def export_virtual_machine_coverage_excel() -> FileResponse:
    _ensure_azure_site()
    rows = _vm_coverage_export_rows()
    now = datetime.now(timezone.utc)
    filename = f"azure_vm_coverage_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "VM Coverage"
    headers = ["SKU", "Region", "VMs", "Reserved Instances (RI)", "Needed / Excess"]
    for column_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for row_index, row in enumerate(rows, 2):
        ws.cell(row=row_index, column=1, value=row["SKU"])
        ws.cell(row=row_index, column=2, value=row["Region"])
        ws.cell(row=row_index, column=3, value=row["VMs"])
        ws.cell(row=row_index, column=4, value=row["Reserved Instances (RI)"])
        ws.cell(row=row_index, column=5, value=row["Needed / Excess"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 18

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(tmp.name)
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(
        tmp.name,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/vms/excess/export.csv")
async def export_virtual_machine_excess_csv() -> FileResponse:
    _ensure_azure_site()
    rows = _vm_excess_export_rows()
    now = datetime.now(timezone.utc)
    filename = f"azure_vm_ri_excess_{now.strftime('%Y%m%d_%H%M')}.csv"
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    fieldnames = ["SKU", "Region", "VMs", "Reserved Instances (RI)", "Excess", "Active Reservation Names"]
    try:
        writer = csv.DictWriter(tmp, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
        tmp.flush()
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(tmp.name, filename, "text/csv; charset=utf-8")


@router.get("/vms/excess/export.xlsx")
async def export_virtual_machine_excess_excel() -> FileResponse:
    _ensure_azure_site()
    rows = _vm_excess_export_rows()
    now = datetime.now(timezone.utc)
    filename = f"azure_vm_ri_excess_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "VM RI Excess"
    headers = ["SKU", "Region", "VMs", "Reserved Instances (RI)", "Excess", "Active Reservation Names"]
    for column_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for row_index, row in enumerate(rows, 2):
        ws.cell(row=row_index, column=1, value=row["SKU"])
        ws.cell(row=row_index, column=2, value=row["Region"])
        ws.cell(row=row_index, column=3, value=row["VMs"])
        ws.cell(row=row_index, column=4, value=row["Reserved Instances (RI)"])
        ws.cell(row=row_index, column=5, value=row["Excess"])
        ws.cell(row=row_index, column=6, value=row["Active Reservation Names"])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 56

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(tmp.name)
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(
        tmp.name,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/directory/users")
async def get_users(search: str = Query(default="")) -> list[dict[str, Any]]:
    _ensure_azure_or_primary_site()
    return azure_cache.list_directory_objects("users", search=search)


@router.get("/directory/groups")
async def get_groups(search: str = Query(default="")) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache.list_directory_objects("groups", search=search)


@router.get("/directory/enterprise-apps")
async def get_enterprise_apps(search: str = Query(default="")) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache.list_directory_objects("service_principals", search=search)


@router.get("/directory/app-registrations")
async def get_app_registrations(search: str = Query(default="")) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache.list_directory_objects("applications", search=search)


@router.get("/directory/roles")
async def get_directory_roles(search: str = Query(default="")) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache.list_directory_objects("directory_roles", search=search)


@router.get("/cost/summary")
async def get_cost_summary() -> dict[str, Any]:
    _ensure_azure_site()
    return _get_cost_summary_payload()


@router.get("/cost/trend")
async def get_cost_trend() -> list[dict[str, Any]]:
    _ensure_azure_site()
    return _get_cost_trend_payload()


@router.get("/cost/breakdown")
async def get_cost_breakdown(group_by: str = Query(default="service")) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return _get_cost_breakdown_payload(group_by)


@router.get("/finops/status")
async def get_finops_status(_session: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_finops_service.get_status()


@router.get("/finops/reconciliation")
async def get_finops_reconciliation(_session: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_finops_service.get_cost_reconciliation(azure_cache.get_cost_summary())


@router.get("/finops/validation")
async def get_finops_validation(_session: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    _ensure_azure_site()
    return _get_finops_validation_payload()


@router.get("/allocations/policy")
async def get_allocation_policy(_session: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_finops_service.get_allocation_policy()


@router.get("/allocations/status")
async def get_allocation_status(_session: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_finops_service.get_allocation_status()


@router.get("/allocations/rules")
async def get_allocation_rules(
    include_inactive: bool = Query(default=False),
    include_all_versions: bool = Query(default=False),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_finops_service.list_allocation_rules(
        include_inactive=include_inactive,
        include_all_versions=include_all_versions,
    )


@router.post("/allocations/rules")
async def create_or_update_allocation_rule(
    body: AzureAllocationRuleRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    try:
        return azure_finops_service.upsert_allocation_rule(
            rule_id=body.rule_id,
            name=body.name,
            description=body.description,
            rule_type=body.rule_type,
            target_dimension=body.target_dimension,
            priority=body.priority,
            enabled=body.enabled,
            condition=body.condition,
            allocation=body.allocation,
            actor_id=str(session.get("email") or ""),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/allocations/rules/{rule_id}/deactivate")
async def deactivate_allocation_rule(
    rule_id: str,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    payload = azure_finops_service.deactivate_allocation_rule(rule_id, actor_id=str(session.get("email") or ""))
    if payload is None:
        raise HTTPException(status_code=404, detail="Allocation rule was not found")
    return payload


@router.get("/allocations/runs")
async def get_allocation_runs(
    limit: int = Query(default=20, ge=1, le=100),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_finops_service.list_allocation_runs(limit=limit)


@router.post("/allocations/runs")
async def create_allocation_run(
    body: AzureAllocationRunRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    try:
        return azure_finops_service.run_allocation(
            actor_id=str(session.get("email") or ""),
            target_dimensions=body.target_dimensions,
            run_label=body.run_label,
            note=body.note,
            trigger_type="manual",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/allocations/runs/{run_id}")
async def get_allocation_run(
    run_id: str,
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, Any]:
    _ensure_azure_site()
    payload = azure_finops_service.get_allocation_run(run_id)
    if payload is None:
        raise HTTPException(status_code=404, detail="Allocation run was not found")
    return payload


@router.get("/allocations/runs/{run_id}/results")
async def get_allocation_run_results(
    run_id: str,
    target_dimension: str = Query(alias="dimension"),
    bucket_type: str = Query(default=""),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    if not azure_finops_service.get_allocation_run(run_id):
        raise HTTPException(status_code=404, detail="Allocation run was not found")
    try:
        return azure_finops_service.list_allocation_results(
            run_id,
            target_dimension=target_dimension,
            bucket_type=bucket_type,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/allocations/runs/{run_id}/residuals")
async def get_allocation_run_residuals(
    run_id: str,
    target_dimension: str = Query(alias="dimension"),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    if not azure_finops_service.get_allocation_run(run_id):
        raise HTTPException(status_code=404, detail="Allocation run was not found")
    try:
        return azure_finops_service.list_allocation_residuals(run_id, target_dimension=target_dimension)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/advisor")
async def get_advisor() -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_cache.get_advisor()


@router.get("/savings/summary")
async def get_savings_summary() -> AzureSavingsSummary:
    _ensure_azure_site()
    return AzureSavingsSummary.model_validate(_get_savings_summary_payload())


@router.get("/savings/opportunities")
async def get_savings_opportunities(
    search: str = Query(default=""),
    category: str = Query(default=""),
    opportunity_type: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    effort: str = Query(default=""),
    risk: str = Query(default=""),
    confidence: str = Query(default=""),
    quantified_only: bool = Query(default=False),
) -> list[AzureSavingsOpportunity]:
    _ensure_azure_site()
    rows = _list_filtered_savings_opportunities(
        search=search,
        category=category,
        opportunity_type=opportunity_type,
        subscription_id=subscription_id,
        resource_group=resource_group,
        effort=effort,
        risk=risk,
        confidence=confidence,
        quantified_only=quantified_only,
    )
    return [AzureSavingsOpportunity.model_validate(item) for item in rows]


@router.get("/recommendations/summary")
async def get_recommendations_summary(_session: dict[str, Any] = Depends(require_authenticated_user)) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    summary = azure_finops_service.get_recommendation_summary()
    if summary:
        return _savings_summary_with_context(summary, source="recommendations")
    return _savings_summary_with_context(
        {
            "currency": str(azure_cache.get_cost_summary().get("currency") or "USD"),
            "total_opportunities": 0,
            "quantified_opportunities": 0,
            "quantified_monthly_savings": 0.0,
            "quick_win_count": 0,
            "quick_win_monthly_savings": 0.0,
            "unquantified_opportunity_count": 0,
            "by_category": [],
            "by_opportunity_type": [],
            "by_effort": [],
            "by_risk": [],
            "by_confidence": [],
            "top_subscriptions": [],
            "top_resource_groups": [],
        },
        source="cache",
    )


@router.get("/recommendations/resource-cost-bridge")
async def get_recommendation_resource_cost_bridge(
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, Any]:
    _ensure_azure_site()
    cache_resources = azure_cache._snapshot("resources") or []
    return azure_finops_service.get_resource_cost_bridge_summary(cache_resources)


@router.get("/recommendations/aks-visibility")
async def get_recommendation_aks_visibility(
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    cache_resources = azure_cache._snapshot("resources") or []
    return azure_finops_service.list_aks_cost_visibility(cache_resources)


@router.get("/recommendations")
async def get_recommendations(
    search: str = Query(default=""),
    category: str = Query(default=""),
    opportunity_type: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    effort: str = Query(default=""),
    risk: str = Query(default=""),
    confidence: str = Query(default=""),
    quantified_only: bool = Query(default=False),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return _list_filtered_recommendations(
        search=search,
        category=category,
        opportunity_type=opportunity_type,
        subscription_id=subscription_id,
        resource_group=resource_group,
        effort=effort,
        risk=risk,
        confidence=confidence,
        quantified_only=quantified_only,
    )


@router.get("/recommendations/export.csv")
async def export_recommendations_csv(
    search: str = Query(default=""),
    category: str = Query(default=""),
    opportunity_type: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    effort: str = Query(default=""),
    risk: str = Query(default=""),
    confidence: str = Query(default=""),
    quantified_only: bool = Query(default=False),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> FileResponse:
    _ensure_azure_site()
    rows = _savings_export_rows(
        _list_filtered_recommendations(
            search=search,
            category=category,
            opportunity_type=opportunity_type,
            subscription_id=subscription_id,
            resource_group=resource_group,
            effort=effort,
            risk=risk,
            confidence=confidence,
            quantified_only=quantified_only,
        )
    )
    now = datetime.now(timezone.utc)
    filename = f"azure_recommendations_{now.strftime('%Y%m%d_%H%M')}.csv"
    headers = [
        "Title",
        "Category",
        "Opportunity Type",
        "Source",
        "Subscription",
        "Resource Group",
        "Location",
        "Resource Name",
        "Resource Type",
        "Current Monthly Cost",
        "Estimated Monthly Savings",
        "Quantified",
        "Effort",
        "Risk",
        "Confidence",
        "Estimate Basis",
        "Summary",
        "Recommended Steps",
        "Evidence",
        "Portal URL",
        "Follow Up Route",
    ]
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    try:
        writer = csv.DictWriter(tmp, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        tmp.flush()
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(tmp.name, filename, "text/csv; charset=utf-8")


@router.get("/recommendations/export.xlsx")
async def export_recommendations_excel(
    search: str = Query(default=""),
    category: str = Query(default=""),
    opportunity_type: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    effort: str = Query(default=""),
    risk: str = Query(default=""),
    confidence: str = Query(default=""),
    quantified_only: bool = Query(default=False),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> FileResponse:
    _ensure_azure_site()
    rows = _savings_export_rows(
        _list_filtered_recommendations(
            search=search,
            category=category,
            opportunity_type=opportunity_type,
            subscription_id=subscription_id,
            resource_group=resource_group,
            effort=effort,
            risk=risk,
            confidence=confidence,
            quantified_only=quantified_only,
        )
    )
    now = datetime.now(timezone.utc)
    filename = f"azure_recommendations_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    headers = [
        "Title",
        "Category",
        "Opportunity Type",
        "Source",
        "Subscription",
        "Resource Group",
        "Location",
        "Resource Name",
        "Resource Type",
        "Current Monthly Cost",
        "Estimated Monthly Savings",
        "Quantified",
        "Effort",
        "Risk",
        "Confidence",
        "Estimate Basis",
        "Summary",
        "Recommended Steps",
        "Evidence",
        "Portal URL",
        "Follow Up Route",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Recommendations"
    for column_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for row_index, row in enumerate(rows, 2):
        for column_index, header in enumerate(headers, 1):
            ws.cell(row=row_index, column=column_index, value=row.get(header, ""))

    for column_name, width in {
        "A": 36,
        "B": 14,
        "C": 26,
        "D": 12,
        "E": 24,
        "F": 24,
        "G": 14,
        "H": 28,
        "I": 34,
        "J": 18,
        "K": 22,
        "L": 12,
        "M": 10,
        "N": 10,
        "O": 12,
        "P": 42,
        "Q": 48,
        "R": 56,
        "S": 64,
        "T": 48,
        "U": 18,
    }.items():
        ws.column_dimensions[column_name].width = width

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(tmp.name)
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(
        tmp.name,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/recommendations/{recommendation_id}")
async def get_recommendation_detail(
    recommendation_id: str,
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    payload = azure_finops_service.get_recommendation(recommendation_id)
    if not payload:
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return payload


@router.get(
    "/recommendations/{recommendation_id}/actions",
    response_model=AzureRecommendationActionContractResponse,
)
async def get_recommendation_action_contract(
    recommendation_id: str,
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    payload = azure_finops_service.get_recommendation_action_contract(recommendation_id)
    if not payload:
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return payload


@router.get("/recommendations/{recommendation_id}/history")
async def get_recommendation_history(
    recommendation_id: str,
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    if not azure_finops_service.get_recommendation(recommendation_id):
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return azure_finops_service.list_recommendation_action_history(recommendation_id)


@router.post(
    "/recommendations/{recommendation_id}/actions/create-ticket",
    response_model=AzureRecommendationCreateTicketResponse,
)
async def create_recommendation_ticket(
    recommendation_id: str,
    body: AzureRecommendationCreateTicketRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()

    recommendation = azure_finops_service.get_recommendation(recommendation_id)
    if not recommendation:
        raise HTTPException(status_code=404, detail="Recommendation was not found")

    contract = azure_finops_service.get_recommendation_action_contract(recommendation_id)
    create_ticket_action = next(
        (
            item
            for item in (contract or {}).get("actions", [])
            if str(item.get("action_type") or "").lower() == "create_ticket"
        ),
        None,
    )
    if create_ticket_action is None:
        raise HTTPException(status_code=404, detail="Create-ticket action is not available for this recommendation")
    if not bool(create_ticket_action.get("can_execute")):
        blocked_reason = str(create_ticket_action.get("blocked_reason") or "").strip()
        raise HTTPException(
            status_code=409,
            detail=blocked_reason or "Create-ticket action is not currently available for this recommendation",
        )

    project_key = (body.project_key or AZURE_FINOPS_RECOMMENDATION_JIRA_PROJECT or "").strip().upper()
    issue_type = (body.issue_type or AZURE_FINOPS_RECOMMENDATION_JIRA_ISSUE_TYPE or "").strip() or "Task"
    summary = (body.summary or "").strip() or _default_recommendation_ticket_summary(recommendation)
    actor_id = str(session.get("email") or "")
    description = _build_recommendation_ticket_description(recommendation, operator_note=body.note)
    labels = [
        "azure-finops",
        str(recommendation.get("category") or "").strip().lower().replace(" ", "-"),
        str(recommendation.get("opportunity_type") or "").strip().lower().replace(" ", "-"),
    ]

    try:
        ctx = get_jira_write_context(session, shared_client=_jira_client)
        if ctx.is_fallback:
            description = append_fallback_actor_block(description, session)
        created_issue = ctx.client.create_issue(
            project_key=project_key,
            issue_type=issue_type,
            summary=summary,
            description=description,
            labels=labels,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("Failed to create Jira ticket for recommendation %s", recommendation_id)
        azure_finops_service.record_recommendation_action_event(
            recommendation_id,
            action_type="create_ticket",
            action_status="failed",
            actor_type="user",
            actor_id=actor_id,
            note=body.note,
            metadata={
                "project_key": project_key,
                "issue_type": issue_type,
                "summary": summary,
                "error": str(exc),
            },
        )
        raise HTTPException(status_code=502, detail="Could not create Jira ticket. Please try again in a moment.") from exc

    ticket_key = str(created_issue.get("key") or "").strip()
    jira_issue_id = str(created_issue.get("id") or "").strip()
    ticket_url = f"{JIRA_BASE_URL}/browse/{ticket_key}" if JIRA_BASE_URL and ticket_key else ""
    note = (body.note or "").strip() or f"Created Jira follow-up {ticket_key}."
    updated_recommendation = azure_finops_service.update_recommendation_action_state(
        recommendation_id,
        action_state="ticket_created",
        action_type="create_ticket",
        actor_type="user",
        actor_id=actor_id,
        note=note,
        metadata={
            "project_key": project_key,
            "issue_type": issue_type,
            "summary": summary,
            "ticket_key": ticket_key,
            "ticket_url": ticket_url,
            "jira_issue_id": jira_issue_id,
        },
    )
    if not updated_recommendation:
        raise HTTPException(status_code=404, detail="Recommendation was not found")

    return {
        "recommendation": updated_recommendation,
        "ticket_key": ticket_key,
        "ticket_url": ticket_url,
        "jira_issue_id": jira_issue_id,
        "project_key": project_key,
        "issue_type": issue_type,
        "summary": summary,
    }


@router.post(
    "/recommendations/{recommendation_id}/actions/send-alert",
    response_model=AzureRecommendationSendAlertResponse,
)
async def send_recommendation_alert(
    recommendation_id: str,
    body: AzureRecommendationSendAlertRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()

    recommendation = azure_finops_service.get_recommendation(recommendation_id)
    if not recommendation:
        raise HTTPException(status_code=404, detail="Recommendation was not found")

    contract = azure_finops_service.get_recommendation_action_contract(recommendation_id)
    send_alert_action = next(
        (
            item
            for item in (contract or {}).get("actions", [])
            if str(item.get("action_type") or "").lower() == "send_alert"
        ),
        None,
    )
    if send_alert_action is None:
        raise HTTPException(status_code=404, detail="Send-alert action is not available for this recommendation")
    if not bool(send_alert_action.get("can_execute")):
        blocked_reason = str(send_alert_action.get("blocked_reason") or "").strip()
        raise HTTPException(
            status_code=409,
            detail=blocked_reason or "Send-alert action is not currently available for this recommendation",
        )

    webhook_url = (body.teams_webhook_url or AZURE_FINOPS_RECOMMENDATION_TEAMS_WEBHOOK_URL or "").strip()
    if not webhook_url:
        raise HTTPException(
            status_code=422,
            detail="A Teams webhook URL is required. Configure a default webhook or provide an override.",
        )
    channel_label = (body.channel or AZURE_FINOPS_RECOMMENDATION_TEAMS_CHANNEL_LABEL or "FinOps").strip() or "FinOps"
    note = (body.note or "").strip()
    actor_id = str(session.get("email") or "")
    site_origin = f"https://{AZURE_APP_HOST}"

    try:
        await send_recommendation_teams_alert(
            webhook_url,
            recommendation,
            site_origin=site_origin,
            channel_label=channel_label,
            operator_note=note,
        )
    except Exception as exc:
        logger.exception("Failed to send Teams alert for recommendation %s", recommendation_id)
        azure_finops_service.record_recommendation_action_event(
            recommendation_id,
            action_type="send_alert",
            action_status="failed",
            actor_type="user",
            actor_id=actor_id,
            note=note,
            metadata={
                "channel": channel_label,
                "delivery_target": "teams_webhook",
                "error": str(exc),
            },
        )
        raise HTTPException(status_code=502, detail="Could not send Teams alert. Please try again in a moment.") from exc

    updated_recommendation = azure_finops_service.update_recommendation_action_state(
        recommendation_id,
        action_state="alert_sent",
        action_type="send_alert",
        actor_type="user",
        actor_id=actor_id,
        note=note or f"Sent Teams alert to {channel_label}.",
        metadata={
            "channel": channel_label,
            "delivery_target": "teams_webhook",
            "alert_status": "sent",
        },
    )
    if not updated_recommendation:
        raise HTTPException(status_code=404, detail="Recommendation was not found")

    return {
        "recommendation": updated_recommendation,
        "alert_status": "sent",
        "delivery_channel": channel_label,
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }


@router.post(
    "/recommendations/{recommendation_id}/actions/run-safe-script",
    response_model=AzureRecommendationRunSafeScriptResponse,
)
async def run_recommendation_safe_script(
    recommendation_id: str,
    body: AzureRecommendationRunSafeScriptRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()

    recommendation = azure_finops_service.get_recommendation(recommendation_id)
    if not recommendation:
        raise HTTPException(status_code=404, detail="Recommendation was not found")

    contract = azure_finops_service.get_recommendation_action_contract(recommendation_id)
    script_action = next(
        (
            item
            for item in (contract or {}).get("actions", [])
            if str(item.get("action_type") or "").lower() == "run_safe_script"
        ),
        None,
    )
    if script_action is None:
        raise HTTPException(status_code=404, detail="Run-safe-script action is not available for this recommendation")
    if not bool(script_action.get("can_execute")):
        blocked_reason = str(script_action.get("blocked_reason") or "").strip()
        raise HTTPException(
            status_code=409,
            detail=blocked_reason or "Run-safe-script action is not currently available for this recommendation",
        )

    actor_id = str(session.get("email") or "")
    try:
        payload = await asyncio.to_thread(
            azure_finops_service.run_recommendation_safe_hook,
            recommendation_id,
            hook_key=body.hook_key,
            dry_run=body.dry_run,
            actor_type="user",
            actor_id=actor_id,
            note=body.note,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    if not payload:
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return payload


@router.post("/recommendations/{recommendation_id}/dismiss")
async def dismiss_recommendation(
    recommendation_id: str,
    body: AzureRecommendationDismissRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    payload = azure_finops_service.dismiss_recommendation(
        recommendation_id,
        reason=body.reason,
        actor_type="user",
        actor_id=str(session.get("email") or ""),
    )
    if not payload:
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return payload


@router.post("/recommendations/{recommendation_id}/reopen")
async def reopen_recommendation(
    recommendation_id: str,
    body: AzureRecommendationReopenRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    payload = azure_finops_service.reopen_recommendation(
        recommendation_id,
        actor_type="user",
        actor_id=str(session.get("email") or ""),
        note=body.note,
    )
    if not payload:
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return payload


@router.post("/recommendations/{recommendation_id}/action-state")
async def update_recommendation_action_state(
    recommendation_id: str,
    body: AzureRecommendationActionStateRequest,
    session: dict[str, Any] = Depends(require_admin),
) -> dict[str, Any]:
    _ensure_azure_site()
    _refresh_finops_recommendations_from_cache()
    try:
        payload = azure_finops_service.update_recommendation_action_state(
            recommendation_id,
            action_state=body.action_state,
            action_type=body.action_type,
            actor_type="user",
            actor_id=str(session.get("email") or ""),
            note=body.note,
            metadata=body.metadata,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not payload:
        raise HTTPException(status_code=404, detail="Recommendation was not found")
    return payload


@router.get("/savings/export.csv")
async def export_savings_csv(
    search: str = Query(default=""),
    category: str = Query(default=""),
    opportunity_type: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    effort: str = Query(default=""),
    risk: str = Query(default=""),
    confidence: str = Query(default=""),
    quantified_only: bool = Query(default=False),
) -> FileResponse:
    _ensure_azure_site()
    rows = _savings_export_rows(
        _list_filtered_savings_opportunities(
            search=search,
            category=category,
            opportunity_type=opportunity_type,
            subscription_id=subscription_id,
            resource_group=resource_group,
            effort=effort,
            risk=risk,
            confidence=confidence,
            quantified_only=quantified_only,
        )
    )
    now = datetime.now(timezone.utc)
    filename = f"azure_savings_{now.strftime('%Y%m%d_%H%M')}.csv"
    headers = [
        "Title",
        "Category",
        "Opportunity Type",
        "Source",
        "Subscription",
        "Resource Group",
        "Location",
        "Resource Name",
        "Resource Type",
        "Current Monthly Cost",
        "Estimated Monthly Savings",
        "Quantified",
        "Effort",
        "Risk",
        "Confidence",
        "Estimate Basis",
        "Summary",
        "Recommended Steps",
        "Evidence",
        "Portal URL",
        "Follow Up Route",
    ]
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8")
    try:
        writer = csv.DictWriter(tmp, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
        tmp.flush()
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(tmp.name, filename, "text/csv; charset=utf-8")


@router.get("/savings/export.xlsx")
async def export_savings_excel(
    search: str = Query(default=""),
    category: str = Query(default=""),
    opportunity_type: str = Query(default=""),
    subscription_id: str = Query(default=""),
    resource_group: str = Query(default=""),
    effort: str = Query(default=""),
    risk: str = Query(default=""),
    confidence: str = Query(default=""),
    quantified_only: bool = Query(default=False),
) -> FileResponse:
    _ensure_azure_site()
    rows = _savings_export_rows(
        _list_filtered_savings_opportunities(
            search=search,
            category=category,
            opportunity_type=opportunity_type,
            subscription_id=subscription_id,
            resource_group=resource_group,
            effort=effort,
            risk=risk,
            confidence=confidence,
            quantified_only=quantified_only,
        )
    )
    now = datetime.now(timezone.utc)
    filename = f"azure_savings_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    headers = [
        "Title",
        "Category",
        "Opportunity Type",
        "Source",
        "Subscription",
        "Resource Group",
        "Location",
        "Resource Name",
        "Resource Type",
        "Current Monthly Cost",
        "Estimated Monthly Savings",
        "Quantified",
        "Effort",
        "Risk",
        "Confidence",
        "Estimate Basis",
        "Summary",
        "Recommended Steps",
        "Evidence",
        "Portal URL",
        "Follow Up Route",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Savings"
    for column_index, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=column_index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    for row_index, row in enumerate(rows, 2):
        for column_index, header in enumerate(headers, 1):
            ws.cell(row=row_index, column=column_index, value=row.get(header, ""))

    for column_name, width in {
        "A": 36,
        "B": 14,
        "C": 26,
        "D": 12,
        "E": 24,
        "F": 24,
        "G": 14,
        "H": 28,
        "I": 34,
        "J": 18,
        "K": 22,
        "L": 12,
        "M": 10,
        "N": 10,
        "O": 12,
        "P": 42,
        "Q": 48,
        "R": 56,
        "S": 64,
        "T": 48,
        "U": 18,
    }.items():
        ws.column_dimensions[column_name].width = width

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb.save(tmp.name)
    finally:
        tmp.close()
    return _build_vm_coverage_file_response(
        tmp.name,
        filename,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/storage")
async def get_storage(
    account_search: str = Query(default=""),
    disk_search: str = Query(default=""),
    snapshot_search: str = Query(default=""),
    disk_unattached_only: bool = Query(default=False),
) -> dict[str, Any]:
    _ensure_azure_site()
    payload = azure_cache.get_storage_summary(
        account_search=account_search,
        disk_search=disk_search,
        snapshot_search=snapshot_search,
        disk_unattached_only=disk_unattached_only,
    )
    payload["cost_context"] = _get_cost_context_payload()
    return payload


@router.get("/compute/optimization")
async def get_compute_optimization(idle_vm_search: str = Query(default="")) -> dict[str, Any]:
    _ensure_azure_site()
    payload = azure_cache.get_compute_optimization(idle_vm_search=idle_vm_search)
    payload["cost_context"] = _get_cost_context_payload()
    return payload


@router.get("/ai/models")
async def get_ai_models(_session: dict[str, Any] = Depends(require_authenticated_user)) -> list[dict[str, str]]:
    _ensure_azure_site()
    return [model.model_dump() for model in get_available_copilot_models()]


@router.get("/ai-costs/summary")
async def get_ai_cost_summary(
    lookback_days: int | None = Query(default=None),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, Any]:
    _ensure_azure_site()
    return azure_finops_service.get_ai_cost_summary(lookback_days=lookback_days) or {
        "lookback_days": int(lookback_days or 0),
        "usage_record_count": 0,
        "request_count": 0,
        "input_tokens": 0,
        "output_tokens": 0,
        "estimated_tokens": 0,
        "estimated_cost": 0.0,
        "currency": "USD",
        "top_model": "",
        "top_feature": "",
        "window_start": "",
        "window_end": "",
    }


@router.get("/ai-costs/trend")
async def get_ai_cost_trend(
    lookback_days: int | None = Query(default=None),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_finops_service.get_ai_cost_trend(lookback_days=lookback_days)


@router.get("/ai-costs/breakdown")
async def get_ai_cost_breakdown(
    group_by: str = Query(default="model"),
    lookback_days: int | None = Query(default=None),
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> list[dict[str, Any]]:
    _ensure_azure_site()
    return azure_finops_service.get_ai_cost_breakdown(group_by=group_by, lookback_days=lookback_days)


@router.post("/ai/cost-chat")
async def post_cost_chat(
    body: AzureCostChatRequest,
    _session: dict[str, Any] = Depends(require_authenticated_user),
) -> dict[str, Any]:
    _ensure_azure_site()
    available = get_available_copilot_models()
    if not available:
        raise HTTPException(
            status_code=400,
            detail="No AI model available for the Azure copilot. Ensure Ollama is running and the configured local model is pulled.",
        )
    available_ids = {model.id for model in available}
    model_id = body.model or get_default_copilot_model_id(available)
    if not model_id:
        raise HTTPException(
            status_code=400,
            detail="No AI model available for the Azure copilot. Ensure Ollama is running and the configured local model is pulled.",
        )
    if model_id not in available_ids:
        raise HTTPException(status_code=400, detail=f"Model '{model_id}' is not available from the active Ollama provider")
    if not body.question.strip():
        raise HTTPException(status_code=400, detail="Question is required")
    result = answer_azure_cost_question(
        body.question,
        _get_cost_copilot_context(),
        model_id,
        actor_type="user",
        actor_id="azure-portal",
    )
    return result.model_dump()
