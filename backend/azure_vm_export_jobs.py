"""Background job queue for Azure VM cost workbook exports."""

from __future__ import annotations

import asyncio
import html
import json
import logging
import os
import sqlite3
import threading
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from azure_cache import azure_cache
from config import AZURE_VM_EXPORT_RETENTION_DAYS, DATA_DIR
from email_service import send_email
from site_context import get_site_origin

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_WORKBOOK_SUBDIR = "azure_vm_exports"
_STATUS_VALUES = {"queued", "running", "completed", "failed"}


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _safe_excel_text(value: Any) -> str:
    text = str(value or "")
    if text and text[0] in ("=", "+", "-", "@"):
        return "\t" + text
    return text


class AzureVMExportJobManager:
    """SQLite-backed FIFO queue for long-running Azure VM cost exports."""

    def __init__(self, db_path: str | None = None) -> None:
        self._db_path = db_path or os.path.join(DATA_DIR, "azure_vm_export_jobs.db")
        self._file_dir = os.path.join(DATA_DIR, _WORKBOOK_SUBDIR)
        os.makedirs(os.path.dirname(self._db_path), exist_ok=True)
        os.makedirs(self._file_dir, exist_ok=True)
        self._lock = threading.Lock()
        self._bg_task: asyncio.Task[None] | None = None
        self._init_db()
        self._requeue_running_jobs()
        self._cleanup_expired_files()

    def _conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self._db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        with self._conn() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS export_jobs (
                    job_id TEXT PRIMARY KEY,
                    status TEXT NOT NULL,
                    recipient_email TEXT NOT NULL,
                    requester_name TEXT NOT NULL,
                    scope TEXT NOT NULL,
                    lookback_days INTEGER NOT NULL,
                    filters_json TEXT NOT NULL,
                    requested_at TEXT NOT NULL,
                    started_at TEXT,
                    completed_at TEXT,
                    progress_current INTEGER NOT NULL DEFAULT 0,
                    progress_total INTEGER NOT NULL DEFAULT 0,
                    progress_message TEXT NOT NULL DEFAULT '',
                    file_name TEXT,
                    file_path TEXT,
                    file_size INTEGER NOT NULL DEFAULT 0,
                    error TEXT,
                    notified_at TEXT,
                    notification_error TEXT
                )
                """
            )
            conn.commit()

    def _coerce_job(self, row: sqlite3.Row | dict[str, Any] | None) -> dict[str, Any] | None:
        if not row:
            return None
        payload = dict(row)
        try:
            filters = json.loads(str(payload.get("filters_json") or "{}"))
        except json.JSONDecodeError:
            filters = {}
        file_path = str(payload.get("file_path") or "")
        status = str(payload.get("status") or "queued")
        return {
            "job_id": str(payload.get("job_id") or ""),
            "status": status if status in _STATUS_VALUES else "queued",
            "recipient_email": str(payload.get("recipient_email") or ""),
            "requester_name": str(payload.get("requester_name") or ""),
            "scope": str(payload.get("scope") or "all"),
            "lookback_days": int(payload.get("lookback_days") or 30),
            "filters": filters,
            "requested_at": str(payload.get("requested_at") or ""),
            "started_at": payload.get("started_at") or None,
            "completed_at": payload.get("completed_at") or None,
            "progress_current": int(payload.get("progress_current") or 0),
            "progress_total": int(payload.get("progress_total") or 0),
            "progress_message": str(payload.get("progress_message") or ""),
            "file_name": payload.get("file_name") or None,
            "file_ready": bool(
                status == "completed" and file_path and os.path.exists(file_path)
            ),
            "error": payload.get("error") or None,
            "file_path": file_path or None,
            "file_size": int(payload.get("file_size") or 0),
            "notified_at": payload.get("notified_at") or None,
            "notification_error": payload.get("notification_error") or None,
        }

    def _update_job(self, job_id: str, **fields: Any) -> None:
        if not fields:
            return
        assignments = ", ".join(f"{key} = ?" for key in fields)
        values = list(fields.values()) + [job_id]
        with self._conn() as conn:
            conn.execute(f"UPDATE export_jobs SET {assignments} WHERE job_id = ?", values)
            conn.commit()

    def _requeue_running_jobs(self) -> None:
        with self._conn() as conn:
            conn.execute(
                """
                UPDATE export_jobs
                SET status = 'queued',
                    started_at = NULL,
                    completed_at = NULL,
                    progress_current = 0,
                    progress_message = 'Re-queued after restart',
                    error = ''
                WHERE status = 'running'
                """
            )
            conn.commit()

    def _cleanup_expired_files(self) -> None:
        cutoff = (_utcnow() - timedelta(days=AZURE_VM_EXPORT_RETENTION_DAYS)).isoformat()
        with self._conn() as conn:
            rows = conn.execute(
                """
                SELECT job_id, file_path
                FROM export_jobs
                WHERE completed_at IS NOT NULL
                  AND completed_at < ?
                  AND file_path IS NOT NULL
                  AND file_path != ''
                """,
                (cutoff,),
            ).fetchall()
            for row in rows:
                path = str(row["file_path"] or "")
                if path:
                    try:
                        os.unlink(path)
                    except FileNotFoundError:
                        pass
                conn.execute(
                    """
                    UPDATE export_jobs
                    SET file_path = NULL,
                        file_name = NULL,
                        file_size = 0
                    WHERE job_id = ?
                    """,
                    (str(row["job_id"]),),
                )
            conn.commit()

    def create_job(
        self,
        *,
        recipient_email: str,
        requester_name: str,
        scope: str,
        lookback_days: int,
        filters: dict[str, Any] | None,
    ) -> dict[str, Any]:
        if scope not in {"all", "filtered"}:
            raise ValueError("Unsupported VM export scope")
        if lookback_days not in {7, 30, 90}:
            raise ValueError("Unsupported VM export lookback window")

        job_id = uuid.uuid4().hex
        requested_at = _utcnow().isoformat()
        filters_json = json.dumps(filters or {})
        with self._conn() as conn:
            conn.execute(
                """
                INSERT INTO export_jobs (
                    job_id,
                    status,
                    recipient_email,
                    requester_name,
                    scope,
                    lookback_days,
                    filters_json,
                    requested_at,
                    progress_current,
                    progress_total,
                    progress_message
                )
                VALUES (?, 'queued', ?, ?, ?, ?, ?, ?, 0, 0, 'Queued')
                """,
                (
                    job_id,
                    recipient_email,
                    requester_name,
                    scope,
                    int(lookback_days),
                    filters_json,
                    requested_at,
                ),
            )
            conn.commit()
        return self.get_job(job_id) or {
            "job_id": job_id,
            "status": "queued",
            "recipient_email": recipient_email,
            "scope": scope,
            "lookback_days": int(lookback_days),
            "filters": filters or {},
            "requested_at": requested_at,
            "started_at": None,
            "completed_at": None,
            "progress_current": 0,
            "progress_total": 0,
            "progress_message": "Queued",
            "file_name": None,
            "file_ready": False,
            "error": None,
        }

    def get_job(self, job_id: str) -> dict[str, Any] | None:
        with self._conn() as conn:
            row = conn.execute("SELECT * FROM export_jobs WHERE job_id = ?", (job_id,)).fetchone()
        return self._coerce_job(row)

    def job_belongs_to(self, job_id: str, email: str, *, is_admin: bool = False) -> bool:
        job = self.get_job(job_id)
        if not job:
            return False
        if is_admin:
            return True
        return str(job.get("recipient_email") or "").lower() == str(email or "").lower()

    async def start_worker(self) -> None:
        if self._bg_task and not self._bg_task.done():
            return
        self._bg_task = asyncio.get_running_loop().create_task(self._background_loop())

    async def stop_worker(self) -> None:
        if not self._bg_task:
            return
        self._bg_task.cancel()
        try:
            await self._bg_task
        except asyncio.CancelledError:
            pass
        self._bg_task = None

    def _claim_next_job(self) -> dict[str, Any] | None:
        with self._lock:
            with self._conn() as conn:
                row = conn.execute(
                    """
                    SELECT *
                    FROM export_jobs
                    WHERE status = 'queued'
                    ORDER BY requested_at ASC
                    LIMIT 1
                    """
                ).fetchone()
                if not row:
                    return None
                job_id = str(row["job_id"])
                conn.execute(
                    """
                    UPDATE export_jobs
                    SET status = 'running',
                        started_at = ?,
                        completed_at = NULL,
                        progress_current = 0,
                        progress_total = 0,
                        progress_message = 'Starting export',
                        error = ''
                    WHERE job_id = ?
                    """,
                    (_utcnow().isoformat(), job_id),
                )
                conn.commit()
            return self.get_job(job_id)

    def _write_workbook(self, job_id: str, payload: dict[str, Any], scope: str, lookback_days: int) -> tuple[str, str, int]:
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "VM Cost Summary"
        detail_sheet = workbook.create_sheet("Associated Resource Costs")
        shared_sheet = workbook.create_sheet("Shared Cost Candidates")

        summary_headers = [
            "VM Name",
            "Resource ID",
            "Subscription",
            "Resource Group",
            "Region",
            "Size",
            "Power State",
            "Lookback Days",
            "Currency",
            "VM Only Cost",
            "Direct Attached Resource Cost",
            "Direct Total Cost",
            "Priced Resource Count",
            "Unpriced Resource Count",
            "Shared Candidate Count",
            "Shared Candidate Amount",
            "Cost Status",
        ]
        detail_headers = [
            "VM Name",
            "VM Resource ID",
            "Associated Resource Name",
            "Associated Resource ID",
            "Relationship",
            "Type",
            "Subscription",
            "Resource Group",
            "Region",
            "Cost",
            "Currency",
            "Pricing Status",
            "Pricing Error",
        ]
        shared_headers = [
            "Resource Name",
            "Resource ID",
            "Type",
            "Subscription",
            "Resource Group",
            "Region",
            "Cost",
            "Currency",
            "Candidate VM Count",
            "Candidate VM Names",
            "Reason",
        ]

        self._populate_sheet(
            summary_sheet,
            summary_headers,
            payload.get("summary_rows") or [],
            [
                "vm_name",
                "resource_id",
                "subscription",
                "resource_group",
                "region",
                "size",
                "power_state",
                "lookback_days",
                "currency",
                "vm_only_cost",
                "direct_attached_resource_cost",
                "direct_total_cost",
                "priced_resource_count",
                "unpriced_resource_count",
                "shared_candidate_count",
                "shared_candidate_amount",
                "cost_status",
            ],
            {
                "A": 28,
                "B": 72,
                "C": 24,
                "D": 24,
                "E": 14,
                "F": 18,
                "G": 16,
                "H": 14,
                "I": 10,
                "J": 14,
                "K": 22,
                "L": 16,
                "M": 18,
                "N": 19,
                "O": 19,
                "P": 20,
                "Q": 28,
            },
        )
        self._populate_sheet(
            detail_sheet,
            detail_headers,
            payload.get("detail_rows") or [],
            [
                "vm_name",
                "vm_resource_id",
                "associated_resource_name",
                "associated_resource_id",
                "relationship",
                "type",
                "subscription",
                "resource_group",
                "region",
                "cost",
                "currency",
                "pricing_status",
                "pricing_error",
            ],
            {
                "A": 26,
                "B": 70,
                "C": 30,
                "D": 70,
                "E": 18,
                "F": 34,
                "G": 22,
                "H": 22,
                "I": 14,
                "J": 14,
                "K": 10,
                "L": 18,
                "M": 44,
            },
        )
        self._populate_sheet(
            shared_sheet,
            shared_headers,
            payload.get("shared_rows") or [],
            [
                "resource_name",
                "resource_id",
                "resource_type",
                "subscription",
                "resource_group",
                "region",
                "cost",
                "currency",
                "candidate_vm_count",
                "candidate_vm_names",
                "reason",
            ],
            {
                "A": 28,
                "B": 70,
                "C": 34,
                "D": 24,
                "E": 24,
                "F": 14,
                "G": 14,
                "H": 10,
                "I": 18,
                "J": 40,
                "K": 30,
            },
        )

        timestamp = _utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"azure_vm_costs_{scope}_{lookback_days}d_{timestamp}.xlsx"
        file_path = os.path.join(self._file_dir, f"{job_id}_{filename}")
        workbook.save(file_path)
        file_size = os.path.getsize(file_path)
        return file_path, filename, file_size

    def _populate_sheet(
        self,
        worksheet: Any,
        headers: list[str],
        rows: list[dict[str, Any]],
        keys: list[str],
        column_widths: dict[str, int],
    ) -> None:
        for column_index, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT

        for row_index, row in enumerate(rows, 2):
            for column_index, key in enumerate(keys, 1):
                value = row.get(key)
                if isinstance(value, str):
                    value = _safe_excel_text(value)
                worksheet.cell(row=row_index, column=column_index, value=value)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

    def _process_job(self, job_id: str) -> None:
        job = self.get_job(job_id)
        if not job:
            return

        def progress(current: int, total: int, message: str) -> None:
            self._update_job(
                job_id,
                progress_current=int(current),
                progress_total=int(total),
                progress_message=message,
            )

        try:
            payload = azure_cache.build_virtual_machine_cost_export(
                scope=str(job.get("scope") or "all"),
                filters=job.get("filters") or {},
                lookback_days=int(job.get("lookback_days") or 30),
                progress_callback=progress,
            )
            latest = self.get_job(job_id) or {}
            total = max(
                int(latest.get("progress_total") or 0),
                int(payload.get("vm_count") or 0),
                1,
            )
            current = min(int(latest.get("progress_current") or 0), max(0, total - 1))
            self._update_job(
                job_id,
                progress_current=current,
                progress_total=total,
                progress_message="Writing Excel workbook",
            )
            file_path, filename, file_size = self._write_workbook(
                job_id,
                payload,
                str(job.get("scope") or "all"),
                int(job.get("lookback_days") or 30),
            )
            self._update_job(
                job_id,
                status="completed",
                completed_at=_utcnow().isoformat(),
                progress_current=max(total, 1),
                progress_total=max(total, 1),
                progress_message="Export ready",
                file_name=filename,
                file_path=file_path,
                file_size=int(file_size),
                error="",
            )
        except Exception as exc:
            logger.exception("Azure VM export job %s failed", job_id)
            self._update_job(
                job_id,
                status="failed",
                completed_at=_utcnow().isoformat(),
                progress_message="Export failed",
                error=str(exc),
            )

    async def _send_notification(self, job: dict[str, Any]) -> None:
        recipient = str(job.get("recipient_email") or "").strip()
        if not recipient:
            return
        job_id = str(job.get("job_id") or "")
        scope = "current filters" if str(job.get("scope") or "") == "filtered" else "all cached VMs"
        download_url = f"{get_site_origin('azure')}/api/azure/vms/cost-export-jobs/{job_id}/download"

        if str(job.get("status") or "") == "completed":
            subject = "Azure VM cost export is ready"
            html_body = (
                "<p>Your Azure VM cost export is ready.</p>"
                f"<p><strong>Scope:</strong> {html.escape(scope)}<br>"
                f"<strong>Lookback:</strong> {int(job.get('lookback_days') or 30)} days<br>"
                f"<strong>Workbook:</strong> {html.escape(str(job.get('file_name') or 'azure_vm_costs.xlsx'))}</p>"
                f"<p><a href=\"{html.escape(download_url)}\">Download the workbook</a></p>"
            )
        else:
            subject = "Azure VM cost export failed"
            html_body = (
                "<p>Your Azure VM cost export could not be completed.</p>"
                f"<p><strong>Scope:</strong> {html.escape(scope)}<br>"
                f"<strong>Lookback:</strong> {int(job.get('lookback_days') or 30)} days</p>"
                f"<p><strong>Error:</strong> {html.escape(str(job.get('error') or 'Unknown error'))}</p>"
            )

        sent = await send_email([recipient], subject, html_body)
        self._update_job(
            job_id,
            notified_at=_utcnow().isoformat(),
            notification_error="" if sent else "Failed to send completion email",
        )

    async def _background_loop(self) -> None:
        while True:
            try:
                await asyncio.get_running_loop().run_in_executor(None, self._cleanup_expired_files)
                job = await asyncio.get_running_loop().run_in_executor(None, self._claim_next_job)
                if not job:
                    await asyncio.sleep(3)
                    continue
                job_id = str(job.get("job_id") or "")
                await asyncio.get_running_loop().run_in_executor(None, self._process_job, job_id)
                latest = await asyncio.get_running_loop().run_in_executor(None, self.get_job, job_id)
                if latest and not latest.get("notified_at"):
                    await self._send_notification(latest)
            except asyncio.CancelledError:
                raise
            except Exception:
                logger.exception("Azure VM export background loop failed")
                await asyncio.sleep(5)


azure_vm_export_jobs = AzureVMExportJobManager()
