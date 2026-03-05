import json, sys, re
from datetime import datetime, timezone

sys.path.insert(0, "/home/gallison/workspace/altlassian/.venv/lib/python3.12/site-packages")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("Loading data...")
with open("/home/gallison/workspace/altlassian/oit_all_issues_full.json") as f:
    issues = json.load(f)
print(f"Loaded {len(issues)} tickets")

def parse_dt(s):
    if not s: return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except:
        return None

def fmt_dt(s):
    dt = parse_dt(s)
    if dt: return dt.strftime("%Y-%m-%d %H:%M:%S")
    return ""

def extract_text(node):
    """Recursively extract plain text from ADF (Atlassian Document Format)."""
    if not node or not isinstance(node, dict):
        return ""
    text = ""
    if node.get("type") == "text":
        text += node.get("text", "")
    for child in node.get("content", []):
        text += extract_text(child)
    return text

def safe_get(d, *keys, default=""):
    """Safely navigate nested dicts."""
    curr = d
    for k in keys:
        if isinstance(curr, dict):
            curr = curr.get(k)
        else:
            return default
        if curr is None:
            return default
    return curr if curr is not None else default

def extract_sla_info(sla_field):
    """Extract SLA completion/ongoing info from JSM SLA fields."""
    if not sla_field or not isinstance(sla_field, dict):
        return "", ""
    
    # completedCycles
    completed = sla_field.get("completedCycles", [])
    ongoing = sla_field.get("ongoingCycle", {})
    
    status = ""
    elapsed = ""
    
    if completed and isinstance(completed, list) and len(completed) > 0:
        last = completed[-1]
        breached = last.get("breached", False)
        status = "BREACHED" if breached else "Met"
        e = last.get("elapsedTime", {})
        if e:
            millis = e.get("millis", 0)
            if millis:
                elapsed = f"{millis/1000/60:.1f}min"
    elif ongoing and isinstance(ongoing, dict):
        breached = ongoing.get("breached", False)
        paused = ongoing.get("paused", False)
        status = "BREACHED" if breached else ("Paused" if paused else "Running")
        e = ongoing.get("elapsedTime", {})
        if e:
            millis = e.get("millis", 0)
            if millis:
                elapsed = f"{millis/1000/60:.1f}min"
    
    return status, elapsed

now = datetime.now(timezone.utc)

# ── Define columns ──
print("Building spreadsheet...")
wb = Workbook()
ws = wb.active
ws.title = "All Tickets"

columns = [
    ("Key", lambda i, f: i.get("key", "")),
    ("ID", lambda i, f: i.get("id", "")),
    ("Summary", lambda i, f: f.get("summary", "")),
    ("Issue Type", lambda i, f: safe_get(f, "issuetype", "name")),
    ("Status", lambda i, f: safe_get(f, "status", "name")),
    ("Status Category", lambda i, f: safe_get(f, "status", "statusCategory", "name")),
    ("Priority", lambda i, f: safe_get(f, "priority", "name")),
    ("Resolution", lambda i, f: safe_get(f, "resolution", "name")),
    ("Assignee", lambda i, f: safe_get(f, "assignee", "displayName")),
    ("Assignee Email", lambda i, f: safe_get(f, "assignee", "emailAddress")),
    ("Reporter", lambda i, f: safe_get(f, "reporter", "displayName")),
    ("Reporter Email", lambda i, f: safe_get(f, "reporter", "emailAddress")),
    ("Creator", lambda i, f: safe_get(f, "creator", "displayName")),
    ("Creator Email", lambda i, f: safe_get(f, "creator", "emailAddress")),
    ("Created", lambda i, f: fmt_dt(f.get("created"))),
    ("Updated", lambda i, f: fmt_dt(f.get("updated"))),
    ("Resolved", lambda i, f: fmt_dt(f.get("resolutiondate"))),
    ("Status Category Changed", lambda i, f: fmt_dt(f.get("statuscategorychangedate"))),
    ("Last Viewed", lambda i, f: fmt_dt(f.get("lastViewed"))),
    ("Description", lambda i, f: extract_text(f.get("description", {}))[:2000]),
    ("Request Type", lambda i, f: safe_get(f, "customfield_11102", "requestType", "name")),
    ("Current Status (JSM)", lambda i, f: safe_get(f, "customfield_11102", "currentStatus", "status")),
    ("Work Category", lambda i, f: safe_get(f, "customfield_11239", "value") if isinstance(f.get("customfield_11239"), dict) else str(f.get("customfield_11239", "") or "")),
    ("SLT Projects", lambda i, f: safe_get(f, "customfield_11117", "value") if isinstance(f.get("customfield_11117"), dict) else str(f.get("customfield_11117", "") or "")),
    ("Applications", lambda i, f: safe_get(f, "customfield_11301", "value") if isinstance(f.get("customfield_11301"), dict) else str(f.get("customfield_11301", "") or "")),
    ("Source", lambda i, f: safe_get(f, "customfield_11249", "value") if isinstance(f.get("customfield_11249"), dict) else str(f.get("customfield_11249", "") or "")),
    ("Request Language", lambda i, f: safe_get(f, "customfield_11217", "value") if isinstance(f.get("customfield_11217"), dict) else str(f.get("customfield_11217", "") or "")),
    ("Steps To Re-Create", lambda i, f: extract_text(f.get("customfield_11121", {}))[:1000]),
    ("Business Priority", lambda i, f: ", ".join(x.get("name","") for x in (f.get("customfield_10200") or []) if isinstance(x, dict))),
    ("Date of First Response", lambda i, f: fmt_dt(f.get("customfield_10001"))),
    ("% Front-end", lambda i, f: f.get("customfield_11323", "")),
    ("% Back-end", lambda i, f: f.get("customfield_11324", "")),
    ("Checklist Progress %", lambda i, f: f.get("customfield_11271", "")),
    ("Work Ratio", lambda i, f: f.get("workratio", "")),
    ("Labels", lambda i, f: ", ".join(f.get("labels", []) or [])),
    ("Components", lambda i, f: ", ".join(c.get("name","") for c in (f.get("components") or []) if isinstance(c, dict))),
    ("Project Key", lambda i, f: safe_get(f, "project", "key")),
    ("Project Name", lambda i, f: safe_get(f, "project", "name")),
    ("Vote Count", lambda i, f: safe_get(f, "votes", "votes")),
    ("Watch Count", lambda i, f: safe_get(f, "watches", "watchCount")),
    ("Comment Count", lambda i, f: safe_get(f, "comment", "total")),
    ("Open Forms", lambda i, f: f.get("customfield_11207", "")),
    ("Submitted Forms", lambda i, f: f.get("customfield_11208", "")),
    ("Locked Forms", lambda i, f: f.get("customfield_11209", "")),
    ("Total Forms", lambda i, f: f.get("customfield_11210", "")),
]

# SLA columns (extract from the 4 SLA fields)
sla_fields = [
    ("customfield_11266", "SLA: First Response"),
    ("customfield_11264", "SLA: Resolution"),
    ("customfield_11267", "SLA: Close After Resolution"),
    ("customfield_11268", "SLA: Review Normal Change"),
]

for cf_id, sla_name in sla_fields:
    columns.append((f"{sla_name} — Status", lambda i, f, _id=cf_id: extract_sla_info(f.get(_id))[0]))
    columns.append((f"{sla_name} — Elapsed", lambda i, f, _id=cf_id: extract_sla_info(f.get(_id))[1]))

# Computed columns
def calc_ttr_hours(i, f):
    c = parse_dt(f.get("created"))
    r = parse_dt(f.get("resolutiondate"))
    if c and r:
        h = (r - c).total_seconds() / 3600
        return round(max(h, 0), 2)
    return ""

def calc_age_days(i, f):
    c = parse_dt(f.get("created"))
    if c:
        return round((now - c).total_seconds() / 86400, 1)
    return ""

def calc_staleness_days(i, f):
    u = parse_dt(f.get("updated"))
    if u:
        return round((now - u).total_seconds() / 86400, 1)
    return ""

def is_open(i, f):
    cat = safe_get(f, "status", "statusCategory", "name")
    return "Open" if cat not in ("Done",) else "Closed"

columns.extend([
    ("Calendar TTR (hours)", calc_ttr_hours),
    ("Age (days)", calc_age_days),
    ("Days Since Update", calc_staleness_days),
    ("Open/Closed", is_open),
])

# ── Write headers ──
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="2F5496")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for c, (name, _) in enumerate(columns, 1):
    cell = ws.cell(1, c, name)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# ── Write data ──
print("Writing rows...")
for idx, iss in enumerate(issues):
    row = idx + 2
    fields = iss.get("fields", {})
    for c, (_, extractor) in enumerate(columns, 1):
        try:
            val = extractor(iss, fields)
        except Exception:
            val = ""
        # Sanitize for Excel
        if isinstance(val, str) and len(val) > 32000:
            val = val[:32000] + "...[truncated]"
        ws.cell(row, c, val)
    
    if idx % 2000 == 0 and idx > 0:
        print(f"  Written {idx:,} rows...")

print(f"  Written {len(issues):,} rows total")

# ── Auto-filter ──
ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(issues)+1}"

# ── Freeze top row ──
ws.freeze_panes = "A2"

# ── Column widths ──
print("Setting column widths...")
# Set reasonable widths based on column name
for c, (name, _) in enumerate(columns, 1):
    if "Description" in name or "Steps" in name:
        ws.column_dimensions[get_column_letter(c)].width = 50
    elif "Summary" in name:
        ws.column_dimensions[get_column_letter(c)].width = 45
    elif "Email" in name:
        ws.column_dimensions[get_column_letter(c)].width = 30
    elif name in ("Key", "ID"):
        ws.column_dimensions[get_column_letter(c)].width = 12
    elif "Date" in name or "Created" in name or "Updated" in name or "Resolved" in name:
        ws.column_dimensions[get_column_letter(c)].width = 20
    else:
        ws.column_dimensions[get_column_letter(c)].width = 16

# ── Save ──
outpath = "/home/gallison/workspace/altlassian/reports/2026-03-03_0825/OIT_All_Tickets_Full_Data_2026-03-03_0825.xlsx"
print(f"Saving to {outpath}...")
wb.save(outpath)
print(f"Done! {len(issues):,} tickets x {len(columns)} columns saved.")
