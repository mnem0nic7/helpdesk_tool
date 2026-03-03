import json, sys, os, re
from datetime import datetime, timezone, timedelta
from collections import Counter, defaultdict
import statistics

sys.path.insert(0, "/home/gallison/workspace/altlassian/.venv/lib/python3.12/site-packages")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

REPORT_DIR = "/home/gallison/workspace/altlassian/reports/2026-03-03_0825"
REPORT_TS = "2026-03-03 08:25 UTC"

# ── Helpers ──
def parse_dt(s):
    if not s: return None
    try: return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except: return None

def fmt_dt(s):
    dt = parse_dt(s)
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""

def extract_text(node):
    if not node or not isinstance(node, dict): return ""
    text = ""
    if node.get("type") == "text": text += node.get("text", "")
    for child in node.get("content", []): text += extract_text(child)
    return text

def safe_get(d, *keys, default=""):
    curr = d
    for k in keys:
        if isinstance(curr, dict): curr = curr.get(k)
        else: return default
        if curr is None: return default
    return curr if curr is not None else default

def extract_sla_info(sla_field):
    if not sla_field or not isinstance(sla_field, dict): return "", ""
    completed = sla_field.get("completedCycles", [])
    ongoing = sla_field.get("ongoingCycle", {})
    status = elapsed = ""
    if completed and isinstance(completed, list) and len(completed) > 0:
        last = completed[-1]
        status = "BREACHED" if last.get("breached", False) else "Met"
        e = last.get("elapsedTime", {})
        if e and e.get("millis"): elapsed = f"{e['millis']/1000/60:.1f}min"
    elif ongoing and isinstance(ongoing, dict):
        breached = ongoing.get("breached", False)
        paused = ongoing.get("paused", False)
        status = "BREACHED" if breached else ("Paused" if paused else "Running")
        e = ongoing.get("elapsedTime", {})
        if e and e.get("millis"): elapsed = f"{e['millis']/1000/60:.1f}min"
    return status, elapsed

def hours_between(a, b):
    if a and b: return max((b - a).total_seconds() / 3600.0, 0)
    return None

def percentile(data, p):
    if not data: return None
    s = sorted(data)
    k = (len(s) - 1) * p / 100.0
    f = int(k); c = f + 1
    if c >= len(s): return s[f]
    return s[f] + (k - f) * (s[c] - s[f])

def fmt_hours(h):
    if h is None: return "N/A"
    if h < 1: return f"{h*60:.0f}m"
    if h < 24: return f"{h:.1f}h"
    return f"{h/24:.1f}d"

# ── Exclusion rule ──
def is_excluded(iss):
    """Returns True if ticket should be excluded per baseline rules."""
    fields = iss.get("fields", {})
    labels = fields.get("labels") or []
    for l in labels:
        if "oasisdev" in l.lower():
            return True
    summary = (fields.get("summary") or "").lower()
    if "oasisdev" in summary:
        return True
    return False

# ── Status mapping ──
ACTIVE_STATUSES = {"new", "open", "assigned", "in progress", "work in progress", "investigating"}
PAUSED_STATUSES = {"waiting for customer", "waiting for support", "pending", "pending customer",
                   "pending vendor", "scheduled", "on hold", "awaiting approval", "waiting for approval"}
TERMINAL_STATUSES = {"resolved", "closed", "done", "cancelled", "declined", "canceled"}

def map_status_bucket(status_name):
    s = (status_name or "").strip().lower()
    if s in TERMINAL_STATUSES: return "Terminal"
    if s in PAUSED_STATUSES: return "Paused"
    if s in ACTIVE_STATUSES: return "Active"
    for p in PAUSED_STATUSES:
        if p in s or s in p: return "Paused"
    for a in ACTIVE_STATUSES:
        if a in s or s in a: return "Active"
    return "Active"

# ── Load data ──
print("Loading data...")
with open("/home/gallison/workspace/altlassian/oit_all_issues_full.json") as f:
    all_issues_raw = json.load(f)

print(f"Loaded {len(all_issues_raw)} total tickets")

now = datetime.now(timezone.utc)

# ── Tag exclusions ──
excluded_count = sum(1 for i in all_issues_raw if is_excluded(i))
included_issues = [i for i in all_issues_raw if not is_excluded(i)]
print(f"Excluded (oasisdev): {excluded_count}")
print(f"Included for metrics: {len(included_issues)}")

# ══════════════════════════════════════════════
# REPORT 1: Full Data Export (ALL tickets with exclusion flag)
# ══════════════════════════════════════════════
print("\n=== Generating Full Data Export ===")
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "All Tickets"

hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="2F5496")
excl_fill = PatternFill("solid", fgColor="FFC7CE")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

sla_fields = [
    ("customfield_11266", "SLA: First Response"),
    ("customfield_11264", "SLA: Resolution"),
    ("customfield_11267", "SLA: Close After Resolution"),
    ("customfield_11268", "SLA: Review Normal Change"),
]

columns = [
    ("Key", lambda i, f: i.get("key", "")),
    ("ID", lambda i, f: i.get("id", "")),
    ("Excluded", lambda i, f: "YES — oasisdev" if is_excluded(i) else ""),
    ("Summary", lambda i, f: f.get("summary", "")),
    ("Issue Type", lambda i, f: safe_get(f, "issuetype", "name")),
    ("Status", lambda i, f: safe_get(f, "status", "name")),
    ("Status Category", lambda i, f: safe_get(f, "status", "statusCategory", "name")),
    ("Priority", lambda i, f: safe_get(f, "priority", "name")),
    ("Resolution", lambda i, f: safe_get(f, "resolution", "name")),
    ("Assignee", lambda i, f: safe_get(f, "assignee", "displayName")),
    ("Assignee Email", lambda i, f: safe_get(f, "assignee", "emailAddress")),
    ("Reporter", lambda i, f: safe_get(f, "reporter", "displayName")),
    ("Reporter Email", lambda i, f: safe_get(f, "reporter", "emailAddress")),
    ("Creator", lambda i, f: safe_get(f, "creator", "displayName")),
    ("Creator Email", lambda i, f: safe_get(f, "creator", "emailAddress")),
    ("Created", lambda i, f: fmt_dt(f.get("created"))),
    ("Updated", lambda i, f: fmt_dt(f.get("updated"))),
    ("Resolved", lambda i, f: fmt_dt(f.get("resolutiondate"))),
    ("Status Category Changed", lambda i, f: fmt_dt(f.get("statuscategorychangedate"))),
    ("Last Viewed", lambda i, f: fmt_dt(f.get("lastViewed"))),
    ("Description", lambda i, f: extract_text(f.get("description", {}))[:2000]),
    ("Request Type", lambda i, f: safe_get(f, "customfield_11102", "requestType", "name")),
    ("Current Status (JSM)", lambda i, f: safe_get(f, "customfield_11102", "currentStatus", "status")),
    ("Work Category", lambda i, f: safe_get(f, "customfield_11239", "value") if isinstance(f.get("customfield_11239"), dict) else str(f.get("customfield_11239", "") or "")),
    ("SLT Projects", lambda i, f: safe_get(f, "customfield_11117", "value") if isinstance(f.get("customfield_11117"), dict) else str(f.get("customfield_11117", "") or "")),
    ("Applications", lambda i, f: safe_get(f, "customfield_11301", "value") if isinstance(f.get("customfield_11301"), dict) else str(f.get("customfield_11301", "") or "")),
    ("Source", lambda i, f: safe_get(f, "customfield_11249", "value") if isinstance(f.get("customfield_11249"), dict) else str(f.get("customfield_11249", "") or "")),
    ("Request Language", lambda i, f: safe_get(f, "customfield_11217", "value") if isinstance(f.get("customfield_11217"), dict) else str(f.get("customfield_11217", "") or "")),
    ("Steps To Re-Create", lambda i, f: extract_text(f.get("customfield_11121", {}))[:1000]),
    ("Business Priority", lambda i, f: ", ".join(x.get("name","") for x in (f.get("customfield_10200") or []) if isinstance(x, dict))),
    ("Date of First Response", lambda i, f: fmt_dt(f.get("customfield_10001"))),
    ("% Front-end", lambda i, f: f.get("customfield_11323", "")),
    ("% Back-end", lambda i, f: f.get("customfield_11324", "")),
    ("Checklist Progress %", lambda i, f: f.get("customfield_11271", "")),
    ("Work Ratio", lambda i, f: f.get("workratio", "")),
    ("Labels", lambda i, f: ", ".join(f.get("labels", []) or [])),
    ("Components", lambda i, f: ", ".join(c.get("name","") for c in (f.get("components") or []) if isinstance(c, dict))),
    ("Project Key", lambda i, f: safe_get(f, "project", "key")),
    ("Project Name", lambda i, f: safe_get(f, "project", "name")),
    ("Vote Count", lambda i, f: safe_get(f, "votes", "votes")),
    ("Watch Count", lambda i, f: safe_get(f, "watches", "watchCount")),
    ("Comment Count", lambda i, f: safe_get(f, "comment", "total")),
    ("Open Forms", lambda i, f: f.get("customfield_11207", "")),
    ("Submitted Forms", lambda i, f: f.get("customfield_11208", "")),
    ("Locked Forms", lambda i, f: f.get("customfield_11209", "")),
    ("Total Forms", lambda i, f: f.get("customfield_11210", "")),
]

for cf_id, sla_name in sla_fields:
    columns.append((f"{sla_name} — Status", lambda i, f, _id=cf_id: extract_sla_info(f.get(_id))[0]))
    columns.append((f"{sla_name} — Elapsed", lambda i, f, _id=cf_id: extract_sla_info(f.get(_id))[1]))

def calc_ttr(i, f):
    c = parse_dt(f.get("created")); r = parse_dt(f.get("resolutiondate"))
    if c and r: return round(max((r-c).total_seconds()/3600, 0), 2)
    return ""
def calc_age(i, f):
    c = parse_dt(f.get("created"))
    return round((now-c).total_seconds()/86400, 1) if c else ""
def calc_stale(i, f):
    u = parse_dt(f.get("updated"))
    return round((now-u).total_seconds()/86400, 1) if u else ""
def is_open_flag(i, f):
    return "Open" if safe_get(f, "status", "statusCategory", "name") not in ("Done",) else "Closed"

columns.extend([
    ("Calendar TTR (hours)", calc_ttr),
    ("Age (days)", calc_age),
    ("Days Since Update", calc_stale),
    ("Open/Closed", is_open_flag),
])

# Write headers
for c, (name, _) in enumerate(columns, 1):
    cell = ws1.cell(1, c, name)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# Write rows (ALL tickets, excluded ones flagged)
for idx, iss in enumerate(all_issues_raw):
    row = idx + 2
    fields = iss.get("fields", {})
    excluded = is_excluded(iss)
    for c, (_, extractor) in enumerate(columns, 1):
        try: val = extractor(iss, fields)
        except: val = ""
        if isinstance(val, str) and len(val) > 32000: val = val[:32000] + "...[truncated]"
        cell = ws1.cell(row, c, val)
        if excluded and c == 3:
            cell.fill = excl_fill
    if idx % 5000 == 0 and idx > 0:
        print(f"  Full data: {idx:,} rows...")

ws1.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(all_issues_raw)+1}"
ws1.freeze_panes = "A2"

for c, (name, _) in enumerate(columns, 1):
    if "Description" in name or "Steps" in name: ws1.column_dimensions[get_column_letter(c)].width = 50
    elif "Summary" in name: ws1.column_dimensions[get_column_letter(c)].width = 45
    elif "Email" in name: ws1.column_dimensions[get_column_letter(c)].width = 30
    elif name in ("Key", "ID", "Excluded"): ws1.column_dimensions[get_column_letter(c)].width = 14
    elif "Date" in name or "Created" in name or "Updated" in name or "Resolved" in name: ws1.column_dimensions[get_column_letter(c)].width = 20
    else: ws1.column_dimensions[get_column_letter(c)].width = 16

path1 = os.path.join(REPORT_DIR, "OIT_All_Tickets_Full_Data_2026-03-03_0825.xlsx")
wb1.save(path1)
print(f"Full data saved: {path1}")

# ══════════════════════════════════════════════
# REPORT 2: Baseline Metrics (EXCLUDED tickets removed)
# ══════════════════════════════════════════════
print("\n=== Generating Baseline Metrics (excl. oasisdev) ===")

# Parse included tickets
tickets = []
for iss in included_issues:
    fields = iss.get("fields", {})
    key = iss.get("key", "")
    created = parse_dt(fields.get("created"))
    resolved = parse_dt(fields.get("resolutiondate"))
    updated = parse_dt(fields.get("updated"))
    status_name = safe_get(fields, "status", "name")
    status_cat = safe_get(fields, "status", "statusCategory", "name")
    priority_name = safe_get(fields, "priority", "name")
    issue_type = safe_get(fields, "issuetype", "name")
    assignee_name = safe_get(fields, "assignee", "displayName")
    reporter_name = safe_get(fields, "reporter", "displayName")
    summary = fields.get("summary", "")
    request_type = safe_get(fields, "customfield_11102", "requestType", "name")
    
    status_bucket = map_status_bucket(status_name)
    is_open = status_bucket != "Terminal"
    calendar_ttr_h = hours_between(created, resolved)
    age_h = hours_between(created, now) if created else None
    staleness_h = hours_between(updated, now) if updated else None
    month_str = created.strftime("%Y-%m") if created else "Unknown"
    
    tickets.append({
        "key": key, "summary": summary, "issue_type": issue_type,
        "request_type": request_type, "status": status_name,
        "status_bucket": status_bucket, "status_cat": status_cat,
        "priority": priority_name, "assignee": assignee_name,
        "reporter": reporter_name, "created": created, "resolved": resolved,
        "updated": updated, "is_open": is_open, "calendar_ttr_h": calendar_ttr_h,
        "age_h": age_h, "staleness_h": staleness_h, "month": month_str,
    })

total = len(tickets)
open_tickets = [t for t in tickets if t["is_open"]]
resolved_tickets = [t for t in tickets if not t["is_open"]]
resolved_with_ttr = [t for t in resolved_tickets if t["calendar_ttr_h"] is not None]
ttr_values = [t["calendar_ttr_h"] for t in resolved_with_ttr]

monthly_created = Counter(t["month"] for t in tickets)
monthly_resolved = Counter(t["month"] for t in resolved_tickets)
all_months = sorted(set(list(monthly_created.keys()) + list(monthly_resolved.keys())))

type_counts = Counter(t["issue_type"] for t in tickets)
priority_counts = Counter(t["priority"] for t in tickets)
priority_open = Counter(t["priority"] for t in open_tickets)
request_type_counts = Counter(t["request_type"] for t in tickets if t["request_type"])

# TTR stats
ttr_stats = {}
if ttr_values:
    ttr_stats = {
        "count": len(ttr_values), "mean": statistics.mean(ttr_values),
        "median": statistics.median(ttr_values), "p90": percentile(ttr_values, 90),
        "p95": percentile(ttr_values, 95), "min": min(ttr_values), "max": max(ttr_values),
    }

# Backlog aging
age_buckets = {"0-2d": 0, "3-7d": 0, "8-14d": 0, "15-30d": 0, "30+d": 0}
for t in open_tickets:
    if t["age_h"] is not None:
        days = t["age_h"] / 24
        if days <= 2: age_buckets["0-2d"] += 1
        elif days <= 7: age_buckets["3-7d"] += 1
        elif days <= 14: age_buckets["8-14d"] += 1
        elif days <= 30: age_buckets["15-30d"] += 1
        else: age_buckets["30+d"] += 1

stale = [t for t in open_tickets if t["staleness_h"] and t["staleness_h"] > 7*24]
stale_by_assignee = Counter(t["assignee"] or "(Unassigned)" for t in stale)
stale_by_status = Counter(t["status"] for t in stale)

ttr_by_type = defaultdict(list)
ttr_by_priority = defaultdict(list)
ttr_by_month = defaultdict(list)
for t in resolved_with_ttr:
    ttr_by_type[t["issue_type"]].append(t["calendar_ttr_h"])
    ttr_by_priority[t["priority"]].append(t["calendar_ttr_h"])
    m = t["resolved"].strftime("%Y-%m") if t["resolved"] else t["month"]
    ttr_by_month[m].append(t["calendar_ttr_h"])

assignee_resolved = Counter(t["assignee"] or "(Unassigned)" for t in resolved_tickets)
assignee_open = Counter(t["assignee"] or "(Unassigned)" for t in open_tickets)
assignee_ttr = defaultdict(list)
for t in resolved_with_ttr:
    assignee_ttr[t["assignee"] or "(Unassigned)"].append(t["calendar_ttr_h"])

# ── Styles ──
section_font = Font(bold=True, size=13, color="1F3864")
subsection_font = Font(bold=True, size=11, color="2F5496")
metric_font = Font(bold=True, size=11)
warn_fill = PatternFill("solid", fgColor="FFC7CE")
good_fill = PatternFill("solid", fgColor="C6EFCE")

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value: mx = max(mx, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx

wb2 = Workbook()
prio_order = ["Highest", "High", "Medium", "Low", "Lowest"]

# ── Sheet 1: Executive Summary ──
ws = wb2.active
ws.title = "Executive Summary"
ws.sheet_properties.tabColor = "2F5496"

r = 1
ws.cell(r, 1, "OIT HELPDESK BASELINE METRICS REPORT").font = Font(bold=True, size=16, color="1F3864")
ws.cell(2, 1, f"Report generated: {REPORT_TS}").font = Font(italic=True, size=10, color="666666")
ws.cell(3, 1, f"Data range: {all_months[0]} to {all_months[-1]} ({len(all_months)} months)").font = Font(italic=True, size=10, color="666666")
ws.cell(4, 1, "Baseline framework: Helpdesk Metrics Baseline v1.0").font = Font(italic=True, size=10, color="666666")

r = 6
ws.cell(r, 1, "EXCLUSION RULES APPLIED").font = Font(bold=True, size=12, color="C00000")
r += 1
ws.cell(r, 1, f"Excluded: {excluded_count} tickets tagged 'oasisdev' (labels or summary)")
ws.cell(r, 1).font = Font(color="C00000")
ws.cell(r+1, 1, f"Total in Jira: {len(all_issues_raw):,} | After exclusions: {total:,}")
ws.cell(r+1, 1).font = Font(color="C00000")

r = 10
ws.cell(r, 1, "A. HEADLINE NUMBERS").font = section_font
r += 1
headlines = [
    ("Total Tickets (after exclusions)", f"{total:,}"),
    ("Open (Backlog)", f"{len(open_tickets):,}"),
    ("Resolved/Closed", f"{len(resolved_tickets):,}"),
    ("Resolution Rate", f"{len(resolved_tickets)/total*100:.1f}%"),
    ("", ""),
    ("Calendar TTR — Median", fmt_hours(ttr_stats.get("median"))),
    ("Calendar TTR — Mean", fmt_hours(ttr_stats.get("mean"))),
    ("Calendar TTR — P90", fmt_hours(ttr_stats.get("p90"))),
    ("Calendar TTR — P95", fmt_hours(ttr_stats.get("p95"))),
    ("", ""),
    ("Open Backlog Aging — 0-2 days", str(age_buckets["0-2d"])),
    ("Open Backlog Aging — 3-7 days", str(age_buckets["3-7d"])),
    ("Open Backlog Aging — 8-14 days", str(age_buckets["8-14d"])),
    ("Open Backlog Aging — 15-30 days", str(age_buckets["15-30d"])),
    ("Open Backlog Aging — 30+ days", str(age_buckets["30+d"])),
    ("", ""),
    ("Stale Tickets (no update 7+ days)", str(len(stale))),
]
for label, val in headlines:
    if not label: r += 1; continue
    ws.cell(r, 1, label).font = metric_font
    ws.cell(r, 2, val).alignment = Alignment(horizontal="right")
    r += 1

r += 2
ws.cell(r, 1, "B. DATA AVAILABILITY NOTES").font = section_font
r += 1
for note in [
    "Calendar TTR: Available. Active TTR: NOT available (needs changelog).",
    "FRT: NOT available (needs changelog). Time to Assign: NOT available.",
    "Reopen/Reassignment Rate: NOT available. CSAT: NOT in data.",
    "SLA Compliance: No targets set. Escalation: NOT available.",
    "RECOMMENDATION: Enable changelog export for Active TTR, FRT, Reopen, Escalation.",
]:
    ws.cell(r, 1, note).font = Font(size=9, color="444444")
    r += 1

ws.column_dimensions["A"].width = 55
ws.column_dimensions["B"].width = 20

# ── Sheet 2: Demand & Intake ──
ws2 = wb2.create_sheet("Demand & Intake")
ws2.sheet_properties.tabColor = "00B050"
r = 1
ws2.cell(r, 1, "DEMAND & INTAKE METRICS (excl. oasisdev)").font = section_font
r += 2

ws2.cell(r, 1, "Ticket Type Mix").font = subsection_font; r += 1
ws2.cell(r, 1, "Issue Type"); ws2.cell(r, 2, "Count"); ws2.cell(r, 3, "% of Total")
style_header_row(ws2, r, 3); r += 1
for typ, cnt in type_counts.most_common():
    ws2.cell(r, 1, typ); ws2.cell(r, 2, cnt); ws2.cell(r, 3, f"{cnt/total*100:.1f}%"); r += 1

r += 1
ws2.cell(r, 1, "Priority Distribution").font = subsection_font; r += 1
ws2.cell(r, 1, "Priority"); ws2.cell(r, 2, "Total"); ws2.cell(r, 3, "Open"); ws2.cell(r, 4, "% of Total")
style_header_row(ws2, r, 4); r += 1
for p in prio_order:
    if p in priority_counts:
        ws2.cell(r, 1, p); ws2.cell(r, 2, priority_counts[p]); ws2.cell(r, 3, priority_open.get(p, 0)); ws2.cell(r, 4, f"{priority_counts[p]/total*100:.1f}%"); r += 1
for p, cnt in priority_counts.most_common():
    if p not in prio_order:
        ws2.cell(r, 1, p or "(None)"); ws2.cell(r, 2, cnt); ws2.cell(r, 3, priority_open.get(p, 0)); ws2.cell(r, 4, f"{cnt/total*100:.1f}%"); r += 1

if request_type_counts:
    r += 1
    ws2.cell(r, 1, "Top JSM Request Types").font = subsection_font; r += 1
    ws2.cell(r, 1, "Request Type"); ws2.cell(r, 2, "Count"); ws2.cell(r, 3, "% of Total")
    style_header_row(ws2, r, 3); r += 1
    for rt, cnt in request_type_counts.most_common(25):
        ws2.cell(r, 1, rt); ws2.cell(r, 2, cnt); ws2.cell(r, 3, f"{cnt/total*100:.1f}%"); r += 1

r += 1
ws2.cell(r, 1, "Monthly Ticket Volume").font = subsection_font; r += 1
hdr_r = r
ws2.cell(r, 1, "Month"); ws2.cell(r, 2, "Created"); ws2.cell(r, 3, "Resolved"); ws2.cell(r, 4, "Net Flow"); ws2.cell(r, 5, "Cumulative Backlog Delta")
style_header_row(ws2, r, 5); r += 1
cum = 0
for m in all_months:
    c = monthly_created.get(m, 0); res = monthly_resolved.get(m, 0); nf = c - res; cum += nf
    ws2.cell(r, 1, m); ws2.cell(r, 2, c); ws2.cell(r, 3, res); ws2.cell(r, 4, nf); ws2.cell(r, 5, cum)
    if nf > 0: ws2.cell(r, 4).fill = warn_fill
    elif nf < 0: ws2.cell(r, 4).fill = good_fill
    r += 1

chart = LineChart(); chart.title = "Monthly Volume (excl. oasisdev)"; chart.y_axis.title = "Tickets"; chart.width = 28; chart.height = 14
data = Reference(ws2, min_col=2, max_col=3, min_row=hdr_r, max_row=r-1)
cats = Reference(ws2, min_col=1, min_row=hdr_r+1, max_row=r-1)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
ws2.add_chart(chart, "G2")
auto_width(ws2)

# ── Sheet 3: Backlog Health ──
ws3 = wb2.create_sheet("Backlog Health")
ws3.sheet_properties.tabColor = "FFC000"
r = 1
ws3.cell(r, 1, "BACKLOG HEALTH (excl. oasisdev)").font = section_font; r += 2
ws3.cell(r, 1, "Total Open"); ws3.cell(r, 2, len(open_tickets)); r += 1

ws3.cell(r, 1, "Aging Distribution").font = subsection_font; r += 1
ws3.cell(r, 1, "Age Bucket"); ws3.cell(r, 2, "Count"); ws3.cell(r, 3, "% of Open")
style_header_row(ws3, r, 3); r += 1
age_start = r
for bucket, cnt in age_buckets.items():
    ws3.cell(r, 1, bucket); ws3.cell(r, 2, cnt); ws3.cell(r, 3, f"{cnt/max(len(open_tickets),1)*100:.1f}%")
    if bucket == "30+d" and cnt > 0: ws3.cell(r, 2).fill = warn_fill
    r += 1

pie = PieChart(); pie.title = "Open Backlog by Age"; pie.width = 16; pie.height = 12
pdata = Reference(ws3, min_col=2, min_row=age_start-1, max_row=r-1)
pcats = Reference(ws3, min_col=1, min_row=age_start, max_row=r-1)
pie.add_data(pdata, titles_from_data=True); pie.set_categories(pcats)
pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
ws3.add_chart(pie, "E3")

r += 1
ws3.cell(r, 1, "Open by Status").font = subsection_font; r += 1
ws3.cell(r, 1, "Status"); ws3.cell(r, 2, "Count"); ws3.cell(r, 3, "Status Bucket")
style_header_row(ws3, r, 3); r += 1
for s, cnt in Counter(t["status"] for t in open_tickets).most_common():
    ws3.cell(r, 1, s); ws3.cell(r, 2, cnt); ws3.cell(r, 3, map_status_bucket(s)); r += 1

r += 1
ws3.cell(r, 1, "Open by Priority").font = subsection_font; r += 1
ws3.cell(r, 1, "Priority"); ws3.cell(r, 2, "Count")
style_header_row(ws3, r, 2); r += 1
for p in prio_order:
    if priority_open.get(p, 0) > 0: ws3.cell(r, 1, p); ws3.cell(r, 2, priority_open[p]); r += 1
for p, cnt in priority_open.most_common():
    if p not in prio_order and cnt > 0: ws3.cell(r, 1, p or "(None)"); ws3.cell(r, 2, cnt); r += 1

if stale:
    r += 1
    ws3.cell(r, 1, f"Stale Tickets (7+ days): {len(stale)}").font = subsection_font; r += 1
    ws3.cell(r, 1, "Key"); ws3.cell(r, 2, "Summary"); ws3.cell(r, 3, "Status"); ws3.cell(r, 4, "Assignee"); ws3.cell(r, 5, "Priority"); ws3.cell(r, 6, "Days Since Update"); ws3.cell(r, 7, "Age (days)")
    style_header_row(ws3, r, 7); r += 1
    for t in sorted(stale, key=lambda x: x["staleness_h"] or 0, reverse=True):
        ws3.cell(r, 1, t["key"]); ws3.cell(r, 2, t["summary"][:80]); ws3.cell(r, 3, t["status"])
        ws3.cell(r, 4, t["assignee"] or "(Unassigned)"); ws3.cell(r, 5, t["priority"])
        ws3.cell(r, 6, f"{t['staleness_h']/24:.0f}" if t["staleness_h"] else "")
        ws3.cell(r, 7, f"{t['age_h']/24:.0f}" if t["age_h"] else ""); r += 1
auto_width(ws3)

# ── Sheet 4: Resolution Speed ──
ws4 = wb2.create_sheet("Resolution Speed")
ws4.sheet_properties.tabColor = "FF0000"
r = 1
ws4.cell(r, 1, "RESOLUTION SPEED — Calendar TTR (excl. oasisdev)").font = section_font
ws4.cell(2, 1, "Calendar TTR = Created → Resolved. Active TTR requires changelog.").font = Font(italic=True, size=9, color="666666")
r = 4

ws4.cell(r, 1, "Overall Calendar TTR").font = subsection_font; r += 1
ws4.cell(r, 1, "Statistic"); ws4.cell(r, 2, "Value (hours)"); ws4.cell(r, 3, "Human-readable")
style_header_row(ws4, r, 3); r += 1
for label, key in [("Count","count"),("Mean","mean"),("Median","median"),("P90","p90"),("P95","p95"),("Min","min"),("Max","max")]:
    ws4.cell(r, 1, label); v = ttr_stats.get(key)
    if key == "count": ws4.cell(r, 2, v); ws4.cell(r, 3, f"{v:,}")
    else: ws4.cell(r, 2, round(v,2) if v else "N/A"); ws4.cell(r, 3, fmt_hours(v))
    r += 1

r += 1
ws4.cell(r, 1, "Calendar TTR by Issue Type").font = subsection_font; r += 1
ws4.cell(r, 1, "Issue Type"); ws4.cell(r, 2, "Count"); ws4.cell(r, 3, "Median (h)"); ws4.cell(r, 4, "Mean (h)"); ws4.cell(r, 5, "P90 (h)"); ws4.cell(r, 6, "P95 (h)")
style_header_row(ws4, r, 6); r += 1
for typ in sorted(ttr_by_type.keys()):
    vals = ttr_by_type[typ]
    ws4.cell(r, 1, typ); ws4.cell(r, 2, len(vals))
    ws4.cell(r, 3, round(statistics.median(vals),2)); ws4.cell(r, 4, round(statistics.mean(vals),2))
    ws4.cell(r, 5, round(percentile(vals,90),2)); ws4.cell(r, 6, round(percentile(vals,95),2)); r += 1

r += 1
ws4.cell(r, 1, "Calendar TTR by Priority").font = subsection_font; r += 1
ws4.cell(r, 1, "Priority"); ws4.cell(r, 2, "Count"); ws4.cell(r, 3, "Median (h)"); ws4.cell(r, 4, "Mean (h)"); ws4.cell(r, 5, "P90 (h)"); ws4.cell(r, 6, "P95 (h)")
style_header_row(ws4, r, 6); r += 1
for p in prio_order + [p for p in ttr_by_priority if p not in prio_order]:
    if p in ttr_by_priority:
        vals = ttr_by_priority[p]
        ws4.cell(r, 1, p or "(None)"); ws4.cell(r, 2, len(vals))
        ws4.cell(r, 3, round(statistics.median(vals),2)); ws4.cell(r, 4, round(statistics.mean(vals),2))
        ws4.cell(r, 5, round(percentile(vals,90),2)); ws4.cell(r, 6, round(percentile(vals,95),2)); r += 1

r += 1
ws4.cell(r, 1, "Monthly Calendar TTR Trend").font = subsection_font; r += 1
hdr_r4 = r
ws4.cell(r, 1, "Month"); ws4.cell(r, 2, "Resolved"); ws4.cell(r, 3, "Median TTR (h)"); ws4.cell(r, 4, "P90 TTR (h)"); ws4.cell(r, 5, "P95 TTR (h)")
style_header_row(ws4, r, 5); r += 1
for m in all_months:
    if m in ttr_by_month and ttr_by_month[m]:
        vals = ttr_by_month[m]
        ws4.cell(r, 1, m); ws4.cell(r, 2, len(vals))
        ws4.cell(r, 3, round(statistics.median(vals),2)); ws4.cell(r, 4, round(percentile(vals,90),2)); ws4.cell(r, 5, round(percentile(vals,95),2)); r += 1

ch4 = LineChart(); ch4.title = "Monthly TTR Median & P90"; ch4.y_axis.title = "Hours"; ch4.width = 28; ch4.height = 14
d4 = Reference(ws4, min_col=3, max_col=4, min_row=hdr_r4, max_row=r-1)
c4 = Reference(ws4, min_col=1, min_row=hdr_r4+1, max_row=r-1)
ch4.add_data(d4, titles_from_data=True); ch4.set_categories(c4)
ws4.add_chart(ch4, "G4")

r += 2
ws4.cell(r, 1, "Resolution Time Distribution").font = subsection_font; r += 1
ws4.cell(r, 1, "Bucket"); ws4.cell(r, 2, "Count"); ws4.cell(r, 3, "% of Resolved"); ws4.cell(r, 4, "Cumulative %")
style_header_row(ws4, r, 4); r += 1
ttr_dist = {"< 1h":0,"1-4h":0,"4-8h":0,"8-24h":0,"1-3d":0,"3-7d":0,"7-14d":0,"14-30d":0,"30+d":0}
for h in ttr_values:
    if h<1: ttr_dist["< 1h"]+=1
    elif h<4: ttr_dist["1-4h"]+=1
    elif h<8: ttr_dist["4-8h"]+=1
    elif h<24: ttr_dist["8-24h"]+=1
    elif h<72: ttr_dist["1-3d"]+=1
    elif h<168: ttr_dist["3-7d"]+=1
    elif h<336: ttr_dist["7-14d"]+=1
    elif h<720: ttr_dist["14-30d"]+=1
    else: ttr_dist["30+d"]+=1
cum=0
for bucket, cnt in ttr_dist.items():
    cum+=cnt; ws4.cell(r,1,bucket); ws4.cell(r,2,cnt)
    ws4.cell(r,3,f"{cnt/max(len(ttr_values),1)*100:.1f}%"); ws4.cell(r,4,f"{cum/max(len(ttr_values),1)*100:.1f}%"); r+=1
auto_width(ws4)

# ── Sheet 5: Assignee Workload ──
ws5 = wb2.create_sheet("Assignee Workload")
ws5.sheet_properties.tabColor = "7030A0"
r = 1
ws5.cell(r, 1, "ASSIGNEE WORKLOAD (excl. oasisdev)").font = section_font; r += 2
ws5.cell(r, 1, "Assignee"); ws5.cell(r, 2, "Total Resolved"); ws5.cell(r, 3, "Currently Open")
ws5.cell(r, 4, "Median TTR (h)"); ws5.cell(r, 5, "P90 TTR (h)"); ws5.cell(r, 6, "Stale Open")
style_header_row(ws5, r, 6); r += 1

all_assignees = set(list(assignee_resolved.keys()) + list(assignee_open.keys()))
adata = []
for a in all_assignees:
    adata.append((a, assignee_resolved.get(a,0), assignee_open.get(a,0),
        statistics.median(assignee_ttr[a]) if assignee_ttr.get(a) else None,
        percentile(assignee_ttr[a], 90) if assignee_ttr.get(a) else None,
        stale_by_assignee.get(a, 0)))
adata.sort(key=lambda x: x[1], reverse=True)
for a, res, opn, med, p90, stl in adata[:30]:
    ws5.cell(r,1,a); ws5.cell(r,2,res); ws5.cell(r,3,opn)
    ws5.cell(r,4,round(med,2) if med else "N/A"); ws5.cell(r,5,round(p90,2) if p90 else "N/A")
    ws5.cell(r,6,stl)
    if stl > 0: ws5.cell(r,6).fill = warn_fill
    r += 1
auto_width(ws5)

# ── Sheet 6: Status Mapping ──
ws6 = wb2.create_sheet("Status Mapping")
ws6.sheet_properties.tabColor = "00B0F0"
r = 1
ws6.cell(r, 1, "STATUS MAPPING — Baseline Framework").font = section_font; r += 2
ws6.cell(r, 1, "Jira Status"); ws6.cell(r, 2, "Baseline Bucket"); ws6.cell(r, 3, "Clock"); ws6.cell(r, 4, "Count (All)"); ws6.cell(r, 5, "Count (Open)")
style_header_row(ws6, r, 5); r += 1
all_statuses = Counter(t["status"] for t in tickets)
open_statuses = Counter(t["status"] for t in open_tickets)
for s, cnt in all_statuses.most_common():
    bucket = map_status_bucket(s)
    ws6.cell(r,1,s); ws6.cell(r,2,bucket)
    ws6.cell(r,3,"Running" if bucket=="Active" else ("Paused" if bucket=="Paused" else "Stopped"))
    ws6.cell(r,4,cnt); ws6.cell(r,5,open_statuses.get(s,0))
    if bucket=="Active": ws6.cell(r,3).fill = good_fill
    elif bucket=="Paused": ws6.cell(r,3).fill = PatternFill("solid", fgColor="FFEB9C")
    r += 1
auto_width(ws6)

# ── Sheet 7: Exclusion Log ──
ws7 = wb2.create_sheet("Exclusion Log")
ws7.sheet_properties.tabColor = "C00000"
r = 1
ws7.cell(r, 1, "EXCLUDED TICKETS").font = section_font; r += 1
ws7.cell(r, 1, f"Total excluded: {excluded_count} tickets matching 'oasisdev'").font = Font(color="C00000"); r += 2
ws7.cell(r, 1, "Key"); ws7.cell(r, 2, "Summary"); ws7.cell(r, 3, "Status"); ws7.cell(r, 4, "Created"); ws7.cell(r, 5, "Labels"); ws7.cell(r, 6, "Exclusion Reason")
style_header_row(ws7, r, 6); r += 1
for iss in all_issues_raw:
    if is_excluded(iss):
        f = iss.get("fields", {})
        labels = f.get("labels") or []
        reason = []
        for l in labels:
            if "oasisdev" in l.lower(): reason.append(f"label:{l}")
        if "oasisdev" in (f.get("summary") or "").lower(): reason.append("summary match")
        ws7.cell(r, 1, iss.get("key",""))
        ws7.cell(r, 2, (f.get("summary") or "")[:80])
        ws7.cell(r, 3, safe_get(f, "status", "name"))
        ws7.cell(r, 4, fmt_dt(f.get("created")))
        ws7.cell(r, 5, ", ".join(labels))
        ws7.cell(r, 6, "; ".join(reason))
        r += 1
auto_width(ws7)

# ── Sheet 8: Metric Dictionary ──
ws8 = wb2.create_sheet("Metric Dictionary")
ws8.sheet_properties.tabColor = "1F3864"
r = 1
ws8.cell(r, 1, "METRIC DICTIONARY").font = section_font; r += 2
dict_headers = ["Metric", "Category", "Formula", "Start", "Stop", "Pauses", "Stat", "Available?"]
for c, h in enumerate(dict_headers, 1): ws8.cell(r, c, h)
style_header_row(ws8, r, len(dict_headers)); r += 1
metrics = [
    ("New Tickets","Demand","Count created in period","N/A","N/A","N/A","Count","YES"),
    ("Ticket Mix","Demand","% by type","N/A","N/A","N/A","Rate","YES"),
    ("Open Backlog","Backlog","Count not Resolved/Closed","N/A","N/A","N/A","Count","YES"),
    ("Backlog Aging","Backlog","Open by age bucket","Created","Now","N/A","Count","YES"),
    ("Stale Tickets","Backlog","Open, no update 7+ days","Updated","Now","N/A","Count","YES"),
    ("Throughput","Backlog","Resolved in period","N/A","N/A","N/A","Count","YES"),
    ("Net Flow","Backlog","New - Resolved","N/A","N/A","N/A","Count","YES"),
    ("FRT","Responsiveness","Created→first response","Created","Response","N/A","Med+P90","NO—changelog"),
    ("Calendar TTR","Resolution","Created→Resolved","Created","Resolved","None","Med+P90/95","YES"),
    ("Active TTR","Resolution","Active states only","Created","Resolved","Pending*","Med+P90/95","NO—changelog"),
    ("FCR","Quality","L1 resolved, no reopen 7d","N/A","N/A","N/A","Rate","NO—changelog"),
    ("Reopen Rate","Quality","Reopened/Resolved","N/A","N/A","N/A","Rate","NO—changelog"),
    ("CSAT","Customer","Score + % satisfied","N/A","N/A","N/A","Avg+Rate","NO—not in data"),
    ("Escalation Rate","Escalation","Escalated/Resolved","N/A","N/A","N/A","Rate","NO—changelog"),
    ("Resolved/Agent","Productivity","Agent count/period","N/A","N/A","N/A","Count","YES"),
]
for row_data in metrics:
    for c, val in enumerate(row_data, 1):
        cell = ws8.cell(r, c, val)
        if "NO" in str(val): cell.fill = warn_fill
        elif val == "YES": cell.fill = good_fill
    r += 1
auto_width(ws8, max_w=50)

path2 = os.path.join(REPORT_DIR, "OIT_Baseline_Metrics_2026-03-03_0825.xlsx")
wb2.save(path2)
print(f"Baseline metrics saved: {path2}")

print(f"\n{'='*50}")
print(f"SUMMARY")
print(f"  Total in Jira:        {len(all_issues_raw):,}")
print(f"  Excluded (oasisdev):  {excluded_count:,}")
print(f"  Included for metrics: {total:,}")
print(f"  Open backlog:         {len(open_tickets):,}")
print(f"  Resolved:             {len(resolved_tickets):,}")
print(f"  Calendar TTR median:  {fmt_hours(ttr_stats.get('median'))}")
print(f"  Calendar TTR P90:     {fmt_hours(ttr_stats.get('p90'))}")
